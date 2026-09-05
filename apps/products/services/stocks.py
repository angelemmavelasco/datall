from dataclasses import dataclass
from decimal import Decimal
from typing import ClassVar
import traceback
import calendar
import datetime
import io
from django.db.models import Q, Sum
from django.db.models.functions import Coalesce
from apps.products.models import ProductClass
from apps.sales.models import SaleTransaction

try:
    import pandas as pd
except ImportError:
    pd = None

from django.db import transaction, IntegrityError
from django.db.models import QuerySet
from django.utils import timezone
from apps.core.services.uploads import BaseETLHelper

from apps.core.services.users import UsersService
from apps.products.models import Product, Stock
from apps.sales.models import Warehouse


class ServiceError(Exception):
    pass


class PermissionsError(ServiceError):
    pass


class StockNotFound(ServiceError):
    pass


@dataclass
class StocksService(UsersService):
    stock_model: type = Stock
    product_model: type = Product
    warehouse_model: type = Warehouse

    ACCESS_CONTEXTS: ClassVar[tuple[str, ...]] = (
        'acceso_total_inventario',
        'inventario',
        'acceso_total_productos',
        'acceso_total',
    )

    def read_stocks(self) -> QuerySet:
        """
        returns the queryset of stocks with product and warehouse select_related.
        """
        return self.stock_model.objects.select_related(
            'product',
            'product__product_class',
            'product__product_class__product_category',
            'warehouse'
        ).order_by('product__name', 'warehouse__name', 'lot_number')

    def read_stock(self, *, pk: int) -> Stock:
        """
        returns a single stock record by primary key.
        """
        stock = self.read_stocks().filter(pk=pk).first()
        if stock:
            return stock

        if self.stock_model.objects.filter(pk=pk).exists():
            raise PermissionsError(f'No tienes permiso para acceder a la existencia con ID "{pk}".')

        raise StockNotFound(f'No se encontró ninguna existencia con el ID "{pk}".')

    def _clean_stocks(self, file_obj) -> tuple[bool, str | object]:
        """
        cleans and normalizes tabular data for Stock imports.
        supports Reference mappings and default fallback column names:
        - cve_prod / cve_producto / producto -> product_id
        - lote -> lot_number
        - lugar / almacen / cedis -> warehouse_id
        - existencia / cantidad -> quantity
        - fech_venc / fecha_caducidad -> expiration_date
        """
        print(f"\n[STOCKS-ETL] initializing etl stock process", flush=True)

        if pd is None:
            err = "La librería 'pandas' no está instalada en el entorno."
            print(f"[STOCKS-ETL ERROR] {err}", flush=True)
            return False, err

        is_valid, df_or_err = BaseETLHelper.read_file_to_dataframe(file_obj)
        if not is_valid:
            print(f"[STOCKS-ETL ERROR] failed to load dataframe from file", flush=True)
            return False, df_or_err

        df = df_or_err
        print(f"[STOCKS-ETL] original columns read from file: {list(df.columns)}", flush=True)

        df = BaseETLHelper.apply_reference_column_mappings(
            df,
            self.stock_model,
            submodule_url_name='core:upload_options_list_view',
            context='columna'
        )

        fallback_map = {
            'cve_prod': 'product_id',
            'cve_producto': 'product_id',
            'cve_art': 'product_id',
            'producto': 'product_id',
            'product': 'product_id',
            'lote': 'lot_number',
            'num_lote': 'lot_number',
            'lugar': 'warehouse_id',
            'almacen': 'warehouse_id',
            'cedis': 'warehouse_id',
            'warehouse': 'warehouse_id',
            'existencia': 'quantity',
            'existencias': 'quantity',
            'cantidad': 'quantity',
            'stock': 'quantity',
            'fech_venc': 'expiration_date',
            'f_venc': 'expiration_date',
            'fecha_venc': 'expiration_date',
            'fecha_vencimiento': 'expiration_date',
            'fecha_caducidad': 'expiration_date',
            'caducidad': 'expiration_date',
            'vencimiento': 'expiration_date',
        }

        rename_dict = {}
        for col in df.columns:
            clean_col = str(col).strip().lower()
            if clean_col in fallback_map and fallback_map[clean_col] not in df.columns:
                rename_dict[col] = fallback_map[clean_col]

        if rename_dict:
            df.rename(columns=rename_dict, inplace=True)
            print(f"[STOCKS-ETL] fallback renamed columns: {rename_dict}", flush=True)

        df = BaseETLHelper.resolve_foreign_key_columns(df, self.stock_model)
        print(f"[STOCKS-ETL] columns after fallback map and resolve: {list(df.columns)}", flush=True)

        is_req_valid, req_msg = BaseETLHelper.validate_required_columns(
            df,
            {
                'product_id': 'Producto (cve_prod / product_id)',
                'warehouse_id': 'Centro de Distribución / Almacén (lugar / warehouse_id)',
                'quantity': 'Existencia / Cantidad (existencia / quantity)',
            }
        )
        if not is_req_valid:
            print(f"[STOCKS-ETL ERROR] required columns validation failed: {req_msg}", flush=True)
            return False, req_msg

        if 'warehouse_id' in df.columns:
            df = BaseETLHelper.apply_reference_value_mappings(
                df,
                column='warehouse_id',
                target_model=self.warehouse_model,
                context='valor_cedis',
                submodule_url_name='core:upload_options_list_view'
            )

        str_cols = ['product_id', 'warehouse_id', 'lot_number']
        for col in str_cols:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace({'nan': None, 'NaN': None, 'None': None, 'none': None, 'null': None, 'NULL': None, '': None})

        if 'lot_number' in df.columns:
            df['lot_number'] = df['lot_number'].fillna('')
        else:
            df['lot_number'] = ''
        df = df.dropna(subset=['product_id', 'warehouse_id'])
        if df.empty:
            err = "The file does not contain valid stock after discarding rows without product or warehouse."
            print(f"[STOCKS-ETL ERROR] {err}", flush=True)
            return False, err

        if 'quantity' in df.columns:
            df['quantity'] = pd.to_numeric(
                df['quantity'].astype(str).str.replace(r'[$, ]', '', regex=True),
                errors='coerce'
            ).fillna(0.0).round(2)
        else:
            df['quantity'] = 0.0

        if 'expiration_date' in df.columns:
            fechas_str = df['expiration_date'].astype(str).str.strip().str.replace(' 00:00:00', '', regex=False)
            fechas_str = fechas_str.replace({'nan': None, 'NaN': None, 'None': None, 'none': None, 'null': None, '': None})

            dates_iso = pd.to_datetime(fechas_str, format='%Y-%m-%d', errors='coerce')
            dates_mx = pd.to_datetime(fechas_str, format='%d/%m/%Y', errors='coerce')
            dates_mx_short = pd.to_datetime(fechas_str, format='%d/%m/%y', errors='coerce')

            parsed_dates = dates_iso.fillna(dates_mx).fillna(dates_mx_short).fillna(pd.to_datetime(fechas_str, errors='coerce'))
            df['expiration_date'] = parsed_dates.dt.date
            df['expiration_date'] = df['expiration_date'].where(df['expiration_date'].notnull(), None)
        else:
            df['expiration_date'] = None

        is_fk_valid, fk_msg = BaseETLHelper.validate_foreign_keys(df, self.stock_model)
        if not is_fk_valid:
            print(f"[STOCKS-ETL ERROR] foreign key validation failed: {fk_msg}", flush=True)
            return False, fk_msg
        initial_count = len(df)
        df = df.groupby(['product_id', 'warehouse_id', 'lot_number'], as_index=False).agg({
            'quantity': 'sum',
            'expiration_date': 'first',
        })
        if len(df) < initial_count:
            print(f"[STOCKS-ETL] consolidated {initial_count - len(df)} duplicate records grouping by Product, Warehouse and Lot.", flush=True)

        df = df.where(pd.notnull(df), None)
        print(f"[STOCKS-ETL SUCCESS] Cleaning completed. {len(df)} stock records ready to process.\n", flush=True)

        return True, df

    def bulk_create_stocks(self, file_obj) -> object:
        """
        Executes bulk import/upsert of Stock records into the database.
        Respects the unique constraint on (product, warehouse, lot_number).
        """
        from apps.core.services.uploads import ImportResult, PermissionsError, BaseETLHelper

        print(f"\n[STOCKS-BULK] initializing bulk_create_stocks", flush=True)

        if not self.has_full_access:
            err_msg = 'No tienes permisos suficientes para realizar cargas masivas de existencias.'
            print(f"[STOCKS-BULK PERMISSIONS ERROR] {err_msg}", flush=True)
            raise PermissionsError(err_msg)

        is_valid, df_or_err = self._clean_stocks(file_obj)
        if not is_valid:
            print(f"[STOCKS-BULK VALIDATION ERROR] {df_or_err}", flush=True)
            return ImportResult(success=False, message=df_or_err)

        df = df_or_err

        if df is None or df.empty:
            return ImportResult(success=False, message="El archivo no contiene existencias para procesar.")

        product_ids = set(df['product_id'].dropna().unique())
        warehouse_ids = set(df['warehouse_id'].dropna().unique())

        print(f"[STOCKS-BULK] Productos únicos: {len(product_ids)} | Almacenes únicos: {len(warehouse_ids)}", flush=True)

        # Lookup existing stocks for the matching products and warehouses
        existing_stocks_qs = self.stock_model.objects.filter(
            product_id__in=product_ids,
            warehouse_id__in=warehouse_ids
        )
        existing_stocks_map = {
            (str(s.product_id).strip(), str(s.warehouse_id).strip(), (s.lot_number or '').strip()): s
            for s in existing_stocks_qs
        }
        print(f"[STOCKS-BULK] Existencias coincidentes en BD: {len(existing_stocks_map)}", flush=True)

        stocks_to_create = []
        stocks_to_update = []
        total_processed = 0

        for _, row in df.iterrows():
            pid = str(row['product_id']).strip()
            wid = str(row['warehouse_id']).strip()
            lot = str(row.get('lot_number', '') or '').strip()
            qty = row['quantity']
            exp_date = row.get('expiration_date')
            if pd.isna(exp_date) or str(exp_date).strip() in ('', 'None', 'nan', 'NaN', 'NaT'):
                exp_date = None

            key = (pid, wid, lot)
            total_processed += 1

            if key in existing_stocks_map:
                stock_instance = existing_stocks_map[key]
                stock_instance.quantity = qty
                stock_instance.expiration_date = exp_date
                stock_instance.updated_at = timezone.now()
                stocks_to_update.append(stock_instance)
            else:
                stocks_to_create.append(
                    self.stock_model(
                        product_id=pid,
                        warehouse_id=wid,
                        lot_number=lot,
                        quantity=qty,
                        expiration_date=exp_date
                    )
                )

        print(
            f"[STOCKS-BULK PLAN] Nuevos a crear: {len(stocks_to_create)} | "
            f"A actualizar: {len(stocks_to_update)}",
            flush=True
        )

        try:
            with transaction.atomic():
                if stocks_to_create:
                    print(f"[STOCKS-BULK DB] Creando {len(stocks_to_create)} existencias...", flush=True)
                    self.stock_model.objects.bulk_create(stocks_to_create, batch_size=1000)

                if stocks_to_update:
                    print(f"[STOCKS-BULK DB] Actualizando {len(stocks_to_update)} existencias...", flush=True)
                    self.stock_model.objects.bulk_update(
                        stocks_to_update,
                        ['quantity', 'expiration_date', 'updated_at'],
                        batch_size=1000
                    )

            success_msg = f"Importación exitosa. Se crearon {len(stocks_to_create)} existencias y se actualizaron {len(stocks_to_update)}."
            print(f"[STOCKS-BULK SUCCESS] {success_msg}\n", flush=True)
            return ImportResult(
                success=True,
                message=success_msg,
                total_processed=total_processed,
                created_count=len(stocks_to_create),
                updated_count=len(stocks_to_update)
            )

        except Exception as e:
            print(f"\n[STOCKS-BULK DATABASE EXCEPTION] Error al ejecutar transacción en base de datos: {str(e)}", flush=True)
            traceback.print_exc()
            humanized_msg = BaseETLHelper.humanize_database_error(e)
            print(f"[STOCKS-BULK HUMANIZED MESSAGE] {humanized_msg}\n", flush=True)
            return ImportResult(
                success=False,
                message=humanized_msg,
                total_processed=total_processed,
                errors=[str(e)]
            )

@dataclass
class StockTransfersService(UsersService):
    stock_model: type = Stock
    product_model: type = Product
    warehouse_model: type = Warehouse
    product_class_model: type = ProductClass
    sale_transaction_model: type = SaleTransaction

    ACCESS_CONTEXTS: ClassVar[tuple[str, ...]] = (
        'acceso_total_inventario',
        'inventario',
        'acceso_total_productos',
        'acceso_total_ventas',
        'acceso_total',
    )

    def __post_init__(self):
        super().__post_init__()
        self.errors = []
        self.warnings = []

    def _parse_month(self, ym_str: str) -> datetime.date | None:
        if not ym_str:
            return None
        try:
            return datetime.datetime.strptime(ym_str.strip(), '%Y-%m').date()
        except (ValueError, TypeError):
            return None

    def _get_end_of_month(self, d: datetime.date) -> datetime.date:
        last_day = calendar.monthrange(d.year, d.month)[1]
        return datetime.date(d.year, d.month, last_day)

    def _months_diff(self, start_date: datetime.date, end_date: datetime.date) -> int:
        return (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month) + 1

    def get_available_warehouses(self) -> QuerySet:
        return self.warehouse_model.objects.all().order_by('name')

    def get_available_product_classes(self) -> QuerySet:
        return self.product_class_model.objects.all().order_by('name')

    def _find_transit_warehouse(self, destination_warehouse_id: str) -> Warehouse | None:
        if not destination_warehouse_id:
            return None
        dest_clean = destination_warehouse_id.strip().lower()
        return self.warehouse_model.objects.filter(
            Q(id=f"t_{dest_clean}") |
            (Q(warehouse_type=self.warehouse_model.WarehouseTypeChoices.TRANSFER) &
             (Q(id__icontains=dest_clean) | Q(name__icontains=dest_clean)))
        ).first()

    def calculate_transfer(
        self,
        origin_warehouse_id: str,
        destination_warehouse_id: str,
        start_date_str: str,
        end_date_str: str,
        product_class_ids: list[str] | None = None,
        rotation_level_ids: list[str] | None = None
    ) -> list[dict] | None:
        """
        calculates stock transfers between two warehouses
        """
        self.errors = []
        self.warnings = []

        if not origin_warehouse_id or not destination_warehouse_id:
            self.errors.append("Falta seleccionar almacén de origen o destino.")
            return None

        if origin_warehouse_id == destination_warehouse_id:
            self.errors.append("El almacén de origen y destino no pueden ser el mismo.")
            return None

        start_date = self._parse_month(start_date_str)
        end_date = self._parse_month(end_date_str)

        if not start_date or not end_date:
            self.errors.append("Las fechas de evaluación no tienen un formato válido (YYYY-MM).")
            return None

        if start_date > end_date:
            self.errors.append("La fecha de inicio debe ser anterior o igual a la fecha de fin.")
            return None

        end_date = self._get_end_of_month(end_date)
        months_count = self._months_diff(start_date, end_date)
        if months_count <= 0:
            months_count = 1

        products_qs = self.product_model.objects.select_related('product_class')
        clean_class_ids = [pid for pid in (product_class_ids or []) if pid]
        if clean_class_ids:
            products_qs = products_qs.filter(product_class_id__in=clean_class_ids)

        products = list(products_qs.order_by('name'))
        product_ids = [p.id for p in products]

        if not product_ids:
            return []

        sales_qs = self.sale_transaction_model.objects.filter(
            warehouse_id=destination_warehouse_id,
            product_id__in=product_ids,
            sale_date__gte=start_date,
            sale_date__lte=end_date
        ).values('product_id').annotate(
            total_qty=Coalesce(Sum('quantity'), Decimal('0.00'))
        )
        sales_map = {item['product_id']: item['total_qty'] for item in sales_qs}

        dest_stock_qs = self.stock_model.objects.filter(
            warehouse_id=destination_warehouse_id,
            product_id__in=product_ids
        ).values('product_id').annotate(
            total_qty=Coalesce(Sum('quantity'), Decimal('0.00'))
        )
        dest_stock_map = {item['product_id']: item['total_qty'] for item in dest_stock_qs}

        origin_stock_qs = self.stock_model.objects.filter(
            warehouse_id=origin_warehouse_id,
            product_id__in=product_ids
        ).values('product_id').annotate(
            total_qty=Coalesce(Sum('quantity'), Decimal('0.00'))
        )
        origin_stock_map = {item['product_id']: item['total_qty'] for item in origin_stock_qs}

        transit_warehouse = self._find_transit_warehouse(destination_warehouse_id)
        transit_stock_map = {}
        if transit_warehouse:
            transit_qs = self.stock_model.objects.filter(
                warehouse=transit_warehouse,
                product_id__in=product_ids
            ).values('product_id').annotate(
                total_qty=Coalesce(Sum('quantity'), Decimal('0.00'))
            )
            transit_stock_map = {item['product_id']: item['total_qty'] for item in transit_qs}

        grouped_results = {}
        valid_rotations = [str(r).strip() for r in (rotation_level_ids or []) if str(r).strip()]

        for p in products:
            sold_qty = sales_map.get(p.id, Decimal('0.00'))
            current_stock = dest_stock_map.get(p.id, Decimal('0.00'))
            origin_stock = origin_stock_map.get(p.id, Decimal('0.00'))
            in_transit = transit_stock_map.get(p.id, Decimal('0.00'))

            if sold_qty == 0 and current_stock == 0 and in_transit == 0:
                continue

            avg_monthly = sold_qty / Decimal(months_count)

            if sold_qty > 50:
                rotation_level = '1'
                rotation_name = 'Alta'
            elif sold_qty >= 10:
                rotation_level = '2'
                rotation_name = 'Media'
            else:
                rotation_level = '3'
                rotation_name = 'Baja'

            if valid_rotations and rotation_level not in valid_rotations:
                continue

            class_id = p.product_class_id if p.product_class_id else 'sin_clase'
            class_name = p.product_class.name.title() if p.product_class and p.product_class.name else 'Sin Clase'

            if class_id not in grouped_results:
                grouped_results[class_id] = {
                    'class_id': class_id,
                    'class_name': class_name,
                    'products': []
                }

            if avg_monthly > 0:
                initial_coverage = (current_stock / avg_monthly) * Decimal('100.0')
            elif current_stock > 0:
                initial_coverage = Decimal('100.0')
            else:
                initial_coverage = Decimal('0.00')

            target_stock = avg_monthly * Decimal('1.0')
            initial_suggestion = max(target_stock - current_stock - in_transit, Decimal('0.00'))

            is_origin_insufficient = initial_suggestion > origin_stock

            grouped_results[class_id]['products'].append({
                'product_id': p.id,
                'product_name': (p.name or "").title(),
                'sold_qty': round(sold_qty, 2),
                'avg_monthly': round(avg_monthly, 2),
                'current_stock': round(current_stock, 2),
                'origin_stock': round(origin_stock, 2),
                'in_transit': round(in_transit, 2),
                'initial_coverage': round(initial_coverage, 2),
                'initial_suggestion': round(initial_suggestion, 0),
                'is_origin_insufficient': is_origin_insufficient,
                'rotation_name': rotation_name,
                'rotation_level': rotation_level,
            })

        final_results = []
        for class_id, data in grouped_results.items():
            data['products'].sort(key=lambda x: x['sold_qty'], reverse=True)
            final_results.append(data)

        final_results.sort(key=lambda x: x['class_name'])
        return final_results


class StockTransferExports:

    @staticmethod
    def export_excel(
        results: list[dict],
        start_date_str: str,
        end_date_str: str,
        origin_name: str,
        destination_name: str,
        coverages: dict | None = None
    ) -> bytes | None:
        if not results:
            return None

        if coverages is None:
            coverages = {}

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transferencias"

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        subheader_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        class_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        white_bold = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
        title_font = Font(name="Calibri", color="0F172A", bold=True, size=13)
        bold_font = Font(name="Calibri", color="0F172A", bold=True, size=10)
        regular_font = Font(name="Calibri", color="0F172A", size=10)
        muted_font = Font(name="Calibri", color="64748B", size=9)

        thin_side = Side(style='thin', color='CBD5E1')
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        period_str = f"{start_date_str} a {end_date_str}"

        headers = [
            "ID Producto",
            "Producto",
            "Unidades Vendidas",
            "Promedio Mensual",
            "Existencias Destino",
            "Existencias Origen",
            "En Tránsito",
            "Cobertura Actual (%)",
            "Cobertura Solicitada (%)",
            "Transferencia Sugerida"
        ]

        ws.append([f"Reporte de Transferencias de Stock (Reposición)"])
        ws.cell(row=ws.max_row, column=1).font = title_font
        ws.append([f"Periodo Evaluado: {period_str} | Fecha de Cálculo: {current_time}"])
        ws.cell(row=ws.max_row, column=1).font = muted_font
        ws.append([f"CEDIS Origen: {origin_name}  ──>  CEDIS Destino: {destination_name}"])
        ws.cell(row=ws.max_row, column=1).font = bold_font
        ws.append([])

        for group in results:
            ws.append(["Clase de Producto:", group['class_name']])
            ws.cell(row=ws.max_row, column=1).font = bold_font
            ws.cell(row=ws.max_row, column=1).fill = class_fill
            ws.cell(row=ws.max_row, column=2).font = bold_font
            ws.cell(row=ws.max_row, column=2).fill = class_fill

            rotations = {'Alta': [], 'Media': [], 'Baja': []}
            for p in group['products']:
                if p['rotation_name'] in rotations:
                    rotations[p['rotation_name']].append(p)

            for rot_name, prods in rotations.items():
                if not prods:
                    continue

                ws.append([])
                ws.append([f"Rotación {rot_name} ({len(prods)} productos)"])
                rot_row = ws.max_row
                ws.cell(row=rot_row, column=1).font = white_bold
                ws.cell(row=rot_row, column=1).fill = subheader_fill

                ws.append(headers)
                h_row = ws.max_row
                for col_num, h_text in enumerate(headers, 1):
                    cell = ws.cell(row=h_row, column=col_num)
                    cell.font = white_bold
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for p in prods:
                    prod_id_str = str(p['product_id'])
                    cov = coverages.get(prod_id_str, 1.0)

                    ws.append([
                        p['product_id'],
                        p['product_name'],
                        p['sold_qty'],
                        p['avg_monthly'],
                        p['current_stock'],
                        p['origin_stock'],
                        p['in_transit'],
                        0,
                        cov,
                        0
                    ])
                    d_row = ws.max_row

                    # Col D: Promedio mensual
                    # Col E: Existencias Destino
                    # Col F: Existencias Origen
                    # Col G: En Tránsito
                    # Col H: Cobertura actual = IF(D>0, (E/D)*100, IF(E>0, 100, 0))
                    # Col I: Cobertura solicitada
                    # Col J: Transferencia = MAX((D*I)-E-G, 0)
                    formula_cov = f"=IF(D{d_row}>0, (E{d_row}/D{d_row})*100, IF(E{d_row}>0, 100, 0))"
                    formula_transfer = f"=MAX((D{d_row}*I{d_row})-E{d_row}-G{d_row}, 0)"

                    ws.cell(row=d_row, column=8, value=formula_cov)
                    ws.cell(row=d_row, column=10, value=formula_transfer)

                    for c_num in range(1, len(headers) + 1):
                        c = ws.cell(row=d_row, column=c_num)
                        c.font = regular_font
                        c.border = thin_border
                        if c_num in (1, 2):
                            c.alignment = Alignment(horizontal="left", vertical="center")
                        elif c_num in (3, 4, 5, 6, 7):
                            c.number_format = '#,##0.00'
                            c.alignment = Alignment(horizontal="right", vertical="center")
                        elif c_num == 8:
                            c.number_format = '0.00"%"'
                            c.alignment = Alignment(horizontal="right", vertical="center")
                        elif c_num == 9:
                            c.number_format = '0.0'
                            c.alignment = Alignment(horizontal="right", vertical="center")
                        elif c_num == 10:
                            c.number_format = '#,##0'
                            c.font = bold_font
                            c.alignment = Alignment(horizontal="right", vertical="center")

            ws.append([])

        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 42
        for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws.column_dimensions[col_letter].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
