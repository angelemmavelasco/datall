import io
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any
from collections import defaultdict
from django.db.models import QuerySet, Sum, Count, Q
from django.utils import timezone
from django.db.models.functions import TruncMonth

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.products.models import ProductClass


@dataclass
class TargetAchievementService:
    user: Any
    targets_qs: QuerySet
    transactions_qs: QuerySet
    customers_qs: QuerySet
    routes_qs: QuerySet
    date_start: date | str | None = None
    date_end: date | str | None = None
    cleaned_data: dict[str, Any] | None = None

    today: date = field(init=False)
    date_start_dt: date = field(init=False)
    date_end_dt: date = field(init=False)
    total_b_days: int = field(init=False)
    elapsed_b_days: int = field(init=False)

    VALID_CORE_CLASSES: tuple[str, ...] = (
        'diamond',
        'diamond naturals',
        'care',
        'taste of the wild',
        'msd',
        'vetoquinol',
        'zoetis'
    )

    def __post_init__(self):
        self._init_dates()
        self.total_b_days, self.elapsed_b_days = self._get_business_days(self.date_start_dt, self.date_end_dt)

    def _init_dates(self) -> None:
        """
        normalize the dates from request, if not provided by default will be the current month
        """
        self.today = timezone.localdate()
        first_day_curr_month = self.today.replace(day=1)
        if self.today.month == 12:
            last_day_curr_month = date(self.today.year, 12, 31)
        else:
            last_day_curr_month = date(self.today.year, self.today.month + 1, 1) - timedelta(days=1)

        def _parse_val(val: Any) -> date | None:
            if not val:
                return None
            if isinstance(val, date) and not isinstance(val, datetime):
                return val
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, str):
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
                    try:
                        return datetime.strptime(val.strip(), fmt).date()
                    except ValueError:
                        continue
            return None

        parsed_start = _parse_val(self.date_start)
        parsed_end = _parse_val(self.date_end)

        self.date_start_dt = parsed_start or first_day_curr_month
        self.date_end_dt = parsed_end or last_day_curr_month
        self.date_start = self.date_start_dt
        self.date_end = self.date_end_dt

    def _get_business_days(self, date_start_dt: date, date_end_dt: date) -> tuple[int, int]:
        """
        calculate total business days (monday to friday) of the range and business days elapsed until yesterday/today.
        """
        if not date_start_dt or not date_end_dt or date_start_dt > date_end_dt:
            return 1, 1

        total_days = 0
        curr = date_start_dt
        while curr <= date_end_dt:
            if curr.weekday() < 5:  # Monday to Friday
                total_days += 1
            curr += timedelta(days=1)

        yesterday = self.today - timedelta(days=1)
        end_elapsed = min(yesterday, date_end_dt)

        elapsed_days = 0
        curr = date_start_dt
        while curr <= end_elapsed:
            if curr.weekday() < 5:
                elapsed_days += 1
            curr += timedelta(days=1)

        total_safe = max(1, total_days)
        elapsed_safe = max(1, elapsed_days)
        return total_safe, elapsed_safe

    def _get_display_product_classes(self) -> list[str]:
        """
        normalize product class names to show in columns
        """
        classes_qs = ProductClass.objects.all()
        if self.cleaned_data and self.cleaned_data.get('product_category'):
            classes_qs = classes_qs.filter(product_category__in=self.cleaned_data['product_category'])
        if self.cleaned_data and self.cleaned_data.get('product_class'):
            classes_qs = classes_qs.filter(pk__in=[c.pk if hasattr(c, 'pk') else c for c in self.cleaned_data['product_class']])

        db_classes = list(classes_qs.values_list('name', flat=True).order_by('name'))
        normalized_db = [c.strip().lower() for c in db_classes if c and c.strip()]

        if self.cleaned_data and (self.cleaned_data.get('product_class') or self.cleaned_data.get('product_category')):
            ordered_list = []
            for c in normalized_db:
                if c not in ordered_list:
                    ordered_list.append(c)
            return ordered_list or ['otros']

        ordered_list = []
        for vc in self.VALID_CORE_CLASSES:
            ordered_list.append(vc)
            
        for c in ('country value', 'nutriforce'):
            if c not in ordered_list:
                ordered_list.append(c)

        for c in normalized_db:
            if c not in ordered_list and c != 'otros':
                ordered_list.append(c)

        if 'otros' not in ordered_list:
            ordered_list.append('otros')

        return ordered_list

    def _get_customer_metrics_by_route(self) -> dict[str, dict[str, Any]]:
        """
        calculate registered customers, new customers in the period and active customers with purchases by route.
        """
        today = self.today
        base_assignments = self.customers_qs.filter(
            Q(assignments__end_date__isnull=True) | Q(assignments__end_date__gte=today)
        )

        customer_qs = base_assignments
        if self.date_end_dt:
            customer_qs = customer_qs.filter(registration_date__lte=self.date_end_dt)

        customers_per_route = customer_qs.order_by().values('assignments__route_id').annotate(
            registered=Count('id', distinct=True)
        )

        new_customers_qs = base_assignments.filter(
            registration_date__gte=self.date_start_dt,
            registration_date__lte=self.date_end_dt
        )
        new_customers_per_route = new_customers_qs.order_by().values('assignments__route_id').annotate(
            nuevos=Count('id', distinct=True)
        )

        active_tx = self.transactions_qs.filter(
            sale_date__gte=self.date_start_dt,
            sale_date__lte=self.date_end_dt
        )
        active_customers_qs = active_tx.order_by().values('route_id').annotate(
            active=Count('customer_id', distinct=True)
        )

        route_customers_dict = defaultdict(lambda: {'registered': 0, 'new': 0, 'active': 0, 'portfolio_scope': 0.0})
        for row in customers_per_route:
            rid = str(row['assignments__route_id']) if row['assignments__route_id'] is not None else 'sn'
            route_customers_dict[rid]['registered'] = row['registered'] or 0

        for row in new_customers_per_route:
            rid = str(row['assignments__route_id']) if row['assignments__route_id'] is not None else 'sn'
            route_customers_dict[rid]['new'] = row['nuevos'] or 0

        for row in active_customers_qs:
            rid = str(row['route_id']) if row['route_id'] is not None else 'sn'
            route_customers_dict[rid]['active'] = row['active'] or 0

        for rid, cdict in route_customers_dict.items():
            reg = cdict['registered']
            act = cdict['active']
            cdict['portfolio_scope'] = round((act / reg * 100.0), 2) if reg > 0 else 0.00

        return route_customers_dict

    def _get_product_class_prf_by_route(self) -> tuple[dict[tuple[str, str], dict[str, Any]], dict[str, str], list[str]]:
        """
        calculate quotas, sales, difference, scope and projection by route and by product class.
        the target's assigned business unit dictates where the route appears.
        if a route has no targets in the period, it appears in its base business unit only if is_active is true.
        """
        display_classes = self._get_display_product_classes()

        target_qs = self.targets_qs.filter(
            period__gte=self.date_start_dt,
            period__lte=self.date_end_dt
        )

        tx_qs = self.transactions_qs.filter(
            sale_date__gte=self.date_start_dt,
            sale_date__lte=self.date_end_dt
        )

        route_targets = target_qs.order_by().values(
            'route_id',
            'route__name',
            'route__is_active',
            'route__business_unit_id',
            'route__business_unit__name',
            'business_unit_id',
            'business_unit__name',
            'product_class__name'
        ).annotate(
            total_target=Sum('target_amount')
        )

        route_sales = tx_qs.order_by().values(
            'route_id',
            'product_class__name'
        ).annotate(
            total_sale=Sum('net_amount')
        )

        def _empty_class_dict():
            return {
                'target': Decimal('0.00'),
                'net_amount': Decimal('0.00'),
                'difference': Decimal('0.00'),
                'scope': Decimal('0.00'),
                'scope_forecast': Decimal('0.00')
            }

        routes_raw_dict: dict[tuple[str, str], dict[str, Any]] = {}
        business_unit_names: dict[str, str] = {}
        routes_with_targets: set[str] = set()
        route_assigned_bus: dict[str, set[str]] = defaultdict(set)

        #process targets grouped by route and target business_unit
        for row in route_targets:
            rid = str(row['route_id']) if row['route_id'] is not None else 'sn'
            target_amount = row['total_target'] or Decimal('0.00')

            buid = str(row['business_unit_id'] or row['route__business_unit_id'] or 'general')
            buname = row['business_unit__name'] or row['route__business_unit__name'] or 'General'
            business_unit_names[buid] = buname

            cname = (row['product_class__name'] or 'otros').strip().lower()
            if cname not in display_classes:
                if 'otros' in display_classes:
                    cname = 'otros'
                else:
                    continue

            key = (rid, buid)
            if key not in routes_raw_dict:
                routes_raw_dict[key] = {
                    'route_id': rid,
                    'route_name': row['route__name'] or rid,
                    'business_unit_id': buid,
                    'classes': defaultdict(_empty_class_dict)
                }

            routes_raw_dict[key]['classes'][cname]['target'] += target_amount
            if target_amount > 0:
                routes_with_targets.add(rid)
                route_assigned_bus[rid].add(buid)

        #process allowed routes without targets (fallback to base business_unit only if is_active is True)
        for r in self.routes_qs:
            rid = str(r.id)
            if rid not in routes_with_targets:
                if getattr(r, 'is_active', True):
                    base_buid = str(r.business_unit_id) if r.business_unit_id is not None else 'general'
                    base_buname = r.business_unit.name if r.business_unit else 'General'
                    business_unit_names[base_buid] = base_buname

                    key = (rid, base_buid)
                    if key not in routes_raw_dict:
                        routes_raw_dict[key] = {
                            'route_id': rid,
                            'route_name': r.name or rid,
                            'business_unit_id': base_buid,
                            'classes': defaultdict(_empty_class_dict)
                        }
                    route_assigned_bus[rid].add(base_buid)

        #map sales transactions to the corresponding (route_id, business_unit_id) entries
        for row in route_sales:
            rid = str(row['route_id']) if row['route_id'] is not None else 'sn'
            sale_amount = row['total_sale'] or Decimal('0.00')

            cname = (row['product_class__name'] or 'otros').strip().lower()
            if cname not in display_classes:
                if 'otros' in display_classes:
                    cname = 'otros'
                else:
                    continue

            # If route has established (rid, buid) entries, apply sales to them
            assigned_bus = route_assigned_bus.get(rid, set())
            for buid in assigned_bus:
                key = (rid, buid)
                if key in routes_raw_dict:
                    routes_raw_dict[key]['classes'][cname]['net_amount'] += sale_amount

        #calculate details by route
        customer_metrics = self._get_customer_metrics_by_route()
        routes_processed: dict[tuple[str, str], dict[str, Any]] = {}

        for (rid, buid), rdata in routes_raw_dict.items():
            c_info = customer_metrics.get(rid, {'registered': 0, 'new': 0, 'active': 0, 'portfolio_scope': 0.0})
            reg = c_info['registered']
            act = c_info['active']
            new = c_info['new']
            port_scope = c_info['portfolio_scope']

            r_target_total = Decimal('0.00')
            r_net_total = Decimal('0.00')
            completed_families = 0
            classes_breakdown = []

            for cname in display_classes:
                cdict = rdata['classes'][cname]
                target = cdict['target']
                net = cdict['net_amount']
                diff = target - net
                cdict['difference'] = diff

                if target > 0:
                    scope = (net / target) * Decimal('100.00')
                else:
                    scope = Decimal('0.00')
                cdict['scope'] = scope

                forecast = (net / Decimal(self.elapsed_b_days)) * Decimal(self.total_b_days)
                if target > 0:
                    scope_forecast = (forecast / target) * Decimal('100.00')
                else:
                    scope_forecast = Decimal('0.00')
                cdict['scope_forecast'] = scope_forecast

                if cname in self.VALID_CORE_CLASSES:
                    if target > 0 and net >= target:
                        completed_families += 1

                r_target_total += target
                r_net_total += net

                classes_breakdown.append({
                    'product_class_name': cname,
                    'target': target,
                    'net_amount': net,
                    'difference': diff,
                    'scope': scope,
                    'scope_forecast': scope_forecast,
                })

            diff_total = r_target_total - r_net_total
            if r_target_total > 0:
                scope_total = (r_net_total / r_target_total) * Decimal('100.00')
            else:
                scope_total = Decimal('0.00')

            proyeccion_r = (r_net_total / Decimal(self.elapsed_b_days)) * Decimal(self.total_b_days)
            if r_target_total > 0:
                scope_forecast_total = (proyeccion_r / r_target_total) * Decimal('100.00')
            else:
                scope_forecast_total = Decimal('0.00')

            routes_processed[(rid, buid)] = {
                'route_id': rid,
                'route_name': rdata['route_name'],
                'business_unit_id': buid,
                'registered_customers': reg,
                'active_customers': act,
                'portfolio_scope': port_scope,
                'new_customers': new,
                'completed_product_classes': completed_families,
                'target': r_target_total,
                'net_amount': r_net_total,
                'difference': diff_total,
                'scope': scope_total,
                'scope_forecast': scope_forecast_total,
                'classes_dict': rdata['classes'],
                'route_product_classes_breakdown': classes_breakdown,
                'rank_sale': 0,
                'rank_scope': 0,
            }

        return routes_processed, business_unit_names, display_classes

    def _get_product_class_prf_by_business_unit(self) -> list[dict[str, Any]]:
        """
        aggregate results by business unit and rank routes by net sales and scope
        """
        routes_processed, business_unit_names, display_classes = self._get_product_class_prf_by_route()

        def _empty_class_dict():
            return {
                'target': Decimal('0.00'),
                'net_amount': Decimal('0.00'),
                'difference': Decimal('0.00'),
                'scope': Decimal('0.00'),
                'scope_forecast': Decimal('0.00')
            }

        bu_summaries = defaultdict(lambda: {
            'business_unit_id': '',
            'business_unit_name': '',
            'registered_customers': 0,
            'active_customers': 0,
            'portfolio_scope': 0.0,
            'new_customers': 0,
            'completed_product_classes': 0,
            'target': Decimal('0.00'),
            'net_amount': Decimal('0.00'),
            'difference': Decimal('0.00'),
            'scope': Decimal('0.00'),
            'scope_forecast': Decimal('0.00'),
            'classes': defaultdict(_empty_class_dict),
            'routes_data': []
        })

        for (rid, buid), rdata in routes_processed.items():
            buid = rdata['business_unit_id']
            bu_sum = bu_summaries[buid]
            bu_sum['business_unit_id'] = buid
            bu_sum['business_unit_name'] = business_unit_names.get(buid, buid).title()

            bu_sum['registered_customers'] += rdata['registered_customers']
            bu_sum['active_customers'] += rdata['active_customers']
            bu_sum['new_customers'] += rdata['new_customers']
            bu_sum['target'] += rdata['target']
            bu_sum['net_amount'] += rdata['net_amount']

            for cname in display_classes:
                cdict = rdata['classes_dict'][cname]
                wdict = bu_sum['classes'][cname]
                wdict['target'] += cdict['target']
                wdict['net_amount'] += cdict['net_amount']
                wdict['difference'] += cdict['difference']

            bu_sum['routes_data'].append(rdata)

        # Consolidate totals and rankings for each business unit
        final_bu_list = []
        for buid, bu_sum in bu_summaries.items():
            reg = bu_sum['registered_customers']
            act = bu_sum['active_customers']
            bu_sum['portfolio_scope'] = round((act / reg * 100.0), 2) if reg > 0 else 0.00

            bu_comp_fams = 0
            bu_classes_breakdown = []

            for cname in display_classes:
                wdict = bu_sum['classes'][cname]
                t = wdict['target']
                n = wdict['net_amount']
                diff = t - n
                wdict['difference'] = diff

                if t > 0:
                    wdict['scope'] = (n / t) * Decimal('100.00')
                else:
                    wdict['scope'] = Decimal('0.00')

                p = (n / Decimal(self.elapsed_b_days)) * Decimal(self.total_b_days)
                if t > 0:
                    wdict['scope_forecast'] = (p / t) * Decimal('100.00')
                else:
                    wdict['scope_forecast'] = Decimal('0.00')

                if cname in self.VALID_CORE_CLASSES:
                    if t > 0 and n >= t:
                        bu_comp_fams += 1

                bu_classes_breakdown.append({
                    'product_class_name': cname,
                    'target': t,
                    'net_amount': n,
                    'difference': diff,
                    'scope': wdict['scope'],
                    'scope_forecast': wdict['scope_forecast']
                })

            bu_sum['completed_product_classes'] = bu_comp_fams

            t_total = bu_sum['target']
            n_total = bu_sum['net_amount']
            bu_sum['difference'] = t_total - n_total
            if t_total > 0:
                bu_sum['scope'] = (n_total / t_total) * Decimal('100.00')
            else:
                bu_sum['scope'] = Decimal('0.00')

            p_total = (n_total / Decimal(self.elapsed_b_days)) * Decimal(self.total_b_days)
            if t_total > 0:
                bu_sum['scope_forecast'] = (p_total / t_total) * Decimal('100.00')
            else:
                bu_sum['scope_forecast'] = Decimal('0.00')

            # calculate internal rankings of routes by sale and by scope
            routes = bu_sum['routes_data']
            routes_by_scope = sorted(routes, key=lambda x: x['scope'], reverse=True)
            routes_by_sale = sorted(routes, key=lambda x: x['net_amount'], reverse=True)

            for rank, r in enumerate(routes_by_scope, start=1):
                r['rank_scope'] = rank
            for rank, r in enumerate(routes_by_sale, start=1):
                r['rank_sale'] = rank

            # sort routes by id
            bu_sum['routes_data'].sort(key=lambda x: int(x['route_id']) if str(x['route_id']).isdigit() else str(x['route_id']))

            final_bu_list.append({
                'total_business_unit_data': {
                    'business_unit_id': buid,
                    'business_unit_name': bu_sum['business_unit_name'],
                    'registered_customers': bu_sum['registered_customers'],
                    'active_customers': bu_sum['active_customers'],
                    'portfolio_scope': bu_sum['portfolio_scope'],
                    'new_customers': bu_sum['new_customers'],
                    'completed_product_classes': bu_sum['completed_product_classes'],
                    'target': bu_sum['target'],
                    'net_amount': bu_sum['net_amount'],
                    'difference': bu_sum['difference'],
                    'scope': bu_sum['scope'],
                    'scope_forecast': bu_sum['scope_forecast'],
                    'business_unit_product_classes_breakdown': bu_classes_breakdown,
                    'rank_sale': 0,
                    'rank_scope': 0,
                },
                'routes_data': bu_sum['routes_data'],
                'display_classes': display_classes,
            })

        # calculate rankings among business units by scope and sale
        bu_by_scope = sorted(final_bu_list, key=lambda x: x['total_business_unit_data']['scope'], reverse=True)
        bu_by_sale = sorted(final_bu_list, key=lambda x: x['total_business_unit_data']['net_amount'], reverse=True)

        for rank, item in enumerate(bu_by_scope, start=1):
            item['total_business_unit_data']['rank_scope'] = rank
        for rank, item in enumerate(bu_by_sale, start=1):
            item['total_business_unit_data']['rank_sale'] = rank

        final_bu_list.sort(key=lambda x: x['total_business_unit_data']['business_unit_name'])
        return final_bu_list

    def get_grand_total_data(self, final_bu_list: list[dict[str, Any]], display_classes: list[str]) -> dict[str, Any]:
        """
        Consolidate grand totals across all business units for the summary table.
        """
        if not final_bu_list:
            return {}

        total_registered = sum(x['total_business_unit_data']['registered_customers'] for x in final_bu_list)
        total_active = sum(x['total_business_unit_data']['active_customers'] for x in final_bu_list)
        total_new = sum(x['total_business_unit_data']['new_customers'] for x in final_bu_list)
        total_target = sum((x['total_business_unit_data']['target'] for x in final_bu_list), Decimal('0.00'))
        total_net = sum((x['total_business_unit_data']['net_amount'] for x in final_bu_list), Decimal('0.00'))
        total_diff = total_target - total_net

        port_scope = round((total_active / total_registered * 100.0), 2) if total_registered > 0 else 0.00
        scope = (total_net / total_target * Decimal('100.00')) if total_target > 0 else Decimal('0.00')

        proj = (total_net / Decimal(self.elapsed_b_days)) * Decimal(self.total_b_days)
        scope_forecast = (proj / total_target * Decimal('100.00')) if total_target > 0 else Decimal('0.00')

        classes_totals = defaultdict(lambda: {'target': Decimal('0.00'), 'net_amount': Decimal('0.00')})
        for x in final_bu_list:
            for c in x['total_business_unit_data']['business_unit_product_classes_breakdown']:
                cname = c['product_class_name']
                classes_totals[cname]['target'] += c['target']
                classes_totals[cname]['net_amount'] += c['net_amount']

        gt_completed_fams = 0
        gt_classes_breakdown = []
        for cname in display_classes:
            t = classes_totals[cname]['target']
            n = classes_totals[cname]['net_amount']
            c_diff = t - n
            c_scope = (n / t * Decimal('100.00')) if t > 0 else Decimal('0.00')
            c_proj = (n / Decimal(self.elapsed_b_days)) * Decimal(self.total_b_days)
            c_scope_forecast = (c_proj / t * Decimal('100.00')) if t > 0 else Decimal('0.00')

            if cname in self.VALID_CORE_CLASSES:
                if t > 0 and n >= t:
                    gt_completed_fams += 1

            gt_classes_breakdown.append({
                'product_class_name': cname,
                'target': t,
                'net_amount': n,
                'difference': c_diff,
                'scope': c_scope,
                'scope_forecast': c_scope_forecast
            })

        return {
            'registered_customers': total_registered,
            'active_customers': total_active,
            'portfolio_scope': port_scope,
            'new_customers': total_new,
            'completed_product_classes': gt_completed_fams,
            'target': total_target,
            'net_amount': total_net,
            'difference': total_diff,
            'scope': scope,
            'scope_forecast': scope_forecast,
            'product_classes_breakdown': gt_classes_breakdown,
        }

    def get_target_achievement_data(self) -> list[dict[str, Any]]:
        """
        Método orquestador principal que coordina el cálculo
        y entrega la estructura consolidada para la vista y templates.
        """
        return self._get_product_class_prf_by_business_unit()


@dataclass
class TargetAchievementExports:
    target_achievement_service: TargetAchievementService

    def export_target_achievement_report(self) -> io.BytesIO:
        service = self.target_achievement_service
        data = service.get_target_achievement_data()
        display_classes = service._get_display_product_classes()
        grand_total = service.get_grand_total_data(data, display_classes) if data else {}

        wb = openpyxl.Workbook()

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        section_font = Font(name="Calibri", size=12, bold=True, color="0F172A")
        title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
        subtitle_font = Font(name="Calibri", size=9, italic=True, color="64748B")
        data_font = Font(name="Calibri", size=10)
        bold_data_font = Font(name="Calibri", size=10, bold=True)

        thin_border_side = Side(style='thin', color='CBD5E1')
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        currency_format = '"$"#,##0.00'
        pct_format = '0.00%'
        int_format = '#,##0'

        ws_summary = wb.active
        ws_summary.title = "Resumen General"
        ws_summary.views.sheetView[0].showGridLines = True

        ws_summary.cell(row=1, column=1, value="REPORTE DE ALCANCE DE OBJETIVOS - RESUMEN GENERAL").font = title_font
        now_str = timezone.localtime().strftime('%Y-%m-%d %H:%M')
        d_start_str = service.date_start_dt.strftime('%Y-%m-%d') if service.date_start_dt else ''
        d_end_str = service.date_end_dt.strftime('%Y-%m-%d') if service.date_end_dt else ''
        ws_summary.cell(row=2, column=1, value=f"Generado el: {now_str} | Periodo de análisis: {d_start_str} al {d_end_str} | Días hábiles: {service.elapsed_b_days} de {service.total_b_days}").font = subtitle_font

        ws_summary.cell(row=4, column=1, value=f"Indicadores Generales ({d_start_str} al {d_end_str})").font = section_font
        kpi_headers = ["Concepto", "Monto", "% Indicador"]
        for col_num, h_text in enumerate(kpi_headers, 1):
            cell = ws_summary.cell(row=5, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_num > 1 else "left")
            cell.border = cell_border

        net_val = float(grand_total.get('net_amount') or 0)
        target_val = float(grand_total.get('target') or 0)
        diff_val = float(grand_total.get('difference') or 0)
        scope_val = float(grand_total.get('scope') or 0) / 100.0
        forecast_val = float(grand_total.get('scope_forecast') or 0) / 100.0

        general_rows = [
            ("Venta Neta", net_val, None),
            ("Objetivo / Cuota", target_val, None),
            ("Diferencia", diff_val, None),
            ("Alcance Global", None, scope_val),
            ("Pronóstico Global", None, forecast_val),
        ]

        for row_idx, (label, val, pct) in enumerate(general_rows, 6):
            c_lbl = ws_summary.cell(row=row_idx, column=1, value=label)
            c_lbl.font = data_font
            c_lbl.border = cell_border

            c_val = ws_summary.cell(row=row_idx, column=2, value=val if val is not None else "-")
            c_val.font = bold_data_font
            if val is not None:
                c_val.number_format = currency_format
            c_val.alignment = Alignment(horizontal="right" if val is not None else "center")
            c_val.border = cell_border

            c_pct = ws_summary.cell(row=row_idx, column=3, value=pct if pct is not None else "-")
            c_pct.font = bold_data_font
            if pct is not None:
                c_pct.number_format = pct_format
            c_pct.alignment = Alignment(horizontal="right" if pct is not None else "center")
            c_pct.border = cell_border

        start_row_bu = 13
        ws_summary.cell(row=start_row_bu, column=1, value="Desglose por Gerencia").font = section_font
        bu_headers = [
            "Gerencia",
            "Clientes Registrados",
            "Clientes con Consumo",
            "% Efectividad",
            "Clientes Nuevos",
            "Objetivo",
            "Venta Neta",
            "Diferencia",
            "% Alcance",
            "% Pronóstico",
            "Familias Completas",
            "Ranking Venta",
            "Ranking Alcance",
        ]

        for col_num, h_text in enumerate(bu_headers, 1):
            cell = ws_summary.cell(row=start_row_bu + 1, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = cell_border

        curr_row = start_row_bu + 2
        for bu_item in data:
            bu_dict = bu_item['total_business_unit_data']
            row_data = [
                (bu_dict['business_unit_name'], None, "left"),
                (int(bu_dict['registered_customers']), int_format, "right"),
                (int(bu_dict['active_customers']), int_format, "right"),
                (float(bu_dict['portfolio_scope']) / 100.0, pct_format, "right"),
                (int(bu_dict['new_customers']), int_format, "right"),
                (float(bu_dict['target']), currency_format, "right"),
                (float(bu_dict['net_amount']), currency_format, "right"),
                (float(bu_dict['difference']), currency_format, "right"),
                (float(bu_dict['scope']) / 100.0, pct_format, "right"),
                (float(bu_dict['scope_forecast']) / 100.0, pct_format, "right"),
                (int(bu_dict['completed_product_classes']), int_format, "right"),
                (int(bu_dict['rank_sale']), int_format, "center"),
                (int(bu_dict['rank_scope']), int_format, "center"),
            ]

            for col_idx, (val, num_fmt, align_h) in enumerate(row_data, 1):
                cell = ws_summary.cell(row=curr_row, column=col_idx, value=val)
                cell.font = data_font
                cell.border = cell_border
                cell.alignment = Alignment(horizontal=align_h)
                if num_fmt:
                    cell.number_format = num_fmt
            curr_row += 1

        if grand_total:
            total_row_data = [
                ("Total General", None, "left"),
                (int(grand_total.get('registered_customers', 0)), int_format, "right"),
                (int(grand_total.get('active_customers', 0)), int_format, "right"),
                (float(grand_total.get('portfolio_scope', 0)) / 100.0, pct_format, "right"),
                (int(grand_total.get('new_customers', 0)), int_format, "right"),
                (float(grand_total.get('target', 0)), currency_format, "right"),
                (float(grand_total.get('net_amount', 0)), currency_format, "right"),
                (float(grand_total.get('difference', 0)), currency_format, "right"),
                (float(grand_total.get('scope', 0)) / 100.0, pct_format, "right"),
                (float(grand_total.get('scope_forecast', 0)) / 100.0, pct_format, "right"),
                (int(grand_total.get('completed_product_classes', 0)), int_format, "right"),
                ("-", None, "center"),
                ("-", None, "center"),
            ]
            for col_idx, (val, num_fmt, align_h) in enumerate(total_row_data, 1):
                cell = ws_summary.cell(row=curr_row, column=col_idx, value=val)
                cell.font = bold_data_font
                cell.border = cell_border
                cell.alignment = Alignment(horizontal=align_h)
                if num_fmt:
                    cell.number_format = num_fmt

        ws_cartera = wb.create_sheet(title="Cartera")
        ws_cartera.views.sheetView[0].showGridLines = True

        cartera_headers = [
            "Ruta",
            "Agente",
            "Gerencia",
            "Clientes Registrados",
            "Clientes con Consumo",
            "% Efectividad",
            "Clientes Nuevos"
        ]

        for col_num, h_text in enumerate(cartera_headers, 1):
            cell = ws_cartera.cell(row=1, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border

        ws_cartera.row_dimensions[1].height = 24

        routes_rows = []
        for bu_item in data:
            bu_name = bu_item['total_business_unit_data']['business_unit_name']
            for r in bu_item['routes_data']:
                routes_rows.append((bu_name, r))

        routes_rows.sort(key=lambda x: int(x[1]['route_id']) if str(x[1]['route_id']).isdigit() else str(x[1]['route_id']))

        for row_idx, (bu_name, r) in enumerate(routes_rows, 2):
            r_id = str(r['route_id'])
            r_name = (r['route_name'] or r_id).title()
            reg_c = int(r['registered_customers'])
            act_c = int(r['active_customers'])
            eff_pct = float(r['portfolio_scope']) / 100.0
            new_c = int(r['new_customers'])

            row_values = [
                (r_id, '@', "center"),
                (r_name, None, "left"),
                (bu_name, None, "left"),
                (reg_c, int_format, "right"),
                (act_c, int_format, "right"),
                (eff_pct, pct_format, "right"),
                (new_c, int_format, "right"),
            ]

            for col_idx, (val, num_fmt, align_h) in enumerate(row_values, 1):
                cell = ws_cartera.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = cell_border
                cell.alignment = Alignment(horizontal=align_h)
                if num_fmt:
                    cell.number_format = num_fmt

        ws_ventas = wb.create_sheet(title="Ventas y Cuotas")
        ws_ventas.views.sheetView[0].showGridLines = True

        ventas_headers = [
            "Ruta",
            "Agente",
            "Gerencia",
            "Periodo",
            "Clase",
            "Cuota",
            "Venta Neta",
            "Diferencia",
            "% Alcance",
        ]

        for col_num, h_text in enumerate(ventas_headers, 1):
            cell = ws_ventas.cell(row=1, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border

        ws_ventas.row_dimensions[1].height = 24

        target_qs = service.targets_qs.filter(
            period__gte=service.date_start_dt,
            period__lte=service.date_end_dt
        )
        tx_qs = service.transactions_qs.filter(
            sale_date__gte=service.date_start_dt,
            sale_date__lte=service.date_end_dt
        )

        ventas_agrupadas = (
            tx_qs.annotate(periodo=TruncMonth('sale_date'))
            .values(
                'route_id',
                'route__name',
                'route__business_unit_id',
                'route__business_unit__name',
                'product_class_id',
                'product_class__name',
                'periodo',
            )
            .annotate(venta_neta=Sum('net_amount'))
        )

        cuotas_agrupadas = (
            target_qs.annotate(periodo=TruncMonth('period'))
            .values(
                'route_id',
                'route__name',
                'business_unit_id',
                'business_unit__name',
                'route__business_unit_id',
                'route__business_unit__name',
                'product_class_id',
                'product_class__name',
                'periodo',
            )
            .annotate(cuota=Sum('target_amount'))
        )

        consolidado = {}
        for v in ventas_agrupadas:
            rid = str(v['route_id'] or 'sn')
            rname = (v['route__name'] or rid).title()
            buname = (v['route__business_unit__name'] or 'General').title()
            cname = (v['product_class__name'] or 'otros').strip().title()
            p_dt = v['periodo']
            p_str = p_dt.strftime('%d/%m/%y') if hasattr(p_dt, 'strftime') else str(p_dt or '')
            key = (rid, buname, cname, p_str)
            consolidado[key] = {
                'ruta': rid,
                'agente': rname,
                'gerencia': buname,
                'periodo': p_str,
                'clase': cname,
                'cuota': Decimal('0.00'),
                'venta_neta': v['venta_neta'] or Decimal('0.00'),
            }

        for c in cuotas_agrupadas:
            rid = str(c['route_id'] or 'sn')
            rname = (c['route__name'] or rid).title()
            buname = (c['business_unit__name'] or c['route__business_unit__name'] or 'General').title()
            cname = (c['product_class__name'] or 'otros').strip().title()
            p_dt = c['periodo']
            p_str = p_dt.strftime('%d/%m/%y') if hasattr(p_dt, 'strftime') else str(p_dt or '')
            key = (rid, buname, cname, p_str)
            if key not in consolidado:
                consolidado[key] = {
                    'ruta': rid,
                    'agente': rname,
                    'gerencia': buname,
                    'periodo': p_str,
                    'clase': cname,
                    'cuota': c['cuota'] or Decimal('0.00'),
                    'venta_neta': Decimal('0.00'),
                }
            else:
                consolidado[key]['cuota'] += (c['cuota'] or Decimal('0.00'))

        ventas_data = list(consolidado.values())
        ventas_data.sort(key=lambda x: (int(x['ruta']) if str(x['ruta']).isdigit() else str(x['ruta']), x['clase'], x['periodo']))

        for row_idx, item in enumerate(ventas_data, 2):
            cuota_f = float(item['cuota'])
            venta_f = float(item['venta_neta'])
            diff_f = cuota_f - venta_f
            scope_f = (venta_f / cuota_f) if cuota_f > 0 else (1.0 if venta_f > 0 else 0.0)

            row_values = [
                (item['ruta'], '@', "center"),
                (item['agente'], None, "left"),
                (item['gerencia'], None, "left"),
                (item['periodo'], '@', "center"),
                (item['clase'], None, "left"),
                (cuota_f, currency_format, "right"),
                (venta_f, currency_format, "right"),
                (diff_f, currency_format, "right"),
                (scope_f, pct_format, "right"),
            ]

            for col_idx, (val, num_fmt, align_h) in enumerate(row_values, 1):
                cell = ws_ventas.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = cell_border
                cell.alignment = Alignment(horizontal=align_h)
                if num_fmt:
                    cell.number_format = num_fmt

        for sheet in [ws_summary, ws_cartera, ws_ventas]:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if cell.number_format == currency_format:
                        val_str = f"${val_str}"
                    max_len = max(max_len, len(val_str))
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


TargetAchievementService.Exports = TargetAchievementExports