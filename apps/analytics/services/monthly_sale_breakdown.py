from dataclasses import dataclass, field
from decimal import Decimal
from collections import defaultdict
import io
import calendar
from datetime import date
from typing import Any
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.db.models import QuerySet, Sum, Count
from django.db.models.functions import ExtractMonth
from django.utils import timezone

from apps.products.models import ProductClass
from apps.human_resources.models import BusinessUnit


@dataclass
class MonthlySaleBreakdownService:
    user: Any
    targets_qs: QuerySet
    transactions_qs: QuerySet
    customers_qs: QuerySet
    ars_qs: QuerySet
    routes_qs: QuerySet
    year: int | str | None = None
    cleaned_data: dict[str, Any] | None = None

    today: date = field(init=False)
    year_int: int = field(init=False)

    def __post_init__(self):
        self.today = timezone.localdate()
        try:
            self.year_int = int(self.year) if self.year else self.today.year
        except (ValueError, TypeError):
            self.year_int = self.today.year
        self.year = self.year_int
        self._apply_cleaned_data_filters()

    @property
    def is_vendor(self) -> bool:
        """
        returns true if user belongs to vendedor group.
        """
        if not self.user or not hasattr(self.user, 'groups'):
            return False
        return self.user.groups.filter(name='vendedor').exists()

    def _get_margin_status(self, margin: Decimal | float) -> str:
        """
        evaluates qualitative margin status based on percentage threshold.
        """
        if margin >= 43:
            return 'Excelente'
        elif margin >= 40:
            return 'Óptimo'
        elif margin >= 37:
            return 'Regular'
        elif margin >= 35:
            return 'Malo'
        else:
            return 'Muy malo'

    def _apply_cleaned_data_filters(self):
        """
        applies filter criteria from cleaned_data to secondary querysets (targets, routes, customers, ars).
        """
        if not self.cleaned_data:
            return

        if self.cleaned_data.get('route'):
            route_val = self.cleaned_data['route']
            r_ids = [r.id if hasattr(r, 'id') else r for r in route_val]
            self.routes_qs = self.routes_qs.filter(id__in=r_ids)
            self.targets_qs = self.targets_qs.filter(route_id__in=r_ids)
            self.customers_qs = self.customers_qs.filter(assignments__route_id__in=r_ids)
            self.ars_qs = self.ars_qs.filter(customer__assignments__route_id__in=r_ids)

        if self.cleaned_data.get('business_unit'):
            bu_val = self.cleaned_data['business_unit']
            bu_ids = [bu.id if hasattr(bu, 'id') else bu for bu in bu_val]
            self.routes_qs = self.routes_qs.filter(business_unit_id__in=bu_ids)
            self.targets_qs = self.targets_qs.filter(route__business_unit_id__in=bu_ids)
            self.customers_qs = self.customers_qs.filter(assignments__route__business_unit_id__in=bu_ids)
            self.ars_qs = self.ars_qs.filter(customer__assignments__route__business_unit_id__in=bu_ids)

        if self.cleaned_data.get('region'):
            reg_val = self.cleaned_data['region']
            reg_ids = [reg.id if hasattr(reg, 'id') else reg for reg in reg_val]
            all_bu_ids = set(reg_ids)
            current_parents = set(reg_ids)
            while current_parents:
                child_ids = set(
                    BusinessUnit.objects.filter(parent_id__in=current_parents).values_list('id', flat=True)
                )
                new_ids = child_ids - all_bu_ids
                if not new_ids:
                    break
                all_bu_ids.update(new_ids)
                current_parents = new_ids

            self.routes_qs = self.routes_qs.filter(business_unit_id__in=all_bu_ids)
            self.targets_qs = self.targets_qs.filter(route__business_unit_id__in=all_bu_ids)
            self.customers_qs = self.customers_qs.filter(assignments__route__business_unit_id__in=all_bu_ids)
            self.ars_qs = self.ars_qs.filter(customer__assignments__route__business_unit_id__in=all_bu_ids)

        if self.cleaned_data.get('product_class'):
            self.targets_qs = self.targets_qs.filter(product_class__in=self.cleaned_data['product_class'])

        if self.cleaned_data.get('product_category'):
            self.targets_qs = self.targets_qs.filter(product_class__product_category__in=self.cleaned_data['product_category'])

        self.transactions_qs = self.transactions_qs.filter(route__in=self.routes_qs)


    def _get_monthly_targets(self) -> dict[tuple[str, str, int], Decimal]:
        """
        calculates monthly targets per route, product class and month.
        key: (route_id, product_class_id_or_name_lower, month_1_to_12) -> Decimal
        """
        targets = self.targets_qs.filter(
            period__year=self.year_int
        ).order_by().annotate(
            month=ExtractMonth('period')
        ).values('route_id', 'product_class__name', 'month').annotate(
            total=Sum('target_amount')
        )

        data = defaultdict(Decimal)
        for row in targets:
            r_id = str(row['route_id']) if row['route_id'] else ''
            cls_name = str(row['product_class__name'] or '').strip().lower()
            m = row['month']
            data[(r_id, cls_name, m)] += row['total'] or Decimal('0.00')
        return data

    def _get_monthly_sales(self) -> dict[tuple[str, str, int], Decimal]:
        """
        calculates monthly net sales per route, product class and month.
        key: (route_id, product_class_name_lower, month_1_to_12) -> Decimal
        """
        sales = self.transactions_qs.filter(
            sale_date__year=self.year_int
        ).order_by().annotate(
            month=ExtractMonth('sale_date')
        ).values('route_id', 'product_class__name', 'month').annotate(
            total=Sum('net_amount')
        )

        data = defaultdict(Decimal)
        for row in sales:
            r_id = str(row['route_id']) if row['route_id'] else ''
            cls_name = str(row['product_class__name'] or '').strip().lower()
            m = row['month']
            data[(r_id, cls_name, m)] += row['total'] or Decimal('0.00')
        return data

    def _get_monthly_margins(self) -> dict[tuple[str, int], dict[str, Decimal]]:
        """
        calculates profit, net sale and profit margin percentage per route and month.
        key: (route_id, month_1_to_12) -> {'margin': Decimal, 'profit': Decimal, 'net': Decimal}
        """
        margins = self.transactions_qs.filter(
            sale_date__year=self.year_int
        ).order_by().annotate(
            month=ExtractMonth('sale_date')
        ).values('route_id', 'month').annotate(
            total_profit=Sum('profit'),
            total_net=Sum('net_amount')
        )

        data = defaultdict(lambda: {'margin': Decimal('0.00'), 'profit': Decimal('0.00'), 'net': Decimal('0.00')})
        for row in margins:
            r_id = str(row['route_id']) if row['route_id'] else ''
            m = row['month']
            t_profit = row['total_profit'] or Decimal('0.00')
            t_net = row['total_net'] or Decimal('0.00')
            if t_net > 0:
                margin = (t_profit / t_net) * Decimal('100.00')
            else:
                margin = Decimal('0.00')

            data[(r_id, m)] = {
                'margin': margin,
                'profit': t_profit,
                'net': t_net
            }
        return data

    def _get_monthly_new_customers(self) -> dict[tuple[str, int], int]:
        """
        calculates new customers registered per route and month in the target year.
        key: (route_id, month_1_to_12) -> int
        """
        new_cust_qs = self.customers_qs.filter(
            registration_date__year=self.year_int
        ).order_by().annotate(
            month=ExtractMonth('registration_date')
        ).values('current_route_id', 'month').annotate(
            total=Count('id', distinct=True)
        )

        data = defaultdict(int)
        for row in new_cust_qs:
            r_id = str(row['current_route_id']) if row['current_route_id'] else ''
            if r_id:
                data[(r_id, row['month'])] += row['total'] or 0
        return data

    def _get_monthly_accounts_receivable(self) -> tuple[dict[tuple[str, int], int], dict[tuple[str, int], Decimal], dict[tuple[str, int], set]]:
        """
        calculates accounts receivable from the perspective of currently assigned customers:
        - unique customer count (distinct customer count)
        - total balance amount ($)
        filtered month by month based on invoices issued on or before the end of each month.
        """
        month_end_dates = {}
        for m in range(1, 13):
            last_day = calendar.monthrange(self.year_int, m)[1]
            month_end_dates[m] = date(self.year_int, m, last_day)

        ar_qs = self.ars_qs.filter(
            customer__assignments__end_date__isnull=True
        ).values(
            'customer_id',
            'customer__assignments__route_id',
            'issue_date',
            'total_balance'
        ).distinct()

        amount_data = defaultdict(Decimal)
        customers_per_route_month = defaultdict(set)

        for row in ar_qs:
            r_id = str(row['customer__assignments__route_id'] or '')
            issue_d = row['issue_date']
            balance = row['total_balance'] or Decimal('0.00')
            cid = str(row['customer_id'])

            if r_id:
                for m in range(1, 13):
                    if issue_d is None or issue_d <= month_end_dates[m]:
                        amount_data[(r_id, m)] += balance
                        customers_per_route_month[(r_id, m)].add(cid)

        count_data = defaultdict(int)
        for key, customers_set in customers_per_route_month.items():
            count_data[key] = len(customers_set)

        return count_data, amount_data, customers_per_route_month

    def _get_monthly_promotions(self) -> dict[tuple[str, int], Decimal]:
        """promotions placeholder: instantiated at zeros for now"""
        return defaultdict(Decimal)

    def _get_monthly_agreements(self) -> dict[tuple[str, int], int]:
        """agreements placeholder: instantiated at zeros for now"""
        return defaultdict(int)

    def get_data(self) -> list[dict[str, Any]]:
        """
        assembles the complete structured monthly breakdown by Gerencia (BusinessUnit) and Route.
        """
        # Determine product classes to display
        if self.cleaned_data and self.cleaned_data.get('product_class'):
            pcs = self.cleaned_data['product_class']
            pc_names = sorted(list(set(str(pc.name or pc.id).strip().lower() for pc in pcs)))
        elif self.cleaned_data and self.cleaned_data.get('product_category'):
            cats = self.cleaned_data['product_category']
            pcs = ProductClass.objects.filter(product_category__in=cats)
            pc_names = sorted(list(set(str(pc.name or pc.id).strip().lower() for pc in pcs)))
        else:
            all_pcs = ProductClass.objects.values_list('name', flat=True)
            pc_names = sorted(list(set(str(pc).strip().lower() for pc in all_pcs if pc)))

        targets = self._get_monthly_targets()
        sales = self._get_monthly_sales()
        margins = self._get_monthly_margins()
        new_customers = self._get_monthly_new_customers()
        ar_counts, ar_amounts, ar_cust_sets = self._get_monthly_accounts_receivable()
        promotions = self._get_monthly_promotions()
        agreements = self._get_monthly_agreements()

        # Group routes by BusinessUnit (Gerencia)
        bu_map: dict[str, dict[str, Any]] = {}
        routes_list = list(self.routes_qs.select_related('business_unit').order_by('business_unit__name', 'id'))

        for route in routes_list:
            bu = route.business_unit
            bu_id = str(bu.id) if bu else 'GENERAL'
            bu_name = str(bu.name).title() if bu else 'General / Sin Gerencia'

            if bu_id not in bu_map:
                bu_map[bu_id] = {
                    'business_unit_id': bu_id,
                    'business_unit_name': bu_name,
                    'routes_objs': []
                }
            bu_map[bu_id]['routes_objs'].append(route)

        results_names = [
            'margen',
            'clientes nuevos',
            'cuentas por cobrar',
            'cuentas por cobrar $',
            'promociones',
            'convenios'
        ]

        breakdown_data = []

        for bu_id, bu_info in bu_map.items():
            bu_name = bu_info['business_unit_name']
            bu_routes = bu_info['routes_objs']

            temp_bu = {
                'business_unit_id': bu_id,
                'business_unit_name': bu_name,
                'routes': []
            }

            #business unit accumulators
            bu_pc_data = {
                pc: {m: {'target': Decimal('0.00'), 'net_sale': Decimal('0.00')} for m in range(1, 13)}
                for pc in pc_names
            }
            bu_total_data = {m: {'target': Decimal('0.00'), 'net_sale': Decimal('0.00')} for m in range(1, 13)}
            bu_results_data = {res: {m: Decimal('0.00') if '$' in res or res == 'promociones' else 0 for m in range(1, 13)} for res in results_names}
            bu_margin_acc = {m: {'profit': Decimal('0.00'), 'net': Decimal('0.00')} for m in range(1, 13)}
            bu_ar_customers = {m: set() for m in range(1, 13)}

            for route in bu_routes:
                route_id = str(route.id)
                route_name = str(route.name or f"Ruta {route.id}").title()

                temp_route = {
                    'route_id': route_id,
                    'route_name': route_name,
                    'product_classes': [],
                    'total': {'name': 'total', 'monthly_data': []},
                    'results': []
                }

                route_monthly_totals = {
                    m: {'target': Decimal('0.00'), 'net_sale': Decimal('0.00')} for m in range(1, 13)
                }

                for pc_name in pc_names:
                    temp_pc = {'name': pc_name, 'monthly_data': []}
                    for month in range(1, 13):
                        t = targets.get((route_id, pc_name, month), Decimal('0.00'))
                        s = sales.get((route_id, pc_name, month), Decimal('0.00'))
                        diff = s - t
                        scope = (s / t * Decimal('100.00')) if t > 0 else (Decimal('100.00') if s > 0 else Decimal('0.00'))

                        temp_pc['monthly_data'].append({
                            'target': t,
                            'net_sale': s,
                            'diff': diff,
                            'scope': scope
                        })

                        route_monthly_totals[month]['target'] += t
                        route_monthly_totals[month]['net_sale'] += s

                        bu_pc_data[pc_name][month]['target'] += t
                        bu_pc_data[pc_name][month]['net_sale'] += s

                    temp_route['product_classes'].append(temp_pc)

                for month in range(1, 13):
                    t = route_monthly_totals[month]['target']
                    s = route_monthly_totals[month]['net_sale']
                    diff = s - t
                    scope = (s / t * Decimal('100.00')) if t > 0 else (Decimal('100.00') if s > 0 else Decimal('0.00'))

                    temp_route['total']['monthly_data'].append({
                        'target': t,
                        'net_sale': s,
                        'diff': diff,
                        'scope': scope
                    })

                    bu_total_data[month]['target'] += t
                    bu_total_data[month]['net_sale'] += s
                    bu_ar_customers[month].update(ar_cust_sets.get((route_id, month), set()))

                for res_name in results_names:
                    temp_res = {'name': res_name, 'monthly_data': []}
                    for month in range(1, 13):
                        val_display = "0"
                        item_margin = 0.0
                        item_status = ""
                        if res_name == 'margen':
                            m_data = margins.get((route_id, month), {'margin': Decimal('0.00'), 'profit': Decimal('0.00'), 'net': Decimal('0.00')})
                            val_display = f"{m_data['margin']:,.2f} %"
                            item_margin = float(m_data['margin'])
                            item_status = self._get_margin_status(m_data['margin'])
                            bu_margin_acc[month]['profit'] += m_data['profit']
                            bu_margin_acc[month]['net'] += m_data['net']
                        elif res_name == 'clientes nuevos':
                            cnt = new_customers.get((route_id, month), 0)
                            val_display = str(cnt)
                            bu_results_data[res_name][month] += cnt
                        elif res_name == 'cuentas por cobrar':
                            cnt = ar_counts.get((route_id, month), 0)
                            val_display = str(cnt)
                            bu_results_data[res_name][month] += cnt
                        elif res_name == 'cuentas por cobrar $':
                            amt = ar_amounts.get((route_id, month), Decimal('0.00'))
                            val_display = f"$ {amt:,.2f}"
                            bu_results_data[res_name][month] += amt
                        elif res_name == 'promociones':
                            p_amt = promotions.get((route_id, month), Decimal('0.00'))
                            val_display = f"$ {p_amt:,.2f}"
                            bu_results_data[res_name][month] += p_amt
                        elif res_name == 'convenios':
                            c_cnt = agreements.get((route_id, month), 0)
                            val_display = str(c_cnt)
                            bu_results_data[res_name][month] += c_cnt

                        data_entry = {'value': val_display}
                        if res_name == 'margen':
                            data_entry['margin'] = item_margin
                            data_entry['status'] = item_status
                        temp_res['monthly_data'].append(data_entry)
                    temp_route['results'].append(temp_res)

                temp_bu['routes'].append(temp_route)

            bu_summary = {
                'route_id': 'TOTAL',
                'route_name': bu_name,
                'product_classes': [],
                'total': {'name': 'total', 'monthly_data': []},
                'results': []
            }

            for pc_name in pc_names:
                temp_pc = {'name': pc_name, 'monthly_data': []}
                for month in range(1, 13):
                    t = bu_pc_data[pc_name][month]['target']
                    s = bu_pc_data[pc_name][month]['net_sale']
                    diff = s - t
                    scope = (s / t * Decimal('100.00')) if t > 0 else (Decimal('100.00') if s > 0 else Decimal('0.00'))
                    temp_pc['monthly_data'].append({
                        'target': t,
                        'net_sale': s,
                        'diff': diff,
                        'scope': scope
                    })
                bu_summary['product_classes'].append(temp_pc)

            for month in range(1, 13):
                t = bu_total_data[month]['target']
                s = bu_total_data[month]['net_sale']
                diff = s - t
                scope = (s / t * Decimal('100.00')) if t > 0 else (Decimal('100.00') if s > 0 else Decimal('0.00'))
                bu_summary['total']['monthly_data'].append({
                    'target': t,
                    'net_sale': s,
                    'diff': diff,
                    'scope': scope
                })

            for res_name in results_names:
                temp_res = {'name': res_name, 'monthly_data': []}
                for month in range(1, 13):
                    item_margin = 0.0
                    item_status = ""
                    if res_name == 'margen':
                        t_profit = bu_margin_acc[month]['profit']
                        t_net = bu_margin_acc[month]['net']
                        if t_net > 0:
                            margin = (t_profit / t_net) * Decimal('100.00')
                        else:
                            margin = Decimal('0.00')
                        val_display = f"{margin:,.2f} %"
                        item_margin = float(margin)
                        item_status = self._get_margin_status(margin)
                    elif res_name == 'clientes nuevos':
                        val_display = str(bu_results_data[res_name][month])
                    elif res_name == 'cuentas por cobrar':
                        val_display = str(bu_results_data[res_name][month])
                    elif res_name == 'cuentas por cobrar $':
                        val_display = f"$ {bu_results_data[res_name][month]:,.2f}"
                    elif res_name == 'promociones':
                        val_display = f"$ {bu_results_data[res_name][month]:,.2f}"
                    elif res_name == 'convenios':
                        val_display = str(bu_results_data[res_name][month])

                    data_entry = {'value': val_display}
                    if res_name == 'margen':
                        data_entry['margin'] = item_margin
                        data_entry['status'] = item_status
                    temp_res['monthly_data'].append(data_entry)
                bu_summary['results'].append(temp_res)

            temp_bu['routes'].insert(0, bu_summary)
            breakdown_data.append(temp_bu)

        return breakdown_data


@dataclass
class MonthlySaleBreakdownExports:
    monthly_sale_breakdown_service: MonthlySaleBreakdownService

    def export_monthly_sale_breakdown_report(self) -> io.BytesIO:
        """
        generate an Excel workbook with detailed monthly breakdown data.
        Follows the legacy design:
        - Creates one sheet per warehouse/gerencia.
        - Alternating row styling with custom teal, cyan, gold and gray fills.
        - Summarizes target sales vs actual sales for each product class.
        - Includes route-level performance metrics.
        - Provides KPIs like new customers, accounts receivable, promotions, and agreements.
        """
        wb = openpyxl.Workbook()
        if wb.worksheets:
            wb.remove(wb.active)

        #style config
        HEADER_FILL = PatternFill(start_color="2D9999", fill_type="solid")
        ALT_CLASS_1 = PatternFill(start_color="FFFFFF", fill_type="solid")
        ALT_CLASS_2 = PatternFill(start_color="A4E0E6", fill_type="solid")
        TOTAL_FILL = PatternFill(start_color="FAC003", fill_type="solid")
        ALT_SUM_1 = PatternFill(start_color="D9D9D9", fill_type="solid")
        ALT_SUM_2 = PatternFill(start_color="FFFFFF", fill_type="solid")

        WHITE_BOLD = Font(color="FFFFFF", bold=True)
        BLACK_BOLD = Font(color="000000", bold=True)
        BLACK = Font(color="000000")
        EXCELLENT_FONT = Font(color="059669", bold=True)
        OPTIMAL_FONT = Font(color="10B981", bold=True)
        REGULAR_FONT = Font(color="D97706", bold=True)
        BAD_FONT = Font(color="EF4444", bold=True)
        CENTER = Alignment(horizontal="center", vertical="center")
        LEFT = Alignment(horizontal="left", vertical="center")

        THIN_BORDER = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        def style_cell(cell, fill, font=None, alignment=None, number_format=None):
            cell.fill = fill
            cell.border = THIN_BORDER
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if number_format:
                cell.number_format = number_format
            return cell

        month_names = [
            'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
            'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE'
        ]

        report_data = self.monthly_sale_breakdown_service.get_data()

        if not report_data:
            ws = wb.create_sheet(title="SIN DATOS")
            ws.cell(row=1, column=1, value="No se encontraron datos o rutas asignadas para el periodo seleccionado.")
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer

        existing_titles = set()

        for w_data in report_data:
            base_title = (w_data.get('business_unit_name') or 'Gerencia')[:31].replace('/', '-').replace('\\', '').replace('?', '').replace('*', '').replace(':', '').replace('[', '').replace(']', '').upper()
            sheet_title = base_title
            counter = 1
            while sheet_title in existing_titles:
                suffix = f"_{counter}"
                sheet_title = f"{base_title[:31-len(suffix)]}{suffix}"
                counter += 1
            existing_titles.add(sheet_title)

            ws = wb.create_sheet(title=sheet_title)
            current_row = 1

            for route in w_data.get('routes', []):
                # headers
                headers = ["AGENTE", "CLASE"]
                for m_name in month_names:
                    headers.extend([f"OBJETIVO {m_name}", f"VENTA {m_name}", f"ALCANCE {m_name}", f"DIF. {m_name}"])

                for col_num, val in enumerate(headers, 1):
                    c = ws.cell(row=current_row, column=col_num, value=val)
                    style_cell(c, HEADER_FILL, WHITE_BOLD, CENTER)
                current_row += 1

                # product class rows
                row_idx = 0
                for pc in route.get('product_classes', []):
                    fill_color = ALT_CLASS_1 if row_idx % 2 == 0 else ALT_CLASS_2

                    style_cell(ws.cell(row=current_row, column=1, value=route.get('route_id')), fill_color, BLACK, CENTER)
                    style_cell(ws.cell(row=current_row, column=2, value=pc.get('name', '').upper()), fill_color, BLACK, LEFT)

                    col = 3
                    for m_data in pc.get('monthly_data', []):
                        style_cell(ws.cell(row=current_row, column=col, value=float(m_data['target'])), fill_color, BLACK, CENTER, '"$"#,##0.00')
                        style_cell(ws.cell(row=current_row, column=col + 1, value=float(m_data['net_sale'])), fill_color, BLACK, CENTER, '"$"#,##0.00')
                        style_cell(ws.cell(row=current_row, column=col + 2, value=float(m_data['scope']) / 100), fill_color, BLACK, CENTER, '0.00%')
                        style_cell(ws.cell(row=current_row, column=col + 3, value=float(m_data['diff'])), fill_color, BLACK, CENTER, '"$"#,##0.00')
                        col += 4

                    current_row += 1
                    row_idx += 1

                # total
                style_cell(ws.cell(row=current_row, column=1, value=route.get('route_id')), TOTAL_FILL, BLACK_BOLD, CENTER)
                style_cell(ws.cell(row=current_row, column=2, value="TOTAL"), TOTAL_FILL, BLACK_BOLD, LEFT)

                col = 3
                for m_data in route.get('total', {}).get('monthly_data', []):
                    style_cell(ws.cell(row=current_row, column=col, value=float(m_data['target'])), TOTAL_FILL, BLACK_BOLD, CENTER, '"$"#,##0.00')
                    style_cell(ws.cell(row=current_row, column=col + 1, value=float(m_data['net_sale'])), TOTAL_FILL, BLACK_BOLD, CENTER, '"$"#,##0.00')
                    style_cell(ws.cell(row=current_row, column=col + 2, value=float(m_data['scope']) / 100), TOTAL_FILL, BLACK_BOLD, CENTER, '0.00%')
                    style_cell(ws.cell(row=current_row, column=col + 3, value=float(m_data['diff'])), TOTAL_FILL, BLACK_BOLD, CENTER, '"$"#,##0.00')
                    col += 4
                current_row += 1

                # results rows
                res_idx = 0
                for res in route.get('results', []):
                    fill_color = ALT_SUM_1 if res_idx % 2 == 0 else ALT_SUM_2

                    style_cell(ws.cell(row=current_row, column=1, value=route.get('route_id')), fill_color, BLACK_BOLD, CENTER)
                    style_cell(ws.cell(row=current_row, column=2, value=res.get('name', '').upper()), fill_color, BLACK_BOLD, LEFT)

                    col = 3
                    for m_data in res.get('monthly_data', []):
                        res_name = res.get('name', '').lower()
                        ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 3)

                        if 'margen' in res_name:
                            margin_val = m_data.get('margin', 0.0)
                            if self.monthly_sale_breakdown_service.is_vendor:
                                status_str = m_data.get('status') or 'Regular'
                                if margin_val >= 43:
                                    status_font = EXCELLENT_FONT
                                elif margin_val >= 40:
                                    status_font = OPTIMAL_FONT
                                elif margin_val >= 37:
                                    status_font = REGULAR_FONT
                                else:
                                    status_font = BAD_FONT
                                c_val = ws.cell(row=current_row, column=col, value=status_str)
                                style_cell(c_val, fill_color, status_font, CENTER)
                            else:
                                val = float(margin_val) / 100.0 if margin_val is not None else 0.0
                                c_val = ws.cell(row=current_row, column=col, value=val)
                                style_cell(c_val, fill_color, BLACK_BOLD, CENTER, '0.00%')
                        else:
                            val_str = str(m_data.get('value', '0')).replace('$', '').replace('%', '').replace(',', '').strip()
                            try:
                                val = float(val_str)
                            except ValueError:
                                val = 0.0

                            num_format = '0'
                            if '$' in res_name or 'promociones' in res_name:
                                num_format = '"$"#,##0.00'

                            c_val = ws.cell(row=current_row, column=col, value=val)
                            style_cell(c_val, fill_color, BLACK_BOLD, CENTER, num_format)

                        # outline
                        style_cell(ws.cell(row=current_row, column=col + 1), fill_color)
                        style_cell(ws.cell(row=current_row, column=col + 2), fill_color)
                        style_cell(ws.cell(row=current_row, column=col + 3), fill_color)

                        col += 4

                    current_row += 1
                    res_idx += 1

                # alternate space
                current_row += 1

            # column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 25
            for col_idx in range(3, 51):
                ws.column_dimensions[get_column_letter(col_idx)].width = 16

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
