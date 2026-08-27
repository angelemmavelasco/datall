import calendar
import io
import math
import statistics
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from typing import Any
from dateutil.relativedelta import relativedelta
from collections import defaultdict
from django.db.models import QuerySet, Sum, Q
from django.utils import timezone
from apps.core.models import Reference

@dataclass
class CommercialRiskService:
    user: Any
    route: Any
    customers_qs: QuerySet
    transactions_qs: QuerySet
    date_start: date | str | None = None
    date_end: date | str | None = None
    cleaned_data: dict[str, Any] | None = None

    today: date = field(init=False)
    route_id: str | None = field(init=False)
    start_q_date: date = field(init=False)
    end_q_date: date = field(init=False)

    def __post_init__(self):
        self._init_dates()
        self._init_route()

    def _init_dates(self) -> None:
        self.today = timezone.localdate()

        def _parse_val(val: Any) -> date | None:
            if not val:
                return None
            if isinstance(val, date) and not isinstance(val, datetime):
                return val
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, str):
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
                    try:
                        return datetime.strptime(val.strip(), fmt).date()
                    except ValueError:
                        continue
            return None

        self.date_start = _parse_val(self.date_start)
        self.date_end = _parse_val(self.date_end)

        first_day_curr_month = self.today.replace(day=1)
        default_end = first_day_curr_month - relativedelta(days=1)
        default_start = default_end.replace(day=1) - relativedelta(months=11)

        if not self.date_end:
            self.date_end = default_end
        if not self.date_start:
            self.date_start = default_start

        if self.date_start > self.date_end:
            self.date_start, self.date_end = self.date_end, self.date_start

        self.end_q_date = self.date_end
        self.start_q_date = self.end_q_date.replace(day=1) - relativedelta(months=2)

    def _init_route(self) -> None:
        if hasattr(self.route, 'id'):
            self.route_id = str(self.route.id)
        elif self.route:
            self.route_id = str(self.route)
        else:
            self.route_id = None

    def stats(self) -> dict[str, Any]:
        return CommercialRiskStats(commercial_risk_service=self).stats()

    def _get_timeline_chart(self) -> dict[str, Any]:
        """
        generates timeline chart data for commercial risk evolution evaluated month by month
        using rolling closed quarters (e.g. for Jan 2025: Oct, Nov, Dec 2024; for Feb 2025: Nov, Dec 2024, Jan 2025).
        Returns:
            - timeline_months: ['2025-08', '2025-09', ...]
            - months_labels: ['Ago 2025', 'Sep 2025', ...]
            - commercial_risk_index: [81.64, 80.88, ...]
            - gini: [76.49, 75.18, ...]
            - portfolio_scope_complement: [86.78, 86.57, ...] (Clientes desatendidos)
            - portfolio_scope: [13.22, 13.43, ...] (Alcance de cartera)
        """
        if not self.route_id:
            return {
                'timeline_months': [],
                'months_labels': [],
                'commercial_risk_index': [],
                'gini': [],
                'portfolio_scope_complement': [],
                'portfolio_scope': [],
            }

        curr = self.date_start.replace(day=1)
        end_m = self.date_end.replace(day=1)
        timeline_months = []
        months_labels = []

        month_abbr_es = {
            1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
            7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
        }

        while curr <= end_m:
            timeline_months.append(curr.strftime('%Y-%m'))
            months_labels.append(f"{month_abbr_es[curr.month]} {curr.year}")
            curr += relativedelta(months=1)

        if not timeline_months:
            return {
                'timeline_months': [],
                'months_labels': [],
                'commercial_risk_index': [],
                'gini': [],
                'portfolio_scope_complement': [],
                'portfolio_scope': [],
            }

        customers_qs = (
            self.customers_qs
            .filter(assignments__route_id=self.route_id)
            .values('id', 'registration_date')
            .distinct()
        )
        customers_map = {c['id']: c['registration_date'] for c in customers_qs}

        extended_start = self.date_start.replace(day=1) - relativedelta(months=4)
        tx_qs = (
            self.transactions_qs
            .filter(
                route_id=self.route_id,
                sale_date__gte=extended_start,
                sale_date__lte=self.date_end,
                net_amount__gt=0
            )
            .values('sale_date', 'customer_id', 'net_amount')
        )

        monthly_customer_sales = defaultdict(lambda: defaultdict(float))
        for tx in tx_qs:
            m_key = tx['sale_date'].strftime('%Y-%m')
            monthly_customer_sales[m_key][tx['customer_id']] += float(tx['net_amount'])

        gini_list = []
        unattended_list = []
        scope_list = []
        irc_list = []

        for m_str in timeline_months:
            target_date = datetime.strptime(m_str, '%Y-%m').date()
            end_q = target_date.replace(day=1) - relativedelta(days=1)
            q_months = [(end_q.replace(day=1) - relativedelta(months=i)).strftime('%Y-%m') for i in (2, 1, 0)]

            portfolio_cids = [
                cid for cid, reg_date in customers_map.items()
                if reg_date and reg_date <= end_q
            ]
            total_portfolio = len(portfolio_cids)

            quarter_customer_sales = defaultdict(float)
            for qm in q_months:
                for cid in portfolio_cids:
                    amt = monthly_customer_sales[qm].get(cid, 0.0)
                    if amt > 0.0:
                        quarter_customer_sales[cid] += amt

            active_count = sum(1 for cid in portfolio_cids if quarter_customer_sales.get(cid, 0.0) > 0.0)

            if total_portfolio > 0:
                scope = min(1.0, active_count / total_portfolio)
            else:
                scope = 0.0

            unattended = max(0.0, 1.0 - scope)

            sales_arr = sorted([quarter_customer_sales[cid] for cid in portfolio_cids if quarter_customer_sales.get(cid, 0.0) > 0.0])
            n = len(sales_arr)
            cum_sales = sum(sales_arr)

            if n > 0 and cum_sales > 0:
                sum_iy = sum((i + 1) * y for i, y in enumerate(sales_arr))
                gini = (2.0 * sum_iy / (n * cum_sales)) - ((n + 1.0) / n)
                gini = max(min(gini, 1.0), 0.0)
            else:
                gini = 0.0

            irc = (0.5 * gini) + (0.5 * unattended)

            gini_list.append(round(gini * 100.0, 2))
            unattended_list.append(round(unattended * 100.0, 2))
            scope_list.append(round(scope * 100.0, 2))
            irc_list.append(round(irc * 100.0, 2))

        return {
            'timeline_months': timeline_months,
            'months_labels': months_labels,
            'commercial_risk_index': irc_list,
            'gini': gini_list,
            'portfolio_scope_complement': unattended_list,
            'portfolio_scope': scope_list,
        }

    def _get_customer_churn(self) -> dict[str, Any]:
        """
        calculates lost and won customers month by month.
        a customer is 'won' in month M if they purchased in month M but not in M-1.
        a customer is 'lost' in month M if they purchased in month M-1 but not in month M.
        also stores customer IDs for each month to allow interactive navigation.
        returns:
            {
                'months': ['ene', 'feb', ...],
                'lost': [2, 5, ...],
                'won': [4, 1, ...],
                'lost_customer_ids': [['id1', 'id2'], ...],
                'won_customer_ids': [['id3', 'id4'], ...],
            }
        """
        if not self.route_id:
            return {'months': [], 'lost': [], 'won': [], 'lost_customer_ids': [], 'won_customer_ids': []}

        curr = self.date_start.replace(day=1)
        end_m = self.date_end.replace(day=1)
        eval_periods: list[tuple[int, int]] = []
        while curr <= end_m:
            eval_periods.append((curr.year, curr.month))
            curr += relativedelta(months=1)

        if not eval_periods:
            return {'months': [], 'lost': [], 'won': [], 'lost_customer_ids': [], 'won_customer_ids': []}

        first_y, first_m = eval_periods[0]
        baseline_period = (first_y - 1, 12) if first_m == 1 else (first_y, first_m - 1)
        all_periods = [baseline_period] + eval_periods

        start_bound = date(all_periods[0][0], all_periods[0][1], 1)
        _, last_day_end = calendar.monthrange(all_periods[-1][0], all_periods[-1][1])
        end_bound = date(all_periods[-1][0], all_periods[-1][1], last_day_end)

        assigned_cids = set(
            self.customers_qs
            .filter(assignments__route_id=self.route_id)
            .values_list('id', flat=True)
            .distinct()
        )

        tx_churn = (
            self.transactions_qs
            .filter(
                route_id=self.route_id,
                customer_id__in=assigned_cids,
                sale_date__gte=start_bound,
                sale_date__lte=end_bound,
                net_amount__gt=0
            )
            .order_by()
            .values('sale_date__year', 'sale_date__month', 'customer_id')
            .distinct()
        )

        cust_by_period: dict[tuple[int, int], set[Any]] = defaultdict(set)
        for row in tx_churn:
            y = row['sale_date__year']
            m = row['sale_date__month']
            cid = row['customer_id']
            if cid:
                cust_by_period[(y, m)].add(cid)

        month_names = {
            1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
            7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
        }
        has_multiple_years = len(set(y for y, _ in eval_periods)) > 1

        months: list[str] = []
        lost_counts: list[int] = []
        won_counts: list[int] = []
        lost_customer_ids: list[list[str]] = []
        won_customer_ids: list[list[str]] = []

        for i in range(1, len(all_periods)):
            prev_p = all_periods[i - 1]
            curr_p = all_periods[i]

            prev_custs = cust_by_period[prev_p]
            curr_custs = cust_by_period[curr_p]

            lost_set = prev_custs - curr_custs
            won_set = curr_custs - prev_custs

            m_label = f"{month_names[curr_p[1]]} '{str(curr_p[0])[-2:]}" if has_multiple_years else month_names[curr_p[1]]

            months.append(m_label)
            lost_counts.append(len(lost_set))
            won_counts.append(len(won_set))
            lost_customer_ids.append(sorted([str(c) for c in lost_set]))
            won_customer_ids.append(sorted([str(c) for c in won_set]))

        return {
            'months': months,
            'lost': lost_counts,
            'won': won_counts,
            'lost_customer_ids': lost_customer_ids,
            'won_customer_ids': won_customer_ids,
        } 

    def _get_monthly_new_customers(self) -> dict[str, Any]:
        """
        calculates the count of new customers registered in this route month by month.
        returns:
            {
                'months': ["Ago '25", "Sep '25", ...],
                'new_customers': [3, 1, 5, ...],
                'new_customer_ids': [['id1', 'id2'], ...],
            }
        """
        if not self.route_id:
            return {'months': [], 'new_customers': [], 'new_customer_ids': []}

        curr = self.date_start.replace(day=1)
        end_m = self.date_end.replace(day=1)
        eval_periods: list[tuple[int, int]] = []
        month_labels: list[str] = []

        month_names = {
            1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
            7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
        }

        while curr <= end_m:
            eval_periods.append((curr.year, curr.month))
            curr += relativedelta(months=1)

        has_multiple_years = len(set(y for y, _ in eval_periods)) > 1
        for y, m in eval_periods:
            m_label = f"{month_names[m]} '{str(y)[-2:]}" if has_multiple_years else month_names[m]
            month_labels.append(m_label)

        customers_qs = (
            self.customers_qs
            .filter(assignments__route_id=self.route_id)
            .values('id', 'registration_date')
            .distinct()
        )

        counts = []
        cids_list = []
        for y, m in eval_periods:
            m_cids = [
                str(c['id']) for c in customers_qs
                if c['registration_date'] and c['registration_date'].year == y and c['registration_date'].month == m
            ]
            counts.append(len(m_cids))
            cids_list.append(m_cids)

        return {
            'months': month_labels,
            'new_customers': counts,
            'new_customer_ids': cids_list,
        }

    def _get_monthly_portafolio_coverage(self) -> dict[str, Any]:
        """
        calculates portfolio coverage percentage (active customers with purchases / cumulative registered portfolio) month by month.
        returns:
            {
                'months': ["Ago '25", "Sep '25", ...],
                'portfolio_coverage': [8.60, 8.33, ...],
                'active_customers': [25, 24, ...],
                'active_customer_ids': [['id1', ...], ...],
                'total_portfolio': [290, 291, ...],
            }
        """
        if not self.route_id:
            return {'months': [], 'portfolio_coverage': [], 'active_customers': [], 'active_customer_ids': [], 'total_portfolio': []}

        curr = self.date_start.replace(day=1)
        end_m = self.date_end.replace(day=1)
        eval_periods: list[tuple[int, int]] = []
        month_labels: list[str] = []

        month_names = {
            1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
            7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
        }

        while curr <= end_m:
            eval_periods.append((curr.year, curr.month))
            curr += relativedelta(months=1)

        has_multiple_years = len(set(y for y, _ in eval_periods)) > 1
        for y, m in eval_periods:
            m_label = f"{month_names[m]} '{str(y)[-2:]}" if has_multiple_years else month_names[m]
            month_labels.append(m_label)

        customers_qs = (
            self.customers_qs
            .filter(assignments__route_id=self.route_id)
            .values('id', 'registration_date')
            .distinct()
        )
        cust_reg = {str(c['id']): c['registration_date'] for c in customers_qs if c['registration_date']}

        tx_qs = (
            self.transactions_qs
            .filter(
                route_id=self.route_id,
                sale_date__gte=self.date_start,
                sale_date__lte=self.date_end,
                net_amount__gt=0
            )
            .values('sale_date__year', 'sale_date__month', 'customer_id')
            .distinct()
        )

        active_by_period = defaultdict(set)
        for tx in tx_qs:
            active_by_period[(tx['sale_date__year'], tx['sale_date__month'])].add(str(tx['customer_id']))

        coverage_list = []
        active_counts = []
        active_cids_list = []
        total_port_list = []

        for y, m in eval_periods:
            _, last_day = calendar.monthrange(y, m)
            end_month_date = date(y, m, last_day)

            portfolio_cids = [cid for cid, reg in cust_reg.items() if reg <= end_month_date]
            tot_port = len(portfolio_cids)
            total_port_list.append(tot_port)

            active_in_month = active_by_period.get((y, m), set())
            active_port = [cid for cid in portfolio_cids if cid in active_in_month]
            active_count = len(active_port)

            cov = round((active_count / tot_port * 100.0), 2) if tot_port > 0 else 0.0

            coverage_list.append(cov)
            active_counts.append(active_count)
            active_cids_list.append(active_port)

        return {
            'months': month_labels,
            'portfolio_coverage': coverage_list,
            'active_customers': active_counts,
            'active_customer_ids': active_cids_list,
            'total_portfolio': total_port_list,
        }

    def _get_sale_volume_by_customer(self) -> dict[str, dict[str, Any]]:
        """
        calculates monthly sales volume, active months, median, and momentum per customer
        respecting registration date.
        """
        if not self.route_id:
            return {}

        curr = self.date_start.replace(day=1)
        end_m = self.date_end.replace(day=1)
        timeline_months = []
        while curr <= end_m:
            timeline_months.append(curr.strftime('%Y-%m'))
            curr += relativedelta(months=1)

        month_names = {
            1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
            7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
        }
        last_month_key = timeline_months[-1] if timeline_months else ''
        if last_month_key:
            ly, lm = int(last_month_key[:4]), int(last_month_key[5:7])
            last_month_label = f"{month_names[lm]} '{str(ly)[-2:]}"
        else:
            last_month_label = 'Último mes'

        customers_qs = (
            self.customers_qs
            .filter(assignments__route_id=self.route_id)
            .values('id', 'name', 'registration_date')
            .distinct()
        )
        cust_map = {
            str(c['id']): {
                'name': (c['name'] or f"Cliente {c['id']}").title(),
                'reg_date': c['registration_date']
            }
            for c in customers_qs
        }

        tx_qs = (
            self.transactions_qs
            .filter(
                route_id=self.route_id,
                sale_date__gte=self.date_start,
                sale_date__lte=self.date_end,
                net_amount__gt=0
            )
            .values('sale_date', 'customer_id', 'net_amount')
        )

        monthly_sales = defaultdict(lambda: defaultdict(float))
        for t in tx_qs:
            m = t['sale_date'].strftime('%Y-%m')
            monthly_sales[m][str(t['customer_id'])] += float(t['net_amount'])

        result = {}
        for cid, info in cust_map.items():
            reg_date = info['reg_date']
            reg_month = reg_date.strftime('%Y-%m') if reg_date else timeline_months[0]
            valid_months = [m for m in timeline_months if m >= reg_month]
            n_months = len(valid_months)
            if n_months == 0:
                continue

            sales_data = [monthly_sales[m].get(cid, 0.0) for m in valid_months]
            total_sales = sum(sales_data)
            if total_sales <= 0:
                continue

            mean_sales = total_sales / n_months
            median_sales = statistics.median(sales_data) if sales_data else 0.0

            recent_months = valid_months[-3:] if n_months >= 3 else valid_months
            recent_mean = sum(monthly_sales[m].get(cid, 0.0) for m in recent_months) / len(recent_months)
            momentum = (recent_mean / mean_sales) if mean_sales > 0 else 0.0
            bias = ((mean_sales - median_sales) / abs(mean_sales)) if mean_sales > 0 else 0.0

            if len(recent_months) >= 3:
                m_start = month_names[int(recent_months[0][5:7])][:3]
                m_end = month_names[int(recent_months[-1][5:7])][:3]
                recent_label = f"Venta prom. trimestral ({m_start} - {m_end})"
            elif len(recent_months) == 2:
                m_start = month_names[int(recent_months[0][5:7])][:3]
                m_end = month_names[int(recent_months[-1][5:7])][:3]
                recent_label = f"Venta prom. bimestral ({m_start} - {m_end})"
            else:
                m_single = month_names[int(recent_months[0][5:7])][:3]
                recent_label = f"Venta {m_single}"

            months_with_purchases = sum(1 for s in sales_data if s > 0)
            last_month_sale = float(monthly_sales[last_month_key].get(cid, 0.0)) if last_month_key else 0.0

            result[cid] = {
                'customer_id': cid,
                'customer_name': info['name'],
                'registration_date': reg_date,
                'total_sales': total_sales,
                'mean_sales': mean_sales,
                'median_sales': median_sales,
                'recent_mean': recent_mean,
                'recent_label': recent_label,
                'momentum': momentum,
                'bias': bias,
                'valid_months': valid_months,
                'sales_data': sales_data,
                'sales_timeline': {m: float(monthly_sales[m].get(cid, 0.0)) for m in timeline_months},
                'months_with_purchases': months_with_purchases,
                'last_month_sale': last_month_sale,
                'last_month_label': last_month_label,
            }

        return result

    def _get_cv_by_customer(self) -> dict[str, dict[str, Any]]:
        """
        calculates coefficient of variation (CV = std_dev / mean) for each customer.
        correctly handles customers with 1 month of purchase history.
        """
        volume_data = self._get_sale_volume_by_customer()
        result = {}

        for cid, vinfo in volume_data.items():
            sales_data = vinfo['sales_data']
            n_months = len(sales_data)
            mean_sales = vinfo['mean_sales']

            if n_months <= 1:
                cv = 1.0 if vinfo['months_with_purchases'] == 1 else 0.0
                std_dev = 0.0
                is_single_month = True
            else:
                variance = sum((s - mean_sales)**2 for s in sales_data) / n_months
                std_dev = math.sqrt(variance)
                cv = std_dev / mean_sales if mean_sales > 0 else 0.0
                is_single_month = (vinfo['months_with_purchases'] == 1)

            result[cid] = {
                'customer_id': cid,
                'customer_name': vinfo['customer_name'],
                'registration_date': vinfo.get('registration_date'),
                'cv': cv,
                'std_dev': std_dev,
                'mean_sales': mean_sales,
                'median_sales': vinfo.get('median_sales', 0.0),
                'momentum': vinfo['momentum'],
                'bias': vinfo['bias'],
                'total_sales': vinfo['total_sales'],
                'is_single_month': is_single_month,
                'last_month_sale': vinfo['last_month_sale'],
                'last_month_label': vinfo['last_month_label'],
                'recent_mean': vinfo['recent_mean'],
                'recent_label': vinfo['recent_label'],
                'sales_timeline': vinfo.get('sales_timeline', {}),
            }

        return result

    def _get_volatility_and_volume(self) -> dict[str, Any]:
        """
        calculates scatter data, thresholds, and quadrant customer IDs for volatility and volume analysis.
        """
        cv_dict = self._get_cv_by_customer()
        if not cv_dict:
            return {
                'scatter_data': [],
                'thresholds': {'volume': 0.0, 'volatility': 0.0},
                'quadrants': {
                    'low_cv_high_vol': [],
                    'low_cv_low_vol': [],
                    'high_cv_high_vol': [],
                    'high_cv_low_vol': [],
                }
            }

        scatter_data = []
        vol_list = []
        cv_list = []

        for cid, info in cv_dict.items():
            mean_sales = info['mean_sales']
            cv = info['cv']
            momentum = info['momentum']
            bias = info['bias']
            name = info['customer_name']
            is_single = info['is_single_month']
            last_sale = info['last_month_sale']
            last_label = info['last_month_label']
            recent_mean = info['recent_mean']
            recent_label = info['recent_label']

            scatter_data.append([
                round(mean_sales, 2),
                round(cv, 4),
                round(momentum, 4),
                cid,
                name,
                is_single,
                round(bias, 4),
                round(last_sale, 2),
                last_label,
                round(recent_mean, 2),
                recent_label,
            ])
            vol_list.append(mean_sales)
            cv_list.append(cv)

        vol_list.sort()
        cv_list.sort()
        vol_threshold = vol_list[len(vol_list)//2] if vol_list else 0.0
        cv_threshold = cv_list[len(cv_list)//2] if cv_list else 0.0

        low_cv_high_vol_cids = [item[3] for item in scatter_data if item[0] >= vol_threshold and item[1] <= cv_threshold]
        low_cv_low_vol_cids = [item[3] for item in scatter_data if item[0] < vol_threshold and item[1] <= cv_threshold]
        high_cv_high_vol_cids = [item[3] for item in scatter_data if item[0] >= vol_threshold and item[1] > cv_threshold]
        high_cv_low_vol_cids = [item[3] for item in scatter_data if item[0] < vol_threshold and item[1] > cv_threshold]

        return {
            'scatter_data': scatter_data,
            'thresholds': {
                'volume': round(vol_threshold, 2),
                'volatility': round(cv_threshold, 4),
            },
            'quadrants': {
                'low_cv_high_vol': low_cv_high_vol_cids,
                'low_cv_low_vol': low_cv_low_vol_cids,
                'high_cv_high_vol': high_cv_high_vol_cids,
                'high_cv_low_vol': high_cv_low_vol_cids,
            }
        }

    def _get_momentum_and_bias(self) -> dict[str, Any]:
        """
        returns scatter data and universal thresholds for growth factor (momentum) vs bias analysis.
        """
        cv_dict = self._get_cv_by_customer()
        if not cv_dict:
            return {
                'scatter_data': [],
                'thresholds': {
                    'momentum': 1.0,
                    'bias': 0.0,
                },
                'quadrants': {
                    'stable_growing': [],
                    'erratic_growing': [],
                    'stable_decreasing': [],
                    'erratic_decreasing': [],
                }
            }

        scatter_data = []
        stable_growing_cids = []
        erratic_growing_cids = []
        stable_decreasing_cids = []
        erratic_decreasing_cids = []

        for cid, info in cv_dict.items():
            momentum = info['momentum']
            bias = info['bias']
            mean_sales = info['mean_sales']
            name = info['customer_name']
            is_single = info['is_single_month']
            last_sale = info['last_month_sale']
            last_label = info['last_month_label']
            recent_mean = info['recent_mean']
            recent_label = info['recent_label']
            cv = info['cv']

            #universal thresholds Momentum = 1.0, Bias = 0.0
            if momentum >= 1.0 and bias <= 0.0:
                stable_growing_cids.append(cid)
            elif momentum >= 1.0 and bias > 0.0:
                erratic_growing_cids.append(cid)
            elif momentum < 1.0 and bias <= 0.0:
                stable_decreasing_cids.append(cid)
            else:
                erratic_decreasing_cids.append(cid)

            scatter_data.append([
                round(momentum, 4),      # 0: X axis (Momentum)
                round(bias, 4),          # 1: Y axis (Bias)
                round(mean_sales, 2),    # 2: Mean sales ($)
                cid,                     # 3: Customer ID
                name,                    # 4: Name
                is_single,               # 5: Single month boolean
                round(last_sale, 2),     # 6: Last month sales ($)
                last_label,              # 7: Last month label
                round(recent_mean, 2),   # 8: Recent mean sales ($)
                recent_label,            # 9: Recent label
                round(cv, 4),            # 10: CV
            ])

        return {
            'scatter_data': scatter_data,
            'thresholds': {
                'momentum': 1.0,
                'bias': 0.0,
            },
            'quadrants': {
                'stable_growing': stable_growing_cids,
                'erratic_growing': erratic_growing_cids,
                'stable_decreasing': stable_decreasing_cids,
                'erratic_decreasing': erratic_decreasing_cids,
            }
        }

    def _init_categories(self) -> None:
        refs = Reference.objects.filter(context='categoria_cliente_monto')
        parsed = []
        for r in refs:
            try:
                parsed.append((r.key.strip().lower(), Decimal(str(r.value))))
            except (ValueError, TypeError):
                continue
        if parsed:
            self.categories = sorted(parsed, key=lambda x: x[1], reverse=True)
        else:
            self.categories = [
                ('diamante', Decimal('250000')),
                ('oro', Decimal('100000')),
                ('aa', Decimal('25000')),
                ('a', Decimal('3000')),
                ('c', Decimal('0')),
            ]

    def _calculate_category(self, sales: Decimal | float | None = None) -> str:
        """returns the category to which the customer belongs given a sales amount"""
        if not hasattr(self, 'categories'):
            self._init_categories()
        if sales is None:
            return 'c'
        sales_dec = Decimal(str(sales))
        for name, min_amount in self.categories:
            if sales_dec >= min_amount:
                return name
        return 'c'

    def _get_monthly_category_composition(self) -> dict[str, Any]:
        """
        Calculates customer category composition over time (month by month).
        Evaluates the rolling 3-month average purchase before each month,
        adjusting fairly for customer registration dates.
        Returns:
            - timeline_months: list of month keys ('2025-08', '2025-09', ...)
            - months_labels: list of month labels ('Ago 2025', 'Sep 2025', ...)
            - categories: ['c', 'a', 'aa', 'oro', 'diamante']
            - category_display: {'c': 'C', 'a': 'A', 'aa': 'AA', 'oro': 'Oro', 'diamante': 'Diamante'}
            - category_counts: {'c': [...], 'a': [...], 'aa': [...], 'oro': [...], 'diamante': [...]}
            - category_pcts: {'c': [...], 'a': [...], ...}
            - category_customer_ids: {'c': [[...], ...], ...}
            - total_portfolio: [total_m0, total_m1, ...]
        """
        if not hasattr(self, 'categories'):
            self._init_categories()

        ordered_cat_keys = [cat[0] for cat in reversed(self.categories)]
        cat_display_map = {
            'c': 'C',
            'a': 'A',
            'aa': 'AA',
            'oro': 'Oro',
            'diamante': 'Diamante'
        }

        if not self.route_id:
            return {
                'timeline_months': [],
                'months_labels': [],
                'categories': ordered_cat_keys,
                'category_display': [cat_display_map.get(k, k.upper()) for k in ordered_cat_keys],
                'category_counts': {k: [] for k in ordered_cat_keys},
                'category_pcts': {k: [] for k in ordered_cat_keys},
                'category_customer_ids': {k: [] for k in ordered_cat_keys},
                'total_portfolio': [],
            }

        curr = self.date_start.replace(day=1)
        end_m = self.date_end.replace(day=1)
        timeline_months = []
        months_labels = []

        month_abbr_es = {
            1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
            7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
        }

        while curr <= end_m:
            timeline_months.append(curr.strftime('%Y-%m'))
            months_labels.append(f"{month_abbr_es[curr.month]} '{str(curr.year)[2:]}")
            curr += relativedelta(months=1)

        if not timeline_months:
            return {
                'timeline_months': [],
                'months_labels': [],
                'categories': ordered_cat_keys,
                'category_display': [cat_display_map.get(k, k.upper()) for k in ordered_cat_keys],
                'category_counts': {k: [] for k in ordered_cat_keys},
                'category_pcts': {k: [] for k in ordered_cat_keys},
                'category_customer_ids': {k: [] for k in ordered_cat_keys},
                'total_portfolio': [],
            }

        customers_qs = (
            self.customers_qs
            .filter(assignments__route_id=self.route_id)
            .values('id', 'registration_date')
            .distinct()
        )
        customers_map = {c['id']: c['registration_date'] for c in customers_qs}

        extended_start = self.date_start.replace(day=1) - relativedelta(months=4)
        tx_qs = (
            self.transactions_qs
            .filter(
                route_id=self.route_id,
                sale_date__gte=extended_start,
                sale_date__lte=self.date_end,
                net_amount__gt=0
            )
            .values('sale_date', 'customer_id', 'net_amount')
        )

        monthly_customer_sales = defaultdict(lambda: defaultdict(float))
        for tx in tx_qs:
            m_key = tx['sale_date'].strftime('%Y-%m')
            monthly_customer_sales[m_key][tx['customer_id']] += float(tx['net_amount'])

        category_counts = {k: [] for k in ordered_cat_keys}
        category_pcts = {k: [] for k in ordered_cat_keys}
        category_customer_ids = {k: [] for k in ordered_cat_keys}
        total_portfolio_list = []

        for m_str in timeline_months:
            target_date = datetime.strptime(m_str, '%Y-%m').date()
            end_q = target_date.replace(day=1) - relativedelta(days=1)
            eval_months = [(end_q.replace(day=1) - relativedelta(months=i)).strftime('%Y-%m') for i in (2, 1, 0)]

            portfolio_cids = [
                cid for cid, reg_date in customers_map.items()
                if reg_date and reg_date <= (target_date.replace(day=1) + relativedelta(months=1) - relativedelta(days=1))
            ]
            total_customers_in_m = len(portfolio_cids)
            total_portfolio_list.append(total_customers_in_m)

            m_cat_counts = {k: 0 for k in ordered_cat_keys}
            m_cat_cids = {k: [] for k in ordered_cat_keys}

            for cid in portfolio_cids:
                reg_date = customers_map.get(cid)
                reg_month = reg_date.strftime('%Y-%m') if reg_date else '1970-01'

                active_eval_months = sum(1 for qm in eval_months if qm >= reg_month)
                if active_eval_months == 0:
                    avg_sales = 0.0
                else:
                    total_eval_sales = sum(monthly_customer_sales[qm].get(cid, 0.0) for qm in eval_months)
                    avg_sales = total_eval_sales / active_eval_months

                cat = self._calculate_category(avg_sales)
                if cat not in m_cat_counts:
                    cat = 'c'
                m_cat_counts[cat] += 1
                m_cat_cids[cat].append(cid)

            for k in ordered_cat_keys:
                cnt = m_cat_counts[k]
                category_counts[k].append(cnt)
                category_customer_ids[k].append(m_cat_cids[k])
                pct = round((cnt / total_customers_in_m * 100.0), 2) if total_customers_in_m > 0 else 0.0
                category_pcts[k].append(pct)

        return {
            'timeline_months': timeline_months,
            'months_labels': months_labels,
            'categories': ordered_cat_keys,
            'category_display': [cat_display_map.get(k, k.upper()) for k in ordered_cat_keys],
            'category_counts': category_counts,
            'category_pcts': category_pcts,
            'category_customer_ids': category_customer_ids,
            'total_portfolio': total_portfolio_list,
        }

    def _get_monetary_contrib_by_category(self) -> dict[str, Any]:
        """
        calculates monetary contribution by customer category month by month.
        evaluates the rolling 3-month average purchase before each month to determine category,
        then aggregates actual month sales ($ and %) for each category.
        returns:
            - timeline_months: list of month keys ('2025-08', '2025-09', ...)
            - months_labels: list of month labels ('Ago 2025', 'Sep 2025', ...)
            - categories: ['c', 'a', 'aa', 'oro', 'diamante']
            - category_display: ['C', 'A', 'AA', 'Oro', 'Diamante']
            - category_amounts: {'c': [...], 'a': [...], 'aa': [...], 'oro': [...], 'diamante': [...]}
            - category_percentages: {'c': [...], 'a': [...], 'aa': [...], 'oro': [...], 'diamante': [...]}
            - category_customer_ids: {'c': [[...], ...], ...}
            - total_sales_by_month: [total_m0, total_m1, ...]
        """
        if not hasattr(self, 'categories'):
            self._init_categories()

        ordered_cat_keys = [cat[0] for cat in reversed(self.categories)]
        cat_display_map = {
            'c': 'C',
            'a': 'A',
            'aa': 'AA',
            'oro': 'Oro',
            'diamante': 'Diamante'
        }

        if not self.route_id:
            return {
                'timeline_months': [],
                'months_labels': [],
                'categories': ordered_cat_keys,
                'category_display': [cat_display_map.get(k, k.upper()) for k in ordered_cat_keys],
                'category_amounts': {k: [] for k in ordered_cat_keys},
                'category_percentages': {k: [] for k in ordered_cat_keys},
                'category_customer_ids': {k: [] for k in ordered_cat_keys},
                'total_sales_by_month': [],
            }

        curr = self.date_start.replace(day=1)
        end_m = self.date_end.replace(day=1)
        timeline_months = []
        months_labels = []

        month_abbr_es = {
            1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
            7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
        }

        while curr <= end_m:
            timeline_months.append(curr.strftime('%Y-%m'))
            months_labels.append(f"{month_abbr_es[curr.month]} '{str(curr.year)[2:]}")
            curr += relativedelta(months=1)

        if not timeline_months:
            return {
                'timeline_months': [],
                'months_labels': [],
                'categories': ordered_cat_keys,
                'category_display': [cat_display_map.get(k, k.upper()) for k in ordered_cat_keys],
                'category_amounts': {k: [] for k in ordered_cat_keys},
                'category_percentages': {k: [] for k in ordered_cat_keys},
                'category_customer_ids': {k: [] for k in ordered_cat_keys},
                'total_sales_by_month': [],
            }

        customers_qs = (
            self.customers_qs
            .filter(assignments__route_id=self.route_id)
            .values('id', 'registration_date')
            .distinct()
        )
        customers_map = {c['id']: c['registration_date'] for c in customers_qs}

        extended_start = self.date_start.replace(day=1) - relativedelta(months=4)
        tx_qs = (
            self.transactions_qs
            .filter(
                route_id=self.route_id,
                sale_date__gte=extended_start,
                sale_date__lte=self.date_end,
                net_amount__gt=0
            )
            .values('sale_date', 'customer_id', 'net_amount')
        )

        monthly_customer_sales = defaultdict(lambda: defaultdict(float))
        for tx in tx_qs:
            m_key = tx['sale_date'].strftime('%Y-%m')
            monthly_customer_sales[m_key][tx['customer_id']] += float(tx['net_amount'])

        category_amounts = {k: [] for k in ordered_cat_keys}
        category_percentages = {k: [] for k in ordered_cat_keys}
        category_customer_ids = {k: [] for k in ordered_cat_keys}
        total_sales_by_month = []

        for m_str in timeline_months:
            target_date = datetime.strptime(m_str, '%Y-%m').date()
            end_q = target_date.replace(day=1) - relativedelta(days=1)
            eval_months = [(end_q.replace(day=1) - relativedelta(months=i)).strftime('%Y-%m') for i in (2, 1, 0)]

            portfolio_cids = [
                cid for cid, reg_date in customers_map.items()
                if reg_date and reg_date <= (target_date.replace(day=1) + relativedelta(months=1) - relativedelta(days=1))
            ]

            m_amounts = {k: 0.0 for k in ordered_cat_keys}
            m_cids = {k: [] for k in ordered_cat_keys}
            total_m_sale = 0.0

            for cid in portfolio_cids:
                reg_date = customers_map.get(cid)
                reg_month = reg_date.strftime('%Y-%m') if reg_date else '1970-01'

                active_eval_months = sum(1 for qm in eval_months if qm >= reg_month)
                if active_eval_months == 0:
                    avg_sales = 0.0
                else:
                    total_eval_sales = sum(monthly_customer_sales[qm].get(cid, 0.0) for qm in eval_months)
                    avg_sales = total_eval_sales / active_eval_months

                cat = self._calculate_category(avg_sales)
                if cat not in m_amounts:
                    cat = 'c'

                actual_sale = float(monthly_customer_sales[m_str].get(cid, 0.0))
                m_amounts[cat] += actual_sale
                total_m_sale += actual_sale
                if actual_sale > 0:
                    m_cids[cat].append(cid)

            total_sales_by_month.append(round(total_m_sale, 2))

            for k in ordered_cat_keys:
                amt = m_amounts[k]
                category_amounts[k].append(round(amt, 2))
                category_customer_ids[k].append(m_cids[k])
                pct = round((amt / total_m_sale * 100.0), 2) if total_m_sale > 0 else 0.0
                category_percentages[k].append(pct)

        return {
            'timeline_months': timeline_months,
            'months_labels': months_labels,
            'categories': ordered_cat_keys,
            'category_display': [cat_display_map.get(k, k.upper()) for k in ordered_cat_keys],
            'category_amounts': category_amounts,
            'category_percentages': category_percentages,
            'category_customer_ids': category_customer_ids,
            'total_sales_by_month': total_sales_by_month,
        }
    


@dataclass
class CommercialRiskStats:
    commercial_risk_service: CommercialRiskService

    def _get_target_customers_registered(self) -> list[str]:
        """Returns customer IDs assigned to this route registered on or before end_q_date"""
        service = self.commercial_risk_service
        if not service.route_id:
            return []

        target_qs = (
            service.customers_qs
            .filter(
                assignments__route_id=service.route_id,
                registration_date__lte=service.end_q_date
            )
            .filter(
                Q(assignments__end_date__isnull=True) | Q(assignments__end_date__gte=service.today)
            )
            .values_list('id', flat=True)
            .distinct()
        )
        return [str(cid) for cid in target_qs]

    def _get_quarter_customer_sales(self) -> dict[str, Decimal]:
        """Returns {customer_id: total_net_amount} for assigned customers in the closed quarter with net_amount > 0"""
        service = self.commercial_risk_service
        if not service.route_id:
            return {}

        target_cids = self._get_target_customers_registered()
        if not target_cids:
            return {}

        sales_qs = (
            service.transactions_qs
            .filter(
                route_id=service.route_id,
                customer_id__in=target_cids,
                sale_date__gte=service.start_q_date,
                sale_date__lte=service.end_q_date,
                net_amount__gt=0
            )
            .values('customer_id')
            .annotate(total=Sum('net_amount'))
        )
        return {str(row['customer_id']): (row['total'] or Decimal('0.00')) for row in sales_qs}

    def _get_gini(self) -> Decimal:
        """
        returns the gini index (0.0 to 100.0) from last closed quarter among purchasing clients.
        Evaluates concentration and dependency on few top clients.
        """
        customer_sales = self._get_quarter_customer_sales()
        sales_list = sorted([float(amt) for amt in customer_sales.values() if amt > 0])
        n = len(sales_list)
        cum_sales = sum(sales_list)

        if n <= 0 or cum_sales <= 0:
            return Decimal('0.00')

        sum_iy = sum((i + 1) * y for i, y in enumerate(sales_list))
        gini_coeff = (2.0 * sum_iy / (n * cum_sales)) - ((n + 1.0) / n)
        gini_pct = max(min(gini_coeff * 100.0, 100.0), 0.0)
        return Decimal(str(round(gini_pct, 2)))

    def _get_portafolio_coverage(self) -> dict[str, Decimal]:
        """
        returns the portfolio coverage from last closed quarter.
        coverage = (customers_with_consumption / registered_customers) * 100
        """
        registered_cids = set(self._get_target_customers_registered())
        total_registered = len(registered_cids)

        customer_sales = self._get_quarter_customer_sales()
        active_customers = sum(1 for cid, amt in customer_sales.items() if amt > 0 and str(cid) in registered_cids)

        if total_registered > 0:
            coverage_pct = (Decimal(active_customers) / Decimal(total_registered)) * Decimal('100.00')
            coverage_pct = min(Decimal('100.00'), max(Decimal('0.00'), coverage_pct))
        else:
            coverage_pct = Decimal('0.00')

        return {
            'registered_customers': Decimal(str(total_registered)),
            'customers_with_consumption': Decimal(str(active_customers)),
            'coverage': Decimal(str(round(coverage_pct, 2))),
        }

    def _get_momentum(self) -> dict[str, Decimal]:
        """
        returns the momentum (growth factor) from last closed quarter compared to historical average.
        momentum = ((quarter_avg / history_avg) - 1) * 100
        """
        service = self.commercial_risk_service
        if not service.route_id:
            return {
                'momentum': Decimal('0.00'),
                'history': Decimal('0.00'),
                'current': Decimal('0.00'),
            }

        tx_historical = (
            service.transactions_qs
            .filter(
                route_id=service.route_id,
                sale_date__lte=service.end_q_date
            )
            .values('sale_date', 'net_amount')
        )

        monthly_sales = defaultdict(Decimal)
        for tx in tx_historical:
            m_key = tx['sale_date'].strftime('%Y-%m')
            monthly_sales[m_key] += (tx['net_amount'] or Decimal('0.00'))

        total_months = len(monthly_sales)
        history_avg = (sum(monthly_sales.values()) / total_months) if total_months > 0 else Decimal('0.00')

        q_months = [
            (service.start_q_date + relativedelta(months=i)).strftime('%Y-%m')
            for i in range(3)
        ]
        quarter_total = sum(monthly_sales.get(m, Decimal('0.00')) for m in q_months)
        quarter_avg = quarter_total / Decimal('3.0')

        if history_avg > Decimal('0.00'):
            momentum_pct = ((quarter_avg / history_avg) - Decimal('1.00')) * Decimal('100.00')
        else:
            momentum_pct = Decimal('0.00')

        return {
            'momentum': Decimal(str(round(momentum_pct, 2))),
            'history': Decimal(str(round(history_avg, 2))),
            'current': Decimal(str(round(quarter_avg, 2))),
        }

    def _get_commercial_risk_index(self, gini: Decimal, coverage: Decimal) -> Decimal:
        """
        calculates commercial risk index as a weighted composite of:
        - structural concentration (Gini index: higher = more risk)
        - portfolio inactivity / non-coverage ((100 - Coverage): higher = more risk)
        formula: 0.5 * (100 - coverage) + 0.5 * gini
        """
        inactivity = Decimal('100.00') - coverage
        risk = (Decimal('0.5') * inactivity) + (Decimal('0.5') * gini)
        bounded_risk = max(min(risk, Decimal('100.00')), Decimal('0.00'))
        return Decimal(str(round(bounded_risk, 2)))

    def stats(self) -> dict[str, Any]:
        gini = self._get_gini()
        portafolio_data = self._get_portafolio_coverage()
        coverage = portafolio_data['coverage']
        momentum_data = self._get_momentum()
        momentum = momentum_data['momentum']
        risk_index = self._get_commercial_risk_index(gini=gini, coverage=coverage)

        churn_data = self.commercial_risk_service._get_customer_churn()
        won_customers = sum(churn_data.get('won', []))
        lost_customers = sum(churn_data.get('lost', []))

        service = self.commercial_risk_service
        new_customers_count = (
            service.customers_qs
            .filter(
                assignments__route_id=service.route_id,
                registration_date__gte=service.date_start,
                registration_date__lte=service.date_end
            )
            .distinct()
            .count()
        ) if service.route_id else 0

        cov_data = self.commercial_risk_service._get_monthly_portafolio_coverage()
        cov_list = cov_data.get('portfolio_coverage', [])
        if cov_list:
            history_cov_avg = sum(cov_list) / len(cov_list)
            last_month_cov = cov_list[-1]
            if history_cov_avg > 0:
                coverage_momentum = Decimal(str(round(((last_month_cov / history_cov_avg) - 1.0) * 100.0, 2)))
            else:
                coverage_momentum = Decimal('0.00')
        else:
            coverage_momentum = Decimal('0.00')

        vol_data = self.commercial_risk_service._get_volatility_and_volume()
        quadrants = vol_data.get('quadrants', {})
        low_cv_high_vol = quadrants.get('low_cv_high_vol', [])
        low_cv_low_vol = quadrants.get('low_cv_low_vol', [])
        high_cv_high_vol = quadrants.get('high_cv_high_vol', [])
        high_cv_low_vol = quadrants.get('high_cv_low_vol', [])

        return {
            'gini_index': gini,
            'portfolio_coverage': coverage,
            'registered_customers': portafolio_data['registered_customers'],
            'customers_with_consumption': portafolio_data['customers_with_consumption'],
            'sales_momentum': momentum,
            'history_avg': momentum_data['history'],
            'quarter_avg': momentum_data['current'],
            'commercial_risk_index': risk_index,
            'won_customers': won_customers,
            'lost_customers': lost_customers,
            'new_customers': new_customers_count,
            'coverage_momentum': coverage_momentum,
            'low_cv_high_vol': len(low_cv_high_vol),
            'low_cv_low_vol': len(low_cv_low_vol),
            'high_cv_high_vol': len(high_cv_high_vol),
            'high_cv_low_vol': len(high_cv_low_vol),
            'low_cv_high_vol_ids': low_cv_high_vol,
            'low_cv_low_vol_ids': low_cv_low_vol,
            'high_cv_high_vol_ids': high_cv_high_vol,
            'high_cv_low_vol_ids': high_cv_low_vol,
            'volatility_thresholds': vol_data.get('thresholds', {}),
            'start_q_date': self.commercial_risk_service.start_q_date,
            'end_q_date': self.commercial_risk_service.end_q_date,
        }


@dataclass
class CommercialRiskExports:
    commercial_risk_service: CommercialRiskService

    def export_commercial_risk_report(self) -> io.BytesIO:
        service = self.commercial_risk_service
        stats = service.stats()
        cv_data = service._get_cv_by_customer()
        vol_data = service._get_volatility_and_volume()
        churn_data = service._get_customer_churn()
        cat_comp_data = service._get_monthly_category_composition()
        mon_contrib_data = service._get_monetary_contrib_by_category()

        wb = openpyxl.Workbook()

        #styles
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        section_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
        title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
        subtitle_font = Font(name="Calibri", size=9, italic=True, color="64748B")
        data_font = Font(name="Calibri", size=10)
        bold_data_font = Font(name="Calibri", size=10, bold=True)

        thin_border_side = Side(style='thin', color='CBD5E1')
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        currency_format = '"$"#,##0.00'
        pct_format = '0.00%'
        int_format = '#,##0'
        dec_format = '0.0000'

        #route and dates
        route_str = f"Ruta {service.route_id}" if service.route_id else "General"
        if hasattr(service.route, 'name') and service.route.name:
            route_str += f" - {service.route.name.title()}"

        d_start_str = service.date_start.strftime('%Y-%m-%d') if service.date_start else ''
        d_end_str = service.date_end.strftime('%Y-%m-%d') if service.date_end else ''
        start_q_str = service.start_q_date.strftime('%Y-%m-%d') if service.start_q_date else ''
        end_q_str = service.end_q_date.strftime('%Y-%m-%d') if service.end_q_date else ''
        now_str = timezone.localtime().strftime('%Y-%m-%d %H:%M')

        # sheet 1 gen summary
        ws_summary = wb.active
        ws_summary.title = "Resumen General"
        ws_summary.views.sheetView[0].showGridLines = True

        ws_summary.cell(row=1, column=1, value="REPORTE DE RIESGO COMERCIAL - RESUMEN EJECUTIVO").font = title_font
        ws_summary.cell(
            row=2, column=1,
            value=f"Ruta: {route_str} | Generado el: {now_str} | Periodo: {d_start_str} al {d_end_str} | Trimestre Cerrado Evaluado: {start_q_str} al {end_q_str}"
        ).font = subtitle_font

        #section 1 closed q
        ws_summary.cell(row=4, column=1, value=f"1. Indicadores del Último Trimestre Cerrado ({start_q_str} al {end_q_str})").font = section_font
        bans_headers = ["Indicador", "Valor", "Referencia / Base"]
        for col_num, h_text in enumerate(bans_headers, 1):
            cell = ws_summary.cell(row=5, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_num > 1 else "left")
            cell.border = cell_border

        bans_rows = [
            ("Índice de Riesgo Comercial (IRC)", float(stats.get('commercial_risk_index') or 0) / 100.0, "50% Concentración Gini + 50% Inactividad"),
            ("Índice de Concentración (Gini)", float(stats.get('gini_index') or 0) / 100.0, "Evaluado en clientes con compra"),
            ("Alcance de Cartera (Cobertura)", float(stats.get('portfolio_coverage') or 0) / 100.0, f"{stats.get('customers_with_consumption', 0)} con consumo / {stats.get('registered_customers', 0)} registrados"),
            ("Factor de Crecimiento (Momentum)", float(stats.get('sales_momentum') or 0) / 100.0, f"Prom. Trimestre: ${float(stats.get('quarter_avg') or 0):,.2f} vs Prom. Histórico: ${float(stats.get('history_avg') or 0):,.2f}"),
        ]

        for r_idx, (lbl, val, ref) in enumerate(bans_rows, 6):
            c_l = ws_summary.cell(row=r_idx, column=1, value=lbl)
            c_l.font = data_font
            c_l.border = cell_border

            c_v = ws_summary.cell(row=r_idx, column=2, value=val)
            c_v.font = bold_data_font
            c_v.number_format = pct_format
            c_v.alignment = Alignment(horizontal="right")
            c_v.border = cell_border

            c_r = ws_summary.cell(row=r_idx, column=3, value=ref)
            c_r.font = data_font
            c_r.border = cell_border

        #section 2 churn 
        start_row_mov = 12
        ws_summary.cell(row=start_row_mov, column=1, value=f"2. Movimiento de Cartera en el Periodo ({d_start_str} al {d_end_str})").font = section_font
        mov_headers = ["Concepto", "Total Clientes", "Observación"]
        for col_num, h_text in enumerate(mov_headers, 1):
            cell = ws_summary.cell(row=start_row_mov + 1, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_num > 1 else "left")
            cell.border = cell_border

        mov_rows = [
            ("Clientes Ganados / Reactivados", int(stats.get('won_customers') or 0), "Clientes que compraron tras haber estado inactivos"),
            ("Clientes Perdidos / Inactivos", int(stats.get('lost_customers') or 0), "Clientes activos que dejaron de comprar"),
            ("Clientes Nuevos Registrados", int(stats.get('new_customers') or 0), "Nuevas altas en la ruta durante el periodo"),
            ("Tendencia de Cobertura (Momentum Cobertura)", float(stats.get('coverage_momentum') or 0) / 100.0, "Variación de alcance del último mes vs promedio"),
        ]

        for r_idx, (lbl, val, obs) in enumerate(mov_rows, start_row_mov + 2):
            c_l = ws_summary.cell(row=r_idx, column=1, value=lbl)
            c_l.font = data_font
            c_l.border = cell_border

            c_v = ws_summary.cell(row=r_idx, column=2, value=val)
            c_v.font = bold_data_font
            if isinstance(val, float):
                c_v.number_format = pct_format
            else:
                c_v.number_format = int_format
            c_v.alignment = Alignment(horizontal="right")
            c_v.border = cell_border

            c_o = ws_summary.cell(row=r_idx, column=3, value=obs)
            c_o.font = data_font
            c_o.border = cell_border

        #section 3 volatility and volume
        start_row_dist = 18
        ws_summary.cell(row=start_row_dist, column=1, value="3. Distribución por Volatilidad y Volumen de Consumo").font = section_font
        dist_headers = ["Cuadrante de Riesgo", "Total Clientes", "Estrategia Recomendada"]
        for col_num, h_text in enumerate(dist_headers, 1):
            cell = ws_summary.cell(row=start_row_dist + 1, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_num > 1 else "left")
            cell.border = cell_border

        dist_rows = [
            ("Baja Volatilidad / Consumo Alto", int(stats.get('low_cv_high_vol') or 0), "Base de la ruta: Clientes estratégicos y conveniables"),
            ("Baja Volatilidad / Consumo Bajo", int(stats.get('low_cv_low_vol') or 0), "Estables de bajo consumo: Acercamiento para incrementar ticket"),
            ("Alta Volatilidad / Consumo Alto", int(stats.get('high_cv_high_vol') or 0), "Grandes erráticos: Estabilizar consumo y planificar pedidos"),
            ("Alta Volatilidad / Consumo Bajo", int(stats.get('high_cv_low_vol') or 0), "Ruido comercial: Compras esporádicas y bajo valor"),
        ]

        for r_idx, (lbl, val, est) in enumerate(dist_rows, start_row_dist + 2):
            c_l = ws_summary.cell(row=r_idx, column=1, value=lbl)
            c_l.font = data_font
            c_l.border = cell_border

            c_v = ws_summary.cell(row=r_idx, column=2, value=val)
            c_v.font = bold_data_font
            c_v.number_format = int_format
            c_v.alignment = Alignment(horizontal="right")
            c_v.border = cell_border

            c_e = ws_summary.cell(row=r_idx, column=3, value=est)
            c_e.font = data_font
            c_e.border = cell_border

        #sheet 2 metrics by customer
        ws_customers = wb.create_sheet("Métricas por Cliente")
        ws_customers.views.sheetView[0].showGridLines = True

        curr_t = service.date_start.replace(day=1)
        end_t = service.date_end.replace(day=1)
        timeline_months = []
        while curr_t <= end_t:
            timeline_months.append(curr_t.strftime('%Y-%m'))
            curr_t += relativedelta(months=1)

        month_headers = [f"Venta {m}" for m in timeline_months]
        ctm_headers = [
            "ID Cliente", "Cliente", "Fecha Registro",
            "Cuadrante Volatilidad", "Cuadrante Sesgo / Crecimiento",
            "Venta Promedio Mensual", "Mediana Mensual", "Venta Trimestre Reciente",
            "Factor Crecimiento", "Crecimiento (%)", "Índice de Sesgo", "Volatilidad (CV)"
        ] + month_headers

        for col_num, h_text in enumerate(ctm_headers, 1):
            cell = ws_customers.cell(row=1, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_num in (1, 3, 4, 5) else ("right" if col_num >= 6 else "left"))
            cell.border = cell_border

        vol_threshold = float(vol_data.get('thresholds', {}).get('volume', 0.0))
        cv_threshold = float(vol_data.get('thresholds', {}).get('volatility', 0.0))

        for r_idx, (cid, info) in enumerate(cv_data.items(), 2):
            mean_s = float(info['mean_sales'])
            median_s = float(info['median_sales'])
            recent_m = float(info['recent_mean'])
            mom = float(info['momentum'])
            bias = float(info['bias'])
            cv = float(info['cv'])
            name = info['customer_name'].title()
            reg_d = info['registration_date'].strftime('%Y-%m-%d') if info['registration_date'] else 'N/A'

            #quadrant volatility
            if mean_s >= vol_threshold and cv <= cv_threshold:
                q_vol = "Baja Vol. / Alto Consumo"
            elif mean_s < vol_threshold and cv <= cv_threshold:
                q_vol = "Baja Vol. / Bajo Consumo"
            elif mean_s >= vol_threshold and cv > cv_threshold:
                q_vol = "Alta Vol. / Alto Consumo"
            else:
                q_vol = "Alta Vol. / Bajo Consumo"

            # quadrant bias
            if mom >= 1.0 and bias <= 0.0:
                q_bias = "Estable y Creciendo"
            elif mom >= 1.0 and bias > 0.0:
                q_bias = "Errático Creciendo"
            elif mom < 1.0 and bias <= 0.0:
                q_bias = "Estable Decreciendo"
            else:
                q_bias = "Errático Decreciendo"

            row_vals = [
                cid, name, reg_d, q_vol, q_bias,
                mean_s, median_s, recent_m, mom, (mom - 1.0), bias, cv
            ]

            sales_timeline = info.get('sales_timeline', {})
            for m in timeline_months:
                row_vals.append(float(sales_timeline.get(m, 0.0)))

            for col_idx, val in enumerate(row_vals, 1):
                c = ws_customers.cell(row=r_idx, column=col_idx, value=val)
                c.font = data_font
                c.border = cell_border

                if col_idx in (1, 3, 4, 5):
                    c.alignment = Alignment(horizontal="center")
                elif col_idx in (6, 7, 8):
                    c.number_format = currency_format
                    c.alignment = Alignment(horizontal="right")
                elif col_idx in (9, 11, 12):
                    c.number_format = dec_format
                    c.alignment = Alignment(horizontal="right")
                elif col_idx == 10:
                    c.number_format = pct_format
                    c.alignment = Alignment(horizontal="right")
                elif col_idx > 12:
                    c.number_format = currency_format
                    c.alignment = Alignment(horizontal="right")
                else:
                    c.alignment = Alignment(horizontal="left")

        #sheet 3 won and lost customers
        ws_churn = wb.create_sheet("Ganados y Perdidos")
        ws_churn.views.sheetView[0].showGridLines = True

        churn_headers = ["Periodo", "ID Cliente", "Cliente", "Movimiento", "Venta Mes Previo", "Venta Mes Actual"]
        for col_num, h_text in enumerate(churn_headers, 1):
            cell = ws_churn.cell(row=1, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_num in (1, 2, 4) else ("right" if col_num >= 5 else "left"))
            cell.border = cell_border

        churn_months = churn_data.get('months', [])
        won_cids_by_m = churn_data.get('won_customer_ids', [])
        lost_cids_by_m = churn_data.get('lost_customer_ids', [])

        curr_churn_row = 2
        for m_idx, m_label in enumerate(churn_months):
            if m_idx == 0:
                continue

            m_curr = timeline_months[m_idx] if m_idx < len(timeline_months) else m_label
            m_prev = timeline_months[m_idx - 1] if m_idx - 1 < len(timeline_months) else ''

            won_list = won_cids_by_m[m_idx] if m_idx < len(won_cids_by_m) else []
            for cid in won_list:
                info = cv_data.get(cid, {})
                cname = info.get('customer_name', f'Cliente #{cid}').title()
                s_prev = float(info.get('sales_timeline', {}).get(m_prev, 0.0))
                s_curr = float(info.get('sales_timeline', {}).get(m_curr, 0.0))

                vals = [m_label, cid, cname, "Ganado / Reactivado", s_prev, s_curr]
                for c_idx, v in enumerate(vals, 1):
                    c = ws_churn.cell(row=curr_churn_row, column=c_idx, value=v)
                    c.font = data_font
                    c.border = cell_border
                    if c_idx in (1, 2, 4):
                        c.alignment = Alignment(horizontal="center")
                    elif c_idx >= 5:
                        c.number_format = currency_format
                        c.alignment = Alignment(horizontal="right")
                    else:
                        c.alignment = Alignment(horizontal="left")
                curr_churn_row += 1

            lost_list = lost_cids_by_m[m_idx] if m_idx < len(lost_cids_by_m) else []
            for cid in lost_list:
                info = cv_data.get(cid, {})
                cname = info.get('customer_name', f'Cliente #{cid}').title()
                s_prev = float(info.get('sales_timeline', {}).get(m_prev, 0.0))
                s_curr = float(info.get('sales_timeline', {}).get(m_curr, 0.0))

                vals = [m_label, cid, cname, "Perdido / Inactivo", s_prev, s_curr]
                for c_idx, v in enumerate(vals, 1):
                    c = ws_churn.cell(row=curr_churn_row, column=c_idx, value=v)
                    c.font = data_font
                    c.border = cell_border
                    if c_idx in (1, 2, 4):
                        c.alignment = Alignment(horizontal="center")
                    elif c_idx >= 5:
                        c.number_format = currency_format
                        c.alignment = Alignment(horizontal="right")
                    else:
                        c.alignment = Alignment(horizontal="left")
                curr_churn_row += 1

        #sheet 4 category downgrade
        ws_downgrade = wb.create_sheet("Baja de Categoría")
        ws_downgrade.views.sheetView[0].showGridLines = True

        downgrade_headers = [
            "Periodo", "ID Cliente", "Cliente", "Categoría Previa",
            "Promedio Evaluado Previo", "Categoría Actual", "Promedio Evaluado Actual", "Diferencia"
        ]
        for col_num, h_text in enumerate(downgrade_headers, 1):
            cell = ws_downgrade.cell(row=1, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_num in (1, 2, 4, 6) else ("right" if col_num in (5, 7, 8) else "left"))
            cell.border = cell_border

        cat_order_rank = {'diamante': 5, 'oro': 4, 'aa': 3, 'a': 2, 'c': 1}
        cat_display_names = {'diamante': 'Diamante', 'oro': 'Oro', 'aa': 'AA', 'a': 'A', 'c': 'C'}

        downgrade_row = 2
        for m_idx, m_str in enumerate(timeline_months):
            if m_idx == 0:
                continue

            prev_m_str = timeline_months[m_idx - 1]
            target_curr = datetime.strptime(m_str, '%Y-%m').date()
            target_prev = datetime.strptime(prev_m_str, '%Y-%m').date()

            end_q_curr = target_curr.replace(day=1) - relativedelta(days=1)
            eval_curr = [(end_q_curr.replace(day=1) - relativedelta(months=i)).strftime('%Y-%m') for i in (2, 1, 0)]

            end_q_prev = target_prev.replace(day=1) - relativedelta(days=1)
            eval_prev = [(end_q_prev.replace(day=1) - relativedelta(months=i)).strftime('%Y-%m') for i in (2, 1, 0)]

            for cid, info in cv_data.items():
                reg_d = info.get('registration_date')
                if reg_d and reg_d > (target_curr.replace(day=1) + relativedelta(months=1) - relativedelta(days=1)):
                    continue

                reg_month = reg_d.strftime('%Y-%m') if reg_d else '1970-01'
                sales_tl = info.get('sales_timeline', {})

                active_prev = sum(1 for m in eval_prev if m >= reg_month)
                avg_prev = (sum(float(sales_tl.get(m, 0.0)) for m in eval_prev) / active_prev) if active_prev > 0 else 0.0
                cat_prev = service._calculate_category(avg_prev)

                active_curr = sum(1 for m in eval_curr if m >= reg_month)
                avg_curr = (sum(float(sales_tl.get(m, 0.0)) for m in eval_curr) / active_curr) if active_curr > 0 else 0.0
                cat_curr = service._calculate_category(avg_curr)

                if cat_order_rank.get(cat_curr, 1) < cat_order_rank.get(cat_prev, 1):
                    cname = info.get('customer_name', f'Cliente #{cid}').title()
                    diff = avg_curr - avg_prev
                    vals = [
                        m_str, cid, cname,
                        cat_display_names.get(cat_prev, cat_prev.upper()), avg_prev,
                        cat_display_names.get(cat_curr, cat_curr.upper()), avg_curr, diff
                    ]
                    for c_idx, v in enumerate(vals, 1):
                        c = ws_downgrade.cell(row=downgrade_row, column=c_idx, value=v)
                        c.font = data_font
                        c.border = cell_border
                        if c_idx in (1, 2, 4, 6):
                            c.alignment = Alignment(horizontal="center")
                        elif c_idx in (5, 7, 8):
                            c.number_format = currency_format
                            c.alignment = Alignment(horizontal="right")
                        else:
                            c.alignment = Alignment(horizontal="left")
                    downgrade_row += 1

        #sheet 5 category evolution
        ws_categories = wb.create_sheet("Evolución de Categorías")
        ws_categories.views.sheetView[0].showGridLines = True

        cat_sheet_headers = [
            "Periodo", "Categoría", "Clientes en Categoría",
            "% Participación Cartera", "Venta Total Categoría", "% Contribución Monetaria"
        ]
        for col_num, h_text in enumerate(cat_sheet_headers, 1):
            cell = ws_categories.cell(row=1, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_num in (1, 2) else "right")
            cell.border = cell_border

        categories_list = cat_comp_data.get('categories', ['diamante', 'oro', 'aa', 'a', 'c'])
        cat_disp_list = cat_comp_data.get('category_display', ['Diamante', 'Oro', 'AA', 'A', 'C'])
        timeline_m_labels = cat_comp_data.get('months_labels', [])
        cat_counts = cat_comp_data.get('category_counts', {})
        cat_pcts = cat_comp_data.get('category_pcts', {})
        cat_amounts = mon_contrib_data.get('category_amounts', {})
        cat_mon_pcts = mon_contrib_data.get('category_percentages', {})

        cat_row_idx = 2
        for m_idx, m_lbl in enumerate(timeline_m_labels):
            for c_idx, cat_key in enumerate(categories_list):
                c_disp = cat_disp_list[c_idx] if c_idx < len(cat_disp_list) else cat_key.upper()
                c_cnt = cat_counts.get(cat_key, [])[m_idx] if m_idx < len(cat_counts.get(cat_key, [])) else 0
                c_pct = (cat_pcts.get(cat_key, [])[m_idx] / 100.0) if m_idx < len(cat_pcts.get(cat_key, [])) else 0.0
                c_amt = cat_amounts.get(cat_key, [])[m_idx] if m_idx < len(cat_amounts.get(cat_key, [])) else 0.0
                c_mon_pct = (cat_mon_pcts.get(cat_key, [])[m_idx] / 100.0) if m_idx < len(cat_mon_pcts.get(cat_key, [])) else 0.0

                vals = [m_lbl, c_disp, c_cnt, c_pct, c_amt, c_mon_pct]
                for col_i, v in enumerate(vals, 1):
                    c = ws_categories.cell(row=cat_row_idx, column=col_i, value=v)
                    c.font = data_font
                    c.border = cell_border
                    if col_i in (1, 2):
                        c.alignment = Alignment(horizontal="center")
                    elif col_i == 3:
                        c.number_format = int_format
                        c.alignment = Alignment(horizontal="right")
                    elif col_i in (4, 6):
                        c.number_format = pct_format
                        c.alignment = Alignment(horizontal="right")
                    elif col_i == 5:
                        c.number_format = currency_format
                        c.alignment = Alignment(horizontal="right")
                cat_row_idx += 1

        # autofit columns
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if '\n' in val_str:
                        val_str = max(val_str.split('\n'), key=len)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer