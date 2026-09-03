import io
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any
from dateutil.relativedelta import relativedelta
from django.db.models import QuerySet, Sum, Q
from django.utils import timezone
from collections import defaultdict
from django.db.models.functions import TruncMonth

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.core.models import Reference
from apps.customers.models import CustomerAssignment
from apps.products.models import ProductClass

@dataclass
class CustomerKpisService:
    user: Any
    customers_qs: QuerySet
    transactions_qs: QuerySet
    ars_qs: QuerySet
    date_start: date | str | None = None
    date_end: date | str | None = None
    cleaned_data: dict[str, Any] | None = None
    # calcs and attrs
    order_by: str = field(default='net_amount', init=False)
    today: date = field(init=False)
    current_year: int = field(init=False)
    previous_year: int = field(init=False)
    first_day_q: date = field(init=False)
    last_day_q: date = field(init=False)
    categories: list[tuple[str, Decimal]] = field(default_factory=list, init=False)
    frequency_categories: list[tuple[str, int]] = field(default_factory=list, init=False)
    relevant_classes: list[tuple[str, str]] = field(default_factory=list, init=False)
    class_names: dict[str, str] = field(default_factory=dict, init=False)
    _cached_stats: dict[str, Any] | None = field(default=None, init=False)
    _cached_target_customers: list[Any] | None = field(default=None, init=False)

    def __post_init__(self):
        self._init_dates()
        self._init_config()

    def _init_dates(self) -> None:
        self.today = timezone.localdate()
        self.current_year = self.today.year
        self.previous_year = self.current_year - 1

        # Last full quarter
        first_day_current_month = self.today.replace(day=1)
        self.last_day_q = first_day_current_month - relativedelta(days=1)
        self.first_day_q = self.last_day_q.replace(day=1) - relativedelta(months=2)

        # Parse or defaults
        d_start = self._parse_date(self.date_start) or self.first_day_q
        d_end = self._parse_date(self.date_end) or self.last_day_q

        # Ensure date_start <= date_end
        if d_start > d_end:
            d_start, d_end = d_end, d_start

        self.date_start = d_start
        self.date_end = d_end

    def _parse_date(self, date_val: Any) -> date | None:
        """Converts str objects to date object if necessary"""
        if not date_val:
            return None
        if isinstance(date_val, date) and not isinstance(date_val, datetime):
            return date_val
        if isinstance(date_val, datetime):
            return date_val.date()
        if isinstance(date_val, str):
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
                try:
                    return datetime.strptime(date_val.strip(), fmt).date()
                except ValueError:
                    continue
        return None

    @property
    def is_vendor(self) -> bool:
        """
        returns true if user belongs to vendedor group.
        """
        if not self.user or not hasattr(self.user, 'groups'):
            return False
        return self.user.groups.filter(name='vendedor').exists()

    def _init_config(self) -> None:
        """init config vars from cleaned_data."""
        config = self.cleaned_data or {}
        self.order_by = config.get('order_contrib') or 'net_amount'
        if self.is_vendor:
            self.order_by = 'net_amount'
        self._init_references()

    class CategoryObj:
        """aux obj to make categories accessible in template"""
        def __init__(self, name: str):
            self.name = name.lower()
        def __str__(self):
            return self.name

    def _init_references(self) -> None:
        refs = list(
            Reference.objects.filter(
                context__in=['categoria_cliente_monto', 'categoria_frecuencia_compra', 'clases_producto_relevantes']
            )
        )
        cat_refs = [r for r in refs if r.context == 'categoria_cliente_monto']
        self.categories = sorted(
            [(r.key.lower(), Decimal(r.value)) for r in cat_refs],
            key=lambda x: x[1],
            reverse=True
        )

        freq_refs = [r for r in refs if r.context == 'categoria_frecuencia_compra']
        parsed_freq = []
        for r in freq_refs:
            try:
                parsed_freq.append((r.key.strip().lower(), int(r.value)))
            except (ValueError, TypeError):
                continue
        if parsed_freq:
            self.frequency_categories = sorted(parsed_freq, key=lambda x: x[1])
        else:
            self.frequency_categories = [('regular', 30), ('irregular', 60)]

        rel_refs = [r for r in refs if r.context == 'clases_producto_relevantes']
        classes = [r.key.strip().lower() for r in rel_refs if r.key]
        if classes:
            self.relevant_classes = classes
        else:
            self.relevant_classes = ['dmd', 'nat', 'tow', 'care', 'msd', 'vtq', 'zts']

        self.class_names = dict(ProductClass.objects.values_list('id', 'name'))

    def _init_categories(self) -> None:
        pass

    def _init_frequency_categories(self) -> None:
        pass

    def _init_relevant_classes(self) -> None:
        pass

    def _calculate_category(self, sales: Decimal | float | None = None) -> str:
        '''returns the category which the customer belongs given a sales amount'''
        if sales is None:
            return 'c'
        sales_dec = Decimal(str(sales))
        for name, min_amount in self.categories:
            if sales_dec >= min_amount:
                return name
        return 'c'

    def _get_target_customers(self) -> list[Any]:
        """
        returns customers from self.customers_qs who were registered on or before self.date_end
        (or historical customers with no registration_date).
        """
        if self._cached_target_customers is not None:
            return self._cached_target_customers

        self._cached_target_customers = [
            c for c in self.customers_qs
            if not c.registration_date or c.registration_date <= self.date_end
        ]
        return self._cached_target_customers

    def _get_sales_metrics(self, customer_ids: list[Any]) -> dict[Any, dict[str, Any]]:
        """
        calculates all key sales amounts using targeted date filtered queries:
        - previous year total
        - previous quarter total
        - previous month total
        - current year total
        - contribution period net amount and profit
        - monthly net sales (m_1 .. m_12) for the current year
        returns {customer_id: {'prev_year': Decimal, 'prev_q': Decimal, 'prev_m': Decimal, 'curr_y': Decimal, 'contrib_net': Decimal, 'contrib_profit': Decimal, 'm_1': Decimal, ...}}
        """
        if not customer_ids:
            return {}

        current_year = self.current_year
        previous_year = self.previous_year
        today = self.today
        clean_tx = self.transactions_qs.select_related(None).order_by()

        sales_metrics = defaultdict(dict)

        #previous year
        for r in clean_tx.filter(customer_id__in=customer_ids, sale_date__year=previous_year).values('customer_id').annotate(total=Sum('net_amount')):
            sales_metrics[r['customer_id']]['prev_year'] = r['total']

        # contribution period
        for r in clean_tx.filter(customer_id__in=customer_ids, sale_date__gte=self.date_start, sale_date__lte=self.date_end).values('customer_id').annotate(net=Sum('net_amount'), profit=Sum('profit')):
            sales_metrics[r['customer_id']]['contrib_net'] = r['net']
            sales_metrics[r['customer_id']]['contrib_profit'] = r['profit']

        # monthly breakdown for current year
        for r in clean_tx.filter(customer_id__in=customer_ids, sale_date__year=current_year).values('customer_id', 'sale_date__month').annotate(total=Sum('net_amount')):
            sales_metrics[r['customer_id']][f"m_{r['sale_date__month']}"] = r['total']

        # derivations:
        last_day_prev_month = today.replace(day=1) - timedelta(days=1)
        first_day_prev_month = last_day_prev_month.replace(day=1)

        prev_q_in_curr_year = (self.first_day_q.year == current_year and self.last_day_q.year == current_year)
        prev_q_months = list(range(self.first_day_q.month, self.last_day_q.month + 1)) if prev_q_in_curr_year else []
        if not prev_q_in_curr_year:
            for r in clean_tx.filter(customer_id__in=customer_ids, sale_date__gte=self.first_day_q, sale_date__lte=self.last_day_q).values('customer_id').annotate(total=Sum('net_amount')):
                sales_metrics[r['customer_id']]['prev_q'] = r['total']

        prev_m_in_curr_year = (first_day_prev_month.year == current_year)
        prev_m_month = first_day_prev_month.month if prev_m_in_curr_year else None
        if not prev_m_in_curr_year:
            for r in clean_tx.filter(customer_id__in=customer_ids, sale_date__gte=first_day_prev_month, sale_date__lte=last_day_prev_month).values('customer_id').annotate(total=Sum('net_amount')):
                sales_metrics[r['customer_id']]['prev_m'] = r['total']

        for cid, c_m in sales_metrics.items():
            c_m['curr_y'] = sum((c_m.get(f'm_{m}') or Decimal('0.00') for m in range(1, 13)), Decimal('0.00'))

            if prev_q_months and 'prev_q' not in c_m:
                c_m['prev_q'] = sum((c_m.get(f'm_{m}') or Decimal('0.00') for m in prev_q_months), Decimal('0.00'))

            if prev_m_month and 'prev_m' not in c_m:
                c_m['prev_m'] = c_m.get(f'm_{prev_m_month}') or Decimal('0.00')

        return sales_metrics

    def _get_prev_year_sales(self, customer_ids: list[Any]) -> dict[Any, Decimal]:
        """returns {customer_id: total_net_sales_prev_year}"""
        metrics = self._get_sales_metrics(customer_ids)
        return {cid: row.get('prev_year') or Decimal('0.00') for cid, row in metrics.items()}

    def _get_prev_quarter_sales(self, customer_ids: list[Any]) -> dict[Any, Decimal]:
        """returns {customer_id: total_net_sales_prev_quarter}"""
        metrics = self._get_sales_metrics(customer_ids)
        return {cid: row.get('prev_q') or Decimal('0.00') for cid, row in metrics.items()}

    def _get_prev_month_sales(self, customer_ids: list[Any]) -> dict[Any, Decimal]:
        """returns {customer_id: total_net_sales_prev_month}"""
        metrics = self._get_sales_metrics(customer_ids)
        return {cid: row.get('prev_m') or Decimal('0.00') for cid, row in metrics.items()}

    def _get_curr_year_sales(self, customer_ids: list[Any]) -> dict[Any, Decimal]:
        """returns {customer_id: total_net_sales_current_year}"""
        metrics = self._get_sales_metrics(customer_ids)
        return {cid: row.get('curr_y') or Decimal('0.00') for cid, row in metrics.items()}

    def _calculate_period_avg(self, total_sales: Decimal | float, start_date: date, end_date: date, reg_date: date | None) -> Decimal:
        """calculates the monthly sales average for any period, respecting active months given registration date"""
        if not total_sales or total_sales <= 0:
            return Decimal('0.00')
        if not reg_date or reg_date <= start_date:
            months = (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month) + 1
        elif reg_date > end_date:
            return Decimal('0.00')
        else:
            months = (end_date.year - reg_date.year) * 12 + (end_date.month - reg_date.month) + 1
        months = max(months, 1)
        return total_sales / Decimal(months)

    def _calculate_sale_frequency(self, customer_ids: list[Any]) -> dict[Any, dict[str, Any]]:
        """calculates purchase frequency per customer based on average days between purchases (> 0)"""
        if not customer_ids:
            return {}
        clean_tx = self.transactions_qs.select_related(None).order_by()
        dates_qs = (
            clean_tx
            .filter(
                customer_id__in=customer_ids,
                sale_date__gte=date(self.previous_year, 1, 1),
                net_amount__gt=0
            )
            .values('customer_id', 'sale_date')
            .distinct()
            .order_by('customer_id', 'sale_date')
        )

        customer_dates = defaultdict(list)
        for row in dates_qs:
            customer_dates[row['customer_id']].append(row['sale_date'])

        freq_map = {}
        for c_id, dates in customer_dates.items():
            if len(dates) < 2:
                freq_map[c_id] = {'name': 'nula', 'days': 0}
                continue

            intervals = [(dates[i] - dates[i-1]).days for i in range(1, len(dates))]
            avg_interval = round(sum(intervals) / len(intervals)) if intervals else 0

            freq_name = 'atipico'
            for name, max_days in self.frequency_categories:
                if avg_interval <= max_days:
                    freq_name = name
                    break

            freq_map[c_id] = {'name': freq_name, 'days': avg_interval}

        return freq_map

    def _calculate_product_classes_consumption(self, customer_ids: list[Any]) -> dict[Any, dict[str, Any]]:
        """
        returns unique consumed product classes per customer (considering only relevant classes).
        returns {customer_id: {class_id: {'name': class_name, 'total': total_net_amount}}}
        """
        if not customer_ids or not self.relevant_classes:
            return {}

        clean_tx = self.transactions_qs.select_related(None).order_by()
        classes_qs = (
            clean_tx
            .filter(
                customer_id__in=customer_ids,
                sale_date__gte=self.first_day_q,
                sale_date__lte=self.last_day_q,
                net_amount__gt=0,
                product_class_id__in=self.relevant_classes
            )
            .values('customer_id', 'product_class_id')
            .annotate(total=Sum('net_amount'))
        )

        classes_map = defaultdict(dict)
        class_names = getattr(self, 'class_names', {})
        for row in classes_qs:
            cid = row['customer_id']
            class_id = row['product_class_id']
            class_name = class_names.get(class_id, class_id)
            classes_map[cid][class_id] = {
                'name': class_name,
                'total': row['total'] or Decimal('0.00')
            }

        return dict(classes_map)

    def _get_collections_info(self, customer_ids: list[Any]) -> dict[Any, dict[str, Decimal]]:
        '''
        returns a dictionary with current balance, overdue balance and total balance for each customer.
        returns {customer_id: {'current_balance': Decimal, 'overdue_balance': Decimal, 'total_balance': Decimal}}
        '''        
        if not customer_ids:
            return {}
        clean_ar = self.ars_qs.select_related(None).order_by()
        ar_data = (
            clean_ar
            .filter(customer_id__in=customer_ids)
            .values('customer_id')
            .annotate(
                total_balance=Sum('total_balance'),
                current_balance=Sum('current_balance')
            )
        )
        collections_map = {}
        for row in ar_data:
            cid = row['customer_id']
            curr_b = row['current_balance'] or Decimal('0.00')
            tot_b = row['total_balance'] or Decimal('0.00')
            overdue_b = max(tot_b - curr_b, Decimal('0.00'))
            collections_map[cid] = {
                'current_balance': curr_b,
                'overdue_balance': overdue_b,
                'total_balance': tot_b,
            }
        return collections_map

    def _get_contrib_metrics(self, customer_ids: list[Any]) -> dict[Any, dict[str, Decimal]]:
        """
        Returns contribution metrics strictly between date_start and date_end for target customers.
        {customer_id: {'net_amount': Decimal, 'profit': Decimal}}
        """
        metrics = self._get_sales_metrics(customer_ids)
        return {
            cid: {
                'net_amount': row.get('contrib_net') or Decimal('0.00'),
                'profit': row.get('contrib_profit') or Decimal('0.00'),
            }
            for cid, row in metrics.items()
        }

    def _get_customer_assignments_map(self, customer_ids: list[Any]) -> dict[Any, dict[str, str]]:
        """returns active route and business unit for each customer"""
        if not customer_ids:
            return {}
        assignments = (
            CustomerAssignment.objects
            .filter(customer_id__in=customer_ids)
            .filter(Q(end_date__isnull=True) | Q(end_date__gte=self.today))
            .values('customer_id', 'route_id', 'route__business_unit__name')
        )
        route_map = {}
        for a in assignments:
            route_map[a['customer_id']] = {
                'route_id': a['route_id'] or '',
                'business_unit': a['route__business_unit__name'] or '',
            }
        return route_map

    def _get_monthly_consumption(self, customer_ids: list[Any]) -> dict[Any, list[dict[str, Any]]]:
        """
        calculates 12 monthly consumption slots for the current year,
        including month sales and growth vs the previous month.
        returns {customer_id: [{month_number, date, sale, growth_vs_previous_month}, ...]}
        """
        if not customer_ids:
            return {}

        metrics_map = self._get_sales_metrics(customer_ids)
        result = {}
        for cid in customer_ids:
            c_metrics = metrics_map.get(cid, {})
            monthly_list = []
            prev_sale = Decimal('0.00')

            for m in range(1, 13):
                current_sale = c_metrics.get(f'm_{m}') or Decimal('0.00')

                if prev_sale > Decimal('0.00'):
                    growth = ((current_sale - prev_sale) / prev_sale) * Decimal('100.00')
                else:
                    growth = Decimal('100.00') if current_sale > Decimal('0.00') else Decimal('0.00')

                monthly_list.append({
                    'month_number': m,
                    'date': date(self.current_year, m, 1),
                    'sale': current_sale,
                    'growth_vs_previous_month': growth,
                })
                prev_sale = current_sale

            result[cid] = monthly_list

        return result

    def get_stats(self) -> dict[str, Any]:
        """
        returns high-level summary KPIs for the header cards in the template.
        only considers target customers registered on or before date_end.
        """
        if self._cached_stats is not None:
            return self._cached_stats

        self.read_customer_kpis()
        return self._cached_stats

    def read_customer_kpis(self) -> list:
        """Builds and returns fully enriched customer records sorted by Pareto criterion"""
        customers = self._get_target_customers()
        customer_ids = [c.id for c in customers]

        #single consolidated sales metrics query + supporting queries
        sales_metrics_map = self._get_sales_metrics(customer_ids)
        freq_sales_map = self._calculate_sale_frequency(customer_ids)
        classes_consumption_map = self._calculate_product_classes_consumption(customer_ids)
        collections_map = self._get_collections_info(customer_ids)
        needs_routes = not bool(customers and hasattr(customers[0], 'current_route_id') and customers[0].current_route_id is not None)
        routes_map = self._get_customer_assignments_map(customer_ids) if needs_routes else {}

        #period ranges
        start_prev_y = date(self.previous_year, 1, 1)
        end_prev_y = date(self.previous_year, 12, 31)
        start_curr_y = date(self.current_year, 1, 1)
        end_curr_y = self.today

        month_dates = [date(self.current_year, m, 1) for m in range(1, 13)]
        empty_monthly = [
            {'month_number': m, 'date': month_dates[m - 1], 'sale': Decimal('0.00'), 'growth_vs_previous_month': Decimal('0.00')}
            for m in range(1, 13)
        ]
        cat_cache = {c[0]: self.CategoryObj(c[0]) for c in self.categories}
        cat_cache['c'] = self.CategoryObj('c')
        default_freq = {'name': 'nula', 'days': 0}
        default_col = {
            'current_balance': Decimal('0.00'),
            'overdue_balance': Decimal('0.00'),
            'total_balance': Decimal('0.00'),
        }

        global_net = Decimal('0.00')
        global_profit = Decimal('0.00')

        active_customers = []
        inactive_customers = []
        is_profit_order = (self.order_by == 'profit')

        for customer in customers:
            c_id = customer.id
            reg_date = customer.registration_date

            # current route and business unit
            if getattr(customer, 'current_route_id', None) is not None:
                customer.current_route_id = customer.current_route_id or '-'
                customer.current_route_business_unit = getattr(customer, 'current_route_business_unit', None) or '-'
            else:
                r_info = routes_map.get(c_id, {})
                customer.current_route_id = r_info.get('route_id', '-')
                customer.current_route_business_unit = r_info.get('business_unit', '-')

            #sales metrics
            c_metrics = sales_metrics_map.get(c_id)
            if c_metrics:
                customer.previous_year_total = c_metrics.get('prev_year') or Decimal('0.00')
                customer.previous_quarter_total = c_metrics.get('prev_q') or Decimal('0.00')
                customer.previous_month_total = c_metrics.get('prev_m') or Decimal('0.00')
                customer.current_year_total = c_metrics.get('curr_y') or Decimal('0.00')
                customer.performance_net_amount = c_metrics.get('contrib_net') or Decimal('0.00')
                customer.performance_profit = c_metrics.get('contrib_profit') or Decimal('0.00')

                monthly_list = []
                prev_sale = Decimal('0.00')
                for m in range(1, 13):
                    current_sale = c_metrics.get(f'm_{m}') or Decimal('0.00')
                    if prev_sale > Decimal('0.00'):
                        growth = ((current_sale - prev_sale) / prev_sale) * Decimal('100.00')
                    else:
                        growth = Decimal('100.00') if current_sale > Decimal('0.00') else Decimal('0.00')

                    monthly_list.append({
                        'month_number': m,
                        'date': month_dates[m - 1],
                        'sale': current_sale,
                        'growth_vs_previous_month': growth,
                    })
                    prev_sale = current_sale

                customer.monthly_consumption = monthly_list
            else:
                customer.previous_year_total = Decimal('0.00')
                customer.previous_quarter_total = Decimal('0.00')
                customer.previous_month_total = Decimal('0.00')
                customer.current_year_total = Decimal('0.00')
                customer.performance_net_amount = Decimal('0.00')
                customer.performance_profit = Decimal('0.00')
                customer.monthly_consumption = empty_monthly

            # prev sale avg
            customer.previous_year_avg = self._calculate_period_avg(customer.previous_year_total, start_prev_y, end_prev_y, reg_date)
            customer.previous_quarter_avg = self._calculate_period_avg(customer.previous_quarter_total, self.first_day_q, self.last_day_q, reg_date)
            customer.current_year_avg = self._calculate_period_avg(customer.current_year_total, start_curr_y, end_curr_y, reg_date)

            # categories according to prev periods sales
            customer.category_prev_year = cat_cache.get(self._calculate_category(customer.previous_year_avg), cat_cache['c'])
            customer.category_prev_quarter = cat_cache.get(self._calculate_category(customer.previous_quarter_avg), cat_cache['c'])
            customer.category_prev_month = cat_cache.get(self._calculate_category(customer.previous_month_total), cat_cache['c'])

            # sale freq
            c_freq = freq_sales_map.get(c_id, default_freq)
            customer.frequency = c_freq['name']
            customer.frequency_days = c_freq['days']

            # collections
            col_info = collections_map.get(c_id, default_col)
            customer.current_balance = col_info['current_balance']
            customer.overdue_balance = col_info['overdue_balance']
            customer.total_balance = col_info['total_balance']

            credit_limit = customer.credit_limit or Decimal('0.00')
            if credit_limit > Decimal('0.00'):
                customer.credit_usage = (customer.total_balance / credit_limit) * Decimal('100.00')
            else:
                customer.credit_usage = Decimal('0.00')

            # agreements
            customer.active_agreements = 0

            # classes consumption
            customer.product_classes_consumed = classes_consumption_map.get(c_id, {})
            customer.product_classes_with_consumption = len(customer.product_classes_consumed)

            # contrib defaults
            customer.selected_contrib_by = 'profit' if is_profit_order else 'net_amount'
            customer.contrib_net_amount = Decimal('0.00')
            customer.net_amount = Decimal('0.00')
            customer.contrib_profit = Decimal('0.00')
            customer.profit = Decimal('0.00')
            customer.cumuled_contrib = Decimal('0.00')
            customer.cumuled_portafolio_count = 0
            customer.cumuled_portafolio_pct = Decimal('0.00')

            global_net += customer.performance_net_amount
            global_profit += customer.performance_profit

            if is_profit_order:
                if customer.performance_profit > Decimal('0.00'):
                    active_customers.append(customer)
                else:
                    inactive_customers.append(customer)
            else:
                if customer.performance_net_amount > Decimal('0.00'):
                    active_customers.append(customer)
                else:
                    inactive_customers.append(customer)

        # pareto sorting & accumulation based on selected criterion
        if is_profit_order:
            active_customers.sort(key=lambda x: x.performance_profit, reverse=True)
        else:
            active_customers.sort(key=lambda x: x.performance_net_amount, reverse=True)

        total_active_customers = len(active_customers)
        cumuled_val = Decimal('0.00')

        for index, customer in enumerate(active_customers, start=1):
            if global_net > Decimal('0.00'):
                customer.contrib_net_amount = (customer.performance_net_amount / global_net) * Decimal('100.00')
            customer.net_amount = customer.contrib_net_amount

            if global_profit > Decimal('0.00'):
                customer.contrib_profit = (customer.performance_profit / global_profit) * Decimal('100.00')
            customer.profit = customer.contrib_profit

            primary_contrib = customer.contrib_profit if is_profit_order else customer.contrib_net_amount
            cumuled_val += primary_contrib
            customer.cumuled_contrib = cumuled_val
            customer.cumuled_portafolio_count = index
            customer.cumuled_portafolio_pct = (Decimal(index) / Decimal(total_active_customers)) * Decimal('100.00') if total_active_customers > 0 else Decimal('0.00')

        registered_customers = len(customers)
        customers_with_consumption = total_active_customers
        customers_without_consumption = max(registered_customers - customers_with_consumption, 0)

        self._cached_stats = {
            'registered_customers': registered_customers,
            'customers_with_consumption': customers_with_consumption,
            'customers_with_consumption_pct': (Decimal(customers_with_consumption) / Decimal(registered_customers) * Decimal('100.00')) if registered_customers > 0 else Decimal('0.00'),
            'customers_without_consumption': customers_without_consumption,
            'customers_without_consumption_pct': (Decimal(customers_without_consumption) / Decimal(registered_customers) * Decimal('100.00')) if registered_customers > 0 else Decimal('0.00'),
            'net_amount': global_net,
            'net_sales': global_net,
            'profit': global_profit,
            'margin': (global_profit / global_net * Decimal('100.00')) if global_net > Decimal('0.00') else Decimal('0.00')
        }

        return active_customers + inactive_customers

@dataclass
class CustomerKpisExports:
    '''dedicated to receive all exports request for customer kpis objects and filters'''
    customer_kpis_service: CustomerKpisService

    def export_customer_kpis_report(self) -> io.BytesIO:
        customers_data = self.customer_kpis_service.read_customer_kpis()
        kpis = self.customer_kpis_service.get_stats()

        wb = openpyxl.Workbook()

        #styles
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        section_font = Font(name="Calibri", size=12, bold=True, color="0F172A")
        title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
        subtitle_font = Font(name="Calibri", size=9, italic=True, color="64748B")
        data_font = Font(name="Calibri", size=10)
        bold_data_font = Font(name="Calibri", size=10, bold=True)

        thin_border_side = Side(style='thin', color='CBD5E1')
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        currency_format = '"$"#,##0.00'
        pct_format = '0.00%'
        int_format = '#,##0'

        # gens
        ws_summary = wb.active
        ws_summary.title = "Resumen General"
        ws_summary.views.sheetView[0].showGridLines = True

        #title
        ws_summary.cell(row=1, column=1, value="REPORTE DE KPIS DE CLIENTES - RESUMEN GENERAL").font = title_font
        now_str = timezone.localtime().strftime('%Y-%m-%d %H:%M')
        d_start_str = self.customer_kpis_service.date_start.strftime('%Y-%m-%d') if self.customer_kpis_service.date_start else ''
        d_end_str = self.customer_kpis_service.date_end.strftime('%Y-%m-%d') if self.customer_kpis_service.date_end else ''
        criterio_str = "Utilidad" if self.customer_kpis_service.order_by == 'profit' else "Venta Neta"
        ws_summary.cell(row=2, column=1, value=f"Generado el: {now_str} | Periodo de análisis: {d_start_str} al {d_end_str} | Criterio de evaluación: {criterio_str}").font = subtitle_font

        #sect 1, gen indicators
        ws_summary.cell(row=4, column=1, value="Indicadores Generales de Cartera").font = section_font
        general_headers = ["Indicador", "Clientes", "% de Cartera"]
        for col_num, h_text in enumerate(general_headers, 1):
            cell = ws_summary.cell(row=5, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_num > 1 else "left")
            cell.border = cell_border

        reg_c = int(kpis.get('registered_customers') or 0)
        with_c = int(kpis.get('customers_with_consumption') or 0)
        with_pct = float(kpis.get('customers_with_consumption_pct') or 0) / 100.0
        without_c = int(kpis.get('customers_without_consumption') or 0)
        without_pct = float(kpis.get('customers_without_consumption_pct') or 0) / 100.0

        general_rows = [
            ("Clientes registrados", reg_c, 1.0),
            ("Clientes con consumo", with_c, with_pct),
            ("Clientes sin consumo", without_c, without_pct),
        ]

        for row_idx, (label, val, pct) in enumerate(general_rows, 6):
            c_lbl = ws_summary.cell(row=row_idx, column=1, value=label)
            c_lbl.font = data_font
            c_lbl.border = cell_border

            c_val = ws_summary.cell(row=row_idx, column=2, value=val)
            c_val.font = bold_data_font
            c_val.number_format = int_format
            c_val.alignment = Alignment(horizontal="right")
            c_val.border = cell_border

            c_pct = ws_summary.cell(row=row_idx, column=3, value=pct)
            c_pct.font = bold_data_font
            c_pct.number_format = pct_format
            c_pct.alignment = Alignment(horizontal="right")
            c_pct.border = cell_border

        #gen 2, performance, net and profit 
        start_row_perf = 11
        ws_summary.cell(row=start_row_perf, column=1, value=f"Rendimiento del Periodo ({d_start_str} al {d_end_str})").font = section_font
        perf_headers = ["Concepto", "Monto Total", "% Margen"]
        for col_num, h_text in enumerate(perf_headers, 1):
            cell = ws_summary.cell(row=start_row_perf + 1, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_num > 1 else "left")
            cell.border = cell_border

        is_vendor = self.customer_kpis_service.is_vendor
        net_s = float(kpis.get('net_amount') or kpis.get('net_sales') or 0)
        prof_s = float(kpis.get('profit') or 0)
        marg_s = float(kpis.get('margin') or 0) / 100.0

        perf_rows = [
            ("Venta Neta", net_s, None),
        ]
        if not is_vendor:
            perf_rows.append(("Utilidad Bruta", prof_s, marg_s))

        for idx, (lbl, amt, marg) in enumerate(perf_rows, start_row_perf + 2):
            c_l = ws_summary.cell(row=idx, column=1, value=lbl)
            c_l.font = data_font
            c_l.border = cell_border

            c_a = ws_summary.cell(row=idx, column=2, value=amt)
            c_a.font = bold_data_font
            c_a.number_format = currency_format
            c_a.alignment = Alignment(horizontal="right")
            c_a.border = cell_border

            c_m = ws_summary.cell(row=idx, column=3, value=marg if marg is not None else "-")
            c_m.font = bold_data_font
            if marg is not None:
                c_m.number_format = pct_format
            c_m.alignment = Alignment(horizontal="right" if marg is not None else "center")
            c_m.border = cell_border

        # sheet 2, complete table
        ws_customers = wb.create_sheet(title="Listado de Clientes")
        ws_customers.views.sheetView[0].showGridLines = True

        #superheaders
        if is_vendor:
            superheaders = [
                ("Identificación", 1, 4),# Cols 1-4 (A-D)
                ("Segmentación", 5, 9), # Cols 5-9 (E-I)
                ("Cobranza", 10, 14),   # Cols 10-14 (J-N)
                ("Métricas de Consumo", 15, 19), # Cols 15-19 (O-S)
                (f"Métricas de Contribución ({d_start_str} al {d_end_str})", 20, 24),  # Cols 20-24 (T-X)
                (f"Desglose de Consumos Mensuales {self.customer_kpis_service.current_year}", 25, 36), # Cols 25-36 (Y-AJ)
            ]
            contrib_headers = [
                "Venta Neta Periodo",
                "% Contribución Venta Neta",
                "% Contribución Acumulada",
                "Clientes Acumulados",
                "% Cartera Acumulada",
            ]
        else:
            superheaders = [
                ("Identificación", 1, 4),# Cols 1-4 (A-D)
                ("Segmentación", 5, 9), # Cols 5-9 (E-I)
                ("Cobranza", 10, 14),   # Cols 10-14 (J-N)
                ("Métricas de Consumo", 15, 19), # Cols 15-19 (O-S)
                (f"Métricas de Contribución ({d_start_str} al {d_end_str})", 20, 26),  # Cols 20-26 (T-Z)
                (f"Desglose de Consumos Mensuales {self.customer_kpis_service.current_year}", 27, 38), # Cols 27-38 (AA-AL)
            ]
            contrib_headers = [
                "Venta Neta Periodo",
                "Utilidad Periodo",
                "% Contribución Venta Neta",
                "% Contribución Utilidad",
                "% Contribución Acumulada",
                "Clientes Acumulados",
                "% Cartera Acumulada",
            ]

        for title, start_col, end_col in superheaders:
            if start_col == end_col:
                cell = ws_customers.cell(row=1, column=start_col, value=title)
            else:
                ws_customers.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                cell = ws_customers.cell(row=1, column=start_col, value=title)
            
            for c_idx in range(start_col, end_col + 1):
                c_head = ws_customers.cell(row=1, column=c_idx)
                c_head.fill = header_fill
                c_head.font = header_font
                c_head.border = cell_border
                c_head.alignment = Alignment(horizontal="center", vertical="center")

        ws_customers.row_dimensions[1].height = 24

        #column headers row 2
        month_names = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        
        column_headers = [
            # ids (1-4)
            "ID Cliente",
            "Nombre Cliente",
            "Ruta",
            "Gerencia",
            # segment (5-9)
            "Tipo de Cliente",
            "Categoría",
            "Frecuencia de Compra",
            "Frecuencia (Días)",
            "Líder de Opinión",
            # collections (10-14)
            "Límite de Crédito",
            "% Uso de Crédito",
            "Saldo al Corriente",
            "Saldo Vencido",
            "Saldo Total",
            # consumption metrics (15-19)
            "Convenios Activos",
            "Clases con Consumo",
            "Promedio Mensual Año Previo",
            "Promedio Mensual Año Actual",
            "Promedio Último Trimestre",
            # contrib
            *contrib_headers,
            # monthly
            *[f"Venta {m}" for m in month_names]
        ]

        for col_num, h_text in enumerate(column_headers, 1):
            cell = ws_customers.cell(row=2, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = cell_border

        ws_customers.row_dimensions[2].height = 26

        # Populate rows
        for row_idx, c in enumerate(customers_data, 3):
            cid = c.id or ''
            cname = (c.name or '').title()
            r_id = getattr(c, 'current_route_id', '-')
            r_bu = (getattr(c, 'current_route_business_unit', '-') or '').title()
            
            c_type = (c.customer_type.name or '').title() if hasattr(c, 'customer_type') and c.customer_type else ''
            cat_obj = getattr(c, 'category_prev_quarter', None)
            cat_name = (getattr(cat_obj, 'name', '') or '').upper()
            freq_name = (getattr(c, 'frequency', '') or '').title()
            freq_days = int(getattr(c, 'frequency_days', 0) or 0)
            op_leader = "Sí" if getattr(c, 'opinion_leader', False) else "No"

            credit_limit = float(getattr(c, 'credit_limit', 0) or 0)
            credit_usage = float(getattr(c, 'credit_usage', 0) or 0) / 100.0
            curr_bal = float(getattr(c, 'current_balance', 0) or 0)
            overdue_bal = float(getattr(c, 'overdue_balance', 0) or 0)
            tot_bal = float(getattr(c, 'total_balance', 0) or 0)

            active_agreements = int(getattr(c, 'active_agreements', 0) or 0)
            prod_classes = int(getattr(c, 'product_classes_with_consumption', 0) or 0)
            prev_y_avg = float(getattr(c, 'previous_year_avg', 0) or 0)
            curr_y_avg = float(getattr(c, 'current_year_avg', 0) or 0)
            prev_q_avg = float(getattr(c, 'previous_quarter_avg', 0) or 0)

            perf_net = float(getattr(c, 'performance_net_amount', 0) or 0)
            perf_profit = float(getattr(c, 'performance_profit', 0) or 0)
            contrib_net_pct = float(getattr(c, 'contrib_net_amount', getattr(c, 'net_amount', 0)) or 0) / 100.0
            contrib_profit_pct = float(getattr(c, 'contrib_profit', getattr(c, 'profit', 0)) or 0) / 100.0
            cumuled_contrib = float(getattr(c, 'cumuled_contrib', 0) or 0) / 100.0
            cumuled_count = int(getattr(c, 'cumuled_portafolio_count', 0) or 0)
            cumuled_pct = float(getattr(c, 'cumuled_portafolio_pct', 0) or 0) / 100.0

            monthly_sales = []
            m_list = getattr(c, 'monthly_consumption', [])
            m_dict = {item['month_number']: float(item.get('sale') or 0) for item in m_list if isinstance(item, dict)}
            for m_num in range(1, 13):
                monthly_sales.append(m_dict.get(m_num, 0.0))

            if is_vendor:
                contrib_values = [
                    (perf_net, currency_format, "right"),
                    (contrib_net_pct, pct_format, "right"),
                    (cumuled_contrib, pct_format, "right"),
                    (cumuled_count, int_format, "right"),
                    (cumuled_pct, pct_format, "right"),
                ]
            else:
                contrib_values = [
                    (perf_net, currency_format, "right"),
                    (perf_profit, currency_format, "right"),
                    (contrib_net_pct, pct_format, "right"),
                    (contrib_profit_pct, pct_format, "right"),
                    (cumuled_contrib, pct_format, "right"),
                    (cumuled_count, int_format, "right"),
                    (cumuled_pct, pct_format, "right"),
                ]

            row_values = [
                # ids
                (cid, '@', "center"),
                (cname, None, "left"),
                (r_id, '@', "center"),
                (r_bu, None, "left"),
                # segs
                (c_type, None, "left"),
                (cat_name, '@', "center"),
                (freq_name, None, "left"),
                (freq_days, int_format, "right"),
                (op_leader, '@', "center"),
                # collections
                (credit_limit, currency_format, "right"),
                (credit_usage, pct_format, "right"),
                (curr_bal, currency_format, "right"),
                (overdue_bal, currency_format, "right"),
                (tot_bal, currency_format, "right"),
                # consumption metrics
                (active_agreements, int_format, "right"),
                (prod_classes, int_format, "right"),
                (prev_y_avg, currency_format, "right"),
                (curr_y_avg, currency_format, "right"),
                (prev_q_avg, currency_format, "right"),
                # contrib
                *contrib_values,
                # monthly
                *[(m_sale, currency_format, "right") for m_sale in monthly_sales]
            ]

            for col_idx, (val, num_fmt, align_h) in enumerate(row_values, 1):
                cell = ws_customers.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = cell_border
                cell.alignment = Alignment(horizontal=align_h)
                if num_fmt:
                    cell.number_format = num_fmt

        # autosize columns
        for sheet in [ws_summary, ws_customers]:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if cell.number_format == currency_format:
                        val_str = f"${val_str}"
                    max_len = max(max_len, len(val_str))
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


@dataclass
class CustomerProfileService(CustomerKpisService):
    '''
    dedicated to calculate dynamic metrics and attributes for a single Customer Profile.
    enriches the customer instance with consumption categories, behavior KPIs,
    filtered collections (Accounts Receivable), and monthly consumption by product class.
    '''
    customer: Any = None
    customers_qs: QuerySet | None = None
    transactions_qs: QuerySet | None = None
    ars_qs: QuerySet | None = None

    def __post_init__(self):
        if self.customer and self.customers_qs is None:
            from apps.customers.models import Customer
            self.customers_qs = Customer.objects.filter(pk=self.customer.pk)

        if self.transactions_qs is None:
            from apps.sales.services.sale_transactions import SaleTransactionsService
            self.transactions_qs = SaleTransactionsService(user=self.user).read_transactions_by_allowed_customers()

        if self.ars_qs is None:
            from apps.customers.services.accounts_receivables import AccountsReceivablesService
            self.ars_qs = AccountsReceivablesService(user=self.user).read_ars_by_allowed_customers()

        super().__post_init__()

        #baseline dates
        first_day_curr_month = self.today.replace(day=1)
        self.last_day_prev_month = first_day_curr_month - timedelta(days=1)
        self.first_day_prev_month = self.last_day_prev_month.replace(day=1)

        self.first_day_prev_year = date(self.previous_year, 1, 1)
        self.last_day_prev_year = date(self.previous_year, 12, 31)

    def build_profile(self) -> Any:
        '''
        enriches and returns the customer object with all profile attributes.
        '''
        if not self.customer:
            return None

        self._set_static_categories()
        self._set_behavior_kpis()
        self._set_collections_kpis()
        self._set_monthly_consumption_by_class()
        return self.customer

    def _set_static_categories(self) -> None:
        '''
        calculates static consumption categories:
        • monthly category (previous full month)
        • quarterly category (previous 3 full months average)
        • annual category (previous year average)
        '''
        from apps.sales.services.sale_transactions import SaleTransactionsService
        base_txs = SaleTransactionsService(user=self.user).read_transactions_by_allowed_customers().filter(customer=self.customer)

        #previous full month
        sales_prev_month = base_txs.filter(
            sale_date__gte=self.first_day_prev_month,
            sale_date__lte=self.last_day_prev_month
        ).aggregate(total=Sum('net_amount'))['total'] or Decimal('0.00')
        self.customer.category_prev_month = self.CategoryObj(self._calculate_category(sales_prev_month))

        #previous full quarter average
        sales_q = base_txs.filter(
            sale_date__gte=self.first_day_q,
            sale_date__lte=self.last_day_q
        ).aggregate(total=Sum('net_amount'))['total'] or Decimal('0.00')
        avg_q = self._calculate_period_avg(sales_q, self.first_day_q, self.last_day_q, self.customer.registration_date)
        self.customer.category_prev_quarter = self.CategoryObj(self._calculate_category(avg_q))

        #previous full year average
        sales_y = base_txs.filter(
            sale_date__gte=self.first_day_prev_year,
            sale_date__lte=self.last_day_prev_year
        ).aggregate(total=Sum('net_amount'))['total'] or Decimal('0.00')
        avg_y = self._calculate_period_avg(sales_y, self.first_day_prev_year, self.last_day_prev_year, self.customer.registration_date)
        self.customer.category_prev_year = self.CategoryObj(self._calculate_category(avg_y))

    def _set_behavior_kpis(self) -> None:
        '''
        calculates purchase frequency, consumed classes count, and top class in previous quarter.
        '''
        filtered_txs = self.transactions_qs.filter(customer=self.customer)

        #purchase frequency
        dates = list(
            filtered_txs.filter(net_amount__gt=0)
            .order_by('sale_date')
            .values_list('sale_date', flat=True)
            .distinct()
        )

        if len(dates) < 2:
            self.customer.frequency = 'nula'
            self.customer.frequency_days = 0
            self.customer.purchase_frequency_category = 'Nula'
            self.customer.purchase_frequency_days = ''
        else:
            intervals = [(dates[i] - dates[i-1]).days for i in range(1, len(dates))]
            avg_interval = round(sum(intervals) / len(intervals)) if intervals else 0

            freq_name = 'atipico'
            for name, max_days in self.frequency_categories:
                if avg_interval <= max_days:
                    freq_name = name
                    break

            self.customer.frequency = freq_name
            self.customer.frequency_days = avg_interval
            self.customer.purchase_frequency_category = freq_name.title()
            self.customer.purchase_frequency_days = f"cada {avg_interval} días"

        #consumed relevant classes count
        classes_count = (
            filtered_txs.filter(
                net_amount__gt=0,
                product_class_id__in=self.relevant_classes
            )
            .values('product_class_id')
            .distinct()
            .count()
        )
        self.customer.consumed_classes = classes_count
        self.customer.product_classes_with_consumption = classes_count

        #most consumed class in previous quarter
        from apps.sales.services.sale_transactions import SaleTransactionsService
        base_txs = SaleTransactionsService(user=self.user).read_transactions_by_allowed_customers().filter(customer=self.customer)
        top_class = (
            base_txs.filter(
                sale_date__gte=self.first_day_q,
                sale_date__lte=self.last_day_q,
                net_amount__gt=0
            )
            .values('product_class__id', 'product_class__name')
            .annotate(total_net=Sum('net_amount'))
            .order_by('-total_net')
            .first()
        )
        if top_class and top_class.get('product_class__name'):
            self.customer.most_consumed_class_last_q = {
                'id': top_class['product_class__id'],
                'name': top_class['product_class__name']
            }
        else:
            self.customer.most_consumed_class_last_q = {'id': '', 'name': 'Sin consumo reciente'}

    def _set_collections_kpis(self) -> None:
        '''
        calculates accounts receivable metrics for the customer respecting route/unit filters.
        '''
        ar_qs = self.ars_qs.filter(customer=self.customer)

        #apply route/business_unit/region filters if present in cleaned_data
        if self.cleaned_data:
            if self.cleaned_data.get('route'):
                ar_qs = ar_qs.filter(route__in=self.cleaned_data['route'])
            if self.cleaned_data.get('business_unit'):
                bu_ids = [bu.pk if hasattr(bu, 'pk') else bu for bu in self.cleaned_data['business_unit']]
                ar_qs = ar_qs.filter(route__business_unit_id__in=bu_ids)
            if self.cleaned_data.get('region'):
                region_objs = self.cleaned_data['region']
                region_ids = set(r.pk if hasattr(r, 'pk') else r for r in region_objs)
                all_bu_ids = set(region_ids)
                current_parents = set(region_ids)
                from apps.human_resources.models import BusinessUnit
                while current_parents:
                    child_ids = set(
                        BusinessUnit.objects.filter(parent_id__in=current_parents)
                        .values_list('id', flat=True)
                    )
                    new_ids = child_ids - all_bu_ids
                    if not new_ids:
                        break
                    all_bu_ids.update(new_ids)
                    current_parents = new_ids
                ar_qs = ar_qs.filter(route__business_unit_id__in=all_bu_ids)

        ar_agg = ar_qs.aggregate(
            total_balance=Sum('total_balance'),
            current_balance=Sum('current_balance'),
            balance_15=Sum('balance_15'),
            balance_30=Sum('balance_30'),
            balance_60=Sum('balance_60'),
            past_due=Sum('past_due'),
        )

        self.customer.total_balance = ar_agg['total_balance'] or Decimal('0.00')
        self.customer.current_balance = ar_agg['current_balance'] or Decimal('0.00')
        self.customer.overdue_balance = max(self.customer.total_balance - self.customer.current_balance, Decimal('0.00'))
        self.customer.balance_15 = ar_agg['balance_15'] or Decimal('0.00')
        self.customer.balance_30 = ar_agg['balance_30'] or Decimal('0.00')
        self.customer.balance_60 = ar_agg['balance_60'] or Decimal('0.00')
        self.customer.past_due = ar_agg['past_due'] or Decimal('0.00')

        credit_limit = self.customer.credit_limit or Decimal('0.00')
        if credit_limit > Decimal('0.00'):
            self.customer.credit_usage = (self.customer.total_balance / credit_limit) * Decimal('100.00')
        else:
            self.customer.credit_usage = Decimal('0.00')

    def _set_monthly_consumption_by_class(self) -> None:
        '''
        builds matrix of monthly consumption by product class for the customer based on filtered transactions.
        Defaults to the current year period if no date range is provided.
        '''

        txs = self.transactions_qs.filter(customer=self.customer)

        start_date = self.date_start or date(self.current_year, 1, 1)
        end_date = self.date_end or self.last_day_prev_month

        if start_date > end_date:
            start_date, end_date = end_date, start_date

        start_m = date(start_date.year, start_date.month, 1)
        end_m = date(end_date.year, end_date.month, 1)

        months_list = []
        curr = start_m
        while curr <= end_m:
            months_list.append(curr)
            curr += relativedelta(months=1)

        grouped = (
            txs.filter(sale_date__gte=start_date, sale_date__lte=end_date)
            .values('product_class__id', 'product_class__name')
            .annotate(
                month=TruncMonth('sale_date'),
                total=Sum('net_amount')
            )
        )

        classes_dict = {}
        for row in grouped:
            c_id = row['product_class__id'] or 'sin_clase'
            c_name = row['product_class__name'] or c_id
            m = row['month']
            if hasattr(m, 'date'):
                m = m.date()
            m = date(m.year, m.month, 1)

            if c_id not in classes_dict:
                classes_dict[c_id] = {
                    'product_class': {'id': c_id, 'name': c_name},
                    'months': {mon: Decimal('0.00') for mon in months_list},
                    'total': Decimal('0.00')
                }

            val = row['total'] or Decimal('0.00')
            if m in classes_dict[c_id]['months']:
                classes_dict[c_id]['months'][m] += val
            classes_dict[c_id]['total'] += val

        sorted_classes = sorted(classes_dict.values(), key=lambda x: x['total'], reverse=True)

        totals_row = {
            'months': {mon: Decimal('0.00') for mon in months_list},
            'grand_total': Decimal('0.00')
        }

        for pc in sorted_classes:
            for mon in months_list:
                totals_row['months'][mon] += pc['months'][mon]
            totals_row['grand_total'] += pc['total']
            pc['monthly_amounts'] = [pc['months'][mon] for mon in months_list]

        totals_row['monthly_amounts'] = [totals_row['months'][mon] for mon in months_list]

        self.customer.monthly_consumption_by_class = sorted_classes
        self.customer.consumption_months = months_list
        self.customer.consumption_totals = totals_row


CustomerKpisService.Profile = CustomerProfileService