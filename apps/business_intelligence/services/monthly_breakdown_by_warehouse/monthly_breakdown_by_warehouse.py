from decimal import Decimal
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


from django.db.models import Sum, Count, F
from django.db.models.functions import ExtractMonth
from django.http import HttpResponse



from apps.core.models import SaleTransaction, SaleTarget, Customer, AccountsReceivable, ProductClass




class MonthlyBreakdownByWarehouse:
    def __init__(self, year, allowed_routes_qs, warehouse_id):
        self.year = int(year)
        self.allowed_routes_qs = allowed_routes_qs
        self.warehouse_id = warehouse_id
        
        self.routes_qs = self.allowed_routes_qs.filter(warehouse_id=self.warehouse_id)
        self.route_ids = list(self.routes_qs.values_list('id', flat=True))

    def _get_monthly_targets(self):
        targets_qs = SaleTarget.objects.filter(
            period__year=self.year,
            route_id__in=self.route_ids
        ).annotate(
            month=ExtractMonth('period')
        ).values('route_id', 'product_class__name', 'month').annotate(
            total=Sum('target_amount')
        )
        
        data = defaultdict(Decimal)
        for row in targets_qs:
            key = (row['route_id'], row['product_class__name'].lower(), row['month'])
            data[key] = row['total'] or Decimal('0.00')
        return data

    def _get_monthly_sales(self):
        sales_qs = SaleTransaction.objects.filter(
            sale_date__year=self.year,
            route_id__in=self.route_ids
        ).annotate(
            month=ExtractMonth('sale_date')
        ).values('route_id', 'product_class__name', 'month').annotate(
            total=Sum('net_amount')
        )
        
        data = defaultdict(Decimal)
        for row in sales_qs:
            key = (row['route_id'], row['product_class__name'].lower(), row['month'])
            data[key] = row['total'] or Decimal('0.00')
        return data

    def _get_monthly_margin(self):
        margin_qs = SaleTransaction.objects.filter(
            sale_date__year=self.year,
            route_id__in=self.route_ids
        ).annotate(
            month=ExtractMonth('sale_date')
        ).values('route_id', 'month').annotate(
            total_profit=Sum('profit'),
            total_net=Sum('net_amount')
        )
        data = defaultdict(lambda: {'margin': Decimal('0.00'), 'profit': Decimal('0.00'), 'net': Decimal('0.00')})
        for row in margin_qs:
            t_profit = row['total_profit'] or Decimal('0.00')
            t_net = row['total_net'] or Decimal('0.00')
            if t_net > 0:
                margin = (t_profit / t_net) * Decimal('100.00')
            else:
                margin = Decimal('0.00')
                
            data[(row['route_id'], row['month'])] = {
                'margin': margin,
                'profit': t_profit,
                'net': t_net
            }
        return data

    def _get_monthly_new_customers(self):
        customers_qs = Customer.objects.filter(
            registration_date__year=self.year,
            route_id__in=self.route_ids
        ).annotate(
            month=ExtractMonth('registration_date')
        ).values('route_id', 'month').annotate(
            total=Count('id')
        )
        data = defaultdict(int)
        for row in customers_qs:
            data[(row['route_id'], row['month'])] = row['total'] or 0
        return data

    def _get_monthly_accounts_receivable(self):
        import calendar
        from datetime import date

        ar_qs = AccountsReceivable.objects.filter(
            customer__route_id__in=self.route_ids
        ).values('customer__route_id', 'issue_date', 'total_balance', 'customer_id')

        amount_data = defaultdict(Decimal)
        count_data = defaultdict(int)

        month_end_dates = {}
        for m in range(1, 13):
            last_day = calendar.monthrange(self.year, m)[1]
            month_end_dates[m] = date(self.year, m, last_day)

        customers_per_route_month = defaultdict(set)

        for row in ar_qs:
            route_id = row['customer__route_id']
            issue_date = row['issue_date']
            balance = row['total_balance'] or Decimal('0.00')
            customer_id = row['customer_id']

            for m in range(1, 13):
                if issue_date is None or issue_date <= month_end_dates[m]:
                    amount_data[(route_id, m)] += balance
                    customers_per_route_month[(route_id, m)].add(customer_id)
        
        for key, customers_set in customers_per_route_month.items():
            count_data[key] = len(customers_set)

        return count_data, amount_data

    def _get_monthly_promotions(self):
        return defaultdict(Decimal)

    def _get_monthly_agreements(self):
        return defaultdict(int)

    def get_data(self):
        """
        Retrieve and structure the monthly breakdown data.

        This method orchestrates the data retrieval by calling helper methods for:
        - Monthly targets
        - Monthly sales
        - Monthly margins
        - Monthly new customers
        - Monthly accounts receivable (both count and amount)
        - Monthly promotions
        - Monthly agreements

        Returns:
            list: A list containing a single dictionary with the structured breakdown data.
        """
        product_classes = ProductClass.objects.values_list('name', flat=True)
        pc_names = sorted(list(set(pc.lower() for pc in product_classes)))
        
        targets = self._get_monthly_targets()
        sales = self._get_monthly_sales()
        margins = self._get_monthly_margin()
        new_customers = self._get_monthly_new_customers()
        ar_counts, ar_amounts = self._get_monthly_accounts_receivable()
        promotions = self._get_monthly_promotions()
        agreements = self._get_monthly_agreements()

        from apps.core.models import Warehouse
        try:
            warehouse = Warehouse.objects.get(id=self.warehouse_id)
            warehouse_name = warehouse.name
        except Warehouse.DoesNotExist:
            warehouse_name = "Desconocida"

        temp_warehouse = {
            'warehouse_id': self.warehouse_id,
            'warehouse_name': warehouse_name,
            'routes': []
        }

        results_names = ['margen', 'clientes nuevos', 'cuentas por cobrar', 'cuentas por cobrar $', 'promociones', 'convenios']

        # warehouse summary accumulators
        w_pc_data = {pc: {m: {'target': Decimal('0.00'), 'net_sale': Decimal('0.00')} for m in range(1, 13)} for pc in pc_names}
        w_total_data = {m: {'target': Decimal('0.00'), 'net_sale': Decimal('0.00')} for m in range(1, 13)}
        w_results_data = {res: {m: 0 for m in range(1, 13)} for res in results_names}
        w_margin_acc = {m: {'profit': Decimal('0.00'), 'net': Decimal('0.00')} for m in range(1, 13)}

        for route in self.routes_qs:
            route_id = route.id
            temp_route = {
                'route_id': route_id,
                'route_name': route.name,
                'product_classes': [],
                'total': {'name': 'total', 'monthly_data': []},
                'results': []
            }
            
            route_monthly_totals = {m: {'target': Decimal('0.00'), 'net_sale': Decimal('0.00')} for m in range(1, 13)}

            for pc_name in pc_names:
                temp_pc = {'name': pc_name, 'monthly_data': []}
                for month in range(1, 13):
                    t = targets[(route_id, pc_name, month)]
                    s = sales[(route_id, pc_name, month)]
                    diff = s - t
                    scope = (s / t * Decimal('100')) if t > 0 else (Decimal('100') if s > 0 else Decimal('0'))
                    
                    temp_pc['monthly_data'].append({
                        'target': t,
                        'net_sale': s,
                        'diff': diff,
                        'scope': scope
                    })

                    route_monthly_totals[month]['target'] += t
                    route_monthly_totals[month]['net_sale'] += s
                    
                    w_pc_data[pc_name][month]['target'] += t
                    w_pc_data[pc_name][month]['net_sale'] += s

                temp_route['product_classes'].append(temp_pc)

            for month in range(1, 13):
                t = route_monthly_totals[month]['target']
                s = route_monthly_totals[month]['net_sale']
                diff = s - t
                scope = (s / t * Decimal('100')) if t > 0 else (Decimal('100') if s > 0 else Decimal('0'))

                temp_route['total']['monthly_data'].append({
                    'target': t,
                    'net_sale': s,
                    'diff': diff,
                    'scope': scope
                })
                
                w_total_data[month]['target'] += t
                w_total_data[month]['net_sale'] += s

            for res_name in results_names:
                temp_res = {'name': res_name, 'monthly_data': []}
                for month in range(1, 13):
                    val = 0
                    if res_name == 'margen':
                        m_data = margins[(route_id, month)]
                        if isinstance(m_data, Decimal):
                            m_data = {'margin': m_data, 'profit': Decimal('0.00'), 'net': Decimal('0.00')}
                        val = f"{m_data['margin']:,.2f} %"
                        w_margin_acc[month]['profit'] += m_data['profit']
                        w_margin_acc[month]['net'] += m_data['net']
                    elif res_name == 'clientes nuevos':
                        val = new_customers[(route_id, month)]
                        w_results_data[res_name][month] += val
                    elif res_name == 'cuentas por cobrar':
                        val = ar_counts[(route_id, month)]
                        w_results_data[res_name][month] += val
                    elif res_name == 'cuentas por cobrar $':
                        val_num = ar_amounts[(route_id, month)]
                        val = f"$ {val_num:,.2f}"
                        w_results_data[res_name][month] += val_num
                    elif res_name == 'promociones':
                        val_num = promotions[(route_id, month)]
                        val = f"{val_num:,.2f}"
                        w_results_data[res_name][month] += val_num
                    elif res_name == 'convenios':
                        val = agreements[(route_id, month)]
                        w_results_data[res_name][month] += val
                    
                    temp_res['monthly_data'].append({'value': val})
                temp_route['results'].append(temp_res)
            
            temp_warehouse['routes'].append(temp_route)
        
        # Build warehouse summary route
        warehouse_summary = {
            'route_id': 'TOTAL',
            'route_name': warehouse_name,
            'product_classes': [],
            'total': {'name': 'total', 'monthly_data': []},
            'results': []
        }

        for pc_name in pc_names:
            temp_pc = {'name': pc_name, 'monthly_data': []}
            for month in range(1, 13):
                t = w_pc_data[pc_name][month]['target']
                s = w_pc_data[pc_name][month]['net_sale']
                diff = s - t
                scope = (s / t * Decimal('100')) if t > 0 else (Decimal('100') if s > 0 else Decimal('0'))
                temp_pc['monthly_data'].append({
                    'target': t,
                    'net_sale': s,
                    'diff': diff,
                    'scope': scope
                })
            warehouse_summary['product_classes'].append(temp_pc)

        for month in range(1, 13):
            t = w_total_data[month]['target']
            s = w_total_data[month]['net_sale']
            diff = s - t
            scope = (s / t * Decimal('100')) if t > 0 else (Decimal('100') if s > 0 else Decimal('0'))
            warehouse_summary['total']['monthly_data'].append({
                'target': t,
                'net_sale': s,
                'diff': diff,
                'scope': scope
            })
            
        for res_name in results_names:
            temp_res = {'name': res_name, 'monthly_data': []}
            for month in range(1, 13):
                if res_name == 'margen':
                    t_profit = w_margin_acc[month]['profit']
                    t_net = w_margin_acc[month]['net']
                    if t_net > 0:
                        margin = (t_profit / t_net) * Decimal('100.00')
                    else:
                        margin = Decimal('0.00')
                    val = f"{margin:,.2f} %"
                elif res_name == 'clientes nuevos':
                    val = w_results_data[res_name][month]
                elif res_name == 'cuentas por cobrar':
                    val = w_results_data[res_name][month]
                elif res_name == 'cuentas por cobrar $':
                    val = f"$ {w_results_data[res_name][month]:,.2f}"
                elif res_name == 'promociones':
                    val = f"{w_results_data[res_name][month]:,.2f}"
                elif res_name == 'convenios':
                    val = w_results_data[res_name][month]
                temp_res['monthly_data'].append({'value': val})
            warehouse_summary['results'].append(temp_res)

        temp_warehouse['routes'].insert(0, warehouse_summary)

        return [temp_warehouse]


    def get_data_report(self):
        """
        Generate an Excel workbook with detailed monthly breakdown data.

        Key Features:
        - Creates one sheet per warehouse.
        - Summarizes target sales vs actual sales for each product class.
        - Includes route-level performance metrics.
        - Provides KPIs like new customers, accounts receivable, promotions, and agreements.

        Returns:
            openpyxl.Workbook: The generated workbook containing all the data.
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        #style config
        HEADER_FILL = PatternFill(start_color="2D9999", fill_type="solid")
        ALT_CLASS_1 = PatternFill(start_color="FFFFFF", fill_type="solid")
        ALT_CLASS_2 = PatternFill(start_color="A4E0E6", fill_type="solid")
        TOTAL_FILL = PatternFill(start_color="FAC003", fill_type="solid")
        ALT_SUM_1 = PatternFill(start_color="D9D9D9", fill_type="solid")
        ALT_SUM_2 = PatternFill(start_color="FFFFFF", fill_type="solid")

        
        WHITE_BOLD = Font(color="FFFFFF", bold=True)
        BLACK_BOLD = Font(color="000000", bold=True)
        BLACK = Font(color="000000")
        CENTER = Alignment(horizontal="center", vertical="center")
        LEFT = Alignment(horizontal="left", vertical="center")

        THIN_BORDER = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )




        def style_cell(cell, fill, font=None, alignment=None, number_format=None):
            cell.fill = fill
            cell.border = THIN_BORDER
            if font: cell.font = font
            if alignment: cell.alignment = alignment
            if number_format: cell.number_format = number_format
            return cell

        month_names = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']

        #reuse internal function
        report_data = self.get_data()

        #iteration over the gotten list
        for w_data in report_data:
            #maximum sheet title length is 31 characters
            sheet_title = (w_data['warehouse_name'] or 'Gerencia')[:31].upper()
            ws = wb.create_sheet(title=sheet_title)
            current_row = 1

            for route in w_data['routes']:
                #headers
                headers = ["AGENTE", "CLASE"]
                for m_name in month_names:
                    # Tus 4 columnas base
                    headers.extend([f"OBJETIVO {m_name}", f"VENTA {m_name}", f"ALCANCE {m_name}", f"DIF. {m_name}"])

                for col_num, val in enumerate(headers, 1):
                    c = ws.cell(row=current_row, column=col_num, value=val)
                    style_cell(c, HEADER_FILL, WHITE_BOLD, CENTER)
                current_row += 1

                #product classe rows
                row_idx = 0
                for pc in route['product_classes']:
                    fill_color = ALT_CLASS_1 if row_idx % 2 == 0 else ALT_CLASS_2

                    style_cell(ws.cell(row=current_row, column=1, value=route['route_id']), fill_color, BLACK, CENTER)
                    style_cell(ws.cell(row=current_row, column=2, value=pc['name'].upper()), fill_color, BLACK, LEFT)

                    col = 3
                    for m_data in pc['monthly_data']:
                        style_cell(ws.cell(row=current_row, column=col, value=float(m_data['target'])), fill_color, BLACK, CENTER, '"$"#,##0.00')
                        style_cell(ws.cell(row=current_row, column=col + 1, value=float(m_data['net_sale'])), fill_color, BLACK, CENTER, '"$"#,##0.00')
                        style_cell(ws.cell(row=current_row, column=col + 2, value=float(m_data['scope']) / 100), fill_color, BLACK, CENTER, '0.00%')
                        style_cell(ws.cell(row=current_row, column=col + 3, value=float(m_data['diff'])), fill_color, BLACK, CENTER, '"$"#,##0.00')
                        col += 4

                    current_row += 1
                    row_idx += 1

                #total
                style_cell(ws.cell(row=current_row, column=1, value=route['route_id']), TOTAL_FILL, BLACK_BOLD, CENTER)
                style_cell(ws.cell(row=current_row, column=2, value="TOTAL"), TOTAL_FILL, BLACK_BOLD, LEFT)

                col = 3
                for m_data in route['total']['monthly_data']:
                    style_cell(ws.cell(row=current_row, column=col, value=float(m_data['target'])), TOTAL_FILL, BLACK_BOLD, CENTER, '"$"#,##0.00')
                    style_cell(ws.cell(row=current_row, column=col + 1, value=float(m_data['net_sale'])), TOTAL_FILL, BLACK_BOLD, CENTER, '"$"#,##0.00')
                    style_cell(ws.cell(row=current_row, column=col + 2, value=float(m_data['scope']) / 100), TOTAL_FILL, BLACK_BOLD, CENTER, '0.00%')
                    style_cell(ws.cell(row=current_row, column=col + 3, value=float(m_data['diff'])), TOTAL_FILL, BLACK_BOLD, CENTER, '"$"#,##0.00')
                    col += 4
                current_row += 1

                #total rows
                res_idx = 0
                for res in route['results']:
                    fill_color = ALT_SUM_1 if res_idx % 2 == 0 else ALT_SUM_2

                    style_cell(ws.cell(row=current_row, column=1, value=route['route_id']), fill_color, BLACK_BOLD, CENTER)
                    style_cell(ws.cell(row=current_row, column=2, value=res['name'].upper()), fill_color, BLACK_BOLD, LEFT)

                    col = 3
                    for m_data in res['monthly_data']:
                        #trasnform the gotten results which have format as currency or percentage to excel recognizes them as numbers
                        val_str = str(m_data['value']).replace('$', '').replace('%', '').replace(',', '').strip()
                        try:
                            val = float(val_str)
                        except ValueError:
                            val = 0.0

                        #currency and percentage format
                        res_name = res['name'].lower()
                        num_format = '0'
                        if 'margen' in res_name:
                            num_format = '0.00%'
                            val = val / 100
                        elif '$' in res_name or 'promociones' in res_name:
                            num_format = '"$"#,##0.00'

                        #colspan 4 inthe total rows
                        ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 3)
                        
                        c_val = ws.cell(row=current_row, column=col, value=val)
                        style_cell(c_val, fill_color, BLACK_BOLD, CENTER, num_format)

                        #outline
                        style_cell(ws.cell(row=current_row, column=col + 1), fill_color)
                        style_cell(ws.cell(row=current_row, column=col + 2), fill_color)
                        style_cell(ws.cell(row=current_row, column=col + 3), fill_color)

                        col += 4

                    current_row += 1
                    res_idx += 1

                #alternate colors
                current_row += 1

            #col width
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 25
            for col_idx in range(3, 51):
                ws.column_dimensions[get_column_letter(col_idx)].width = 16

        #http response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="desglose_mensual_gerencia_{self.year}_{self.warehouse_id}.xlsx"'
        wb.save(response)

        return response


    def summary_for_assistant(self) -> dict:
        """
        Generates a dictionary for llm to analyze
        """

        raw_data = self.get_data()
        if not raw_data:
            return {}

        warehouse_info = raw_data[0]
        
        #cut on the current month if necesary to avoid sending future months with no data
        current_date = date.today()
        if self.year < current_date.year:
            max_month = 12
        elif self.year == current_date.year:
            max_month = current_date.month -1
        else:
            max_month = 0
            
        month_names = {
            1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril', 
            5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto', 
            9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
        }

        #means by warehouse
        sum_venta = 0.0
        sum_objetivo = 0.0
        sum_margen_pct = 0.0
        sum_clientes = 0
        sum_cxc = 0.0
        sum_convenios = 0
        rutas_meses_count = 0

        metricas_mensuales = {}

        for route in warehouse_info['routes']:
            route_key = f"{route['route_id']} - {route['route_name']}"
            metricas_mensuales[route_key] = {}
            res_lookup = {r['name']: r['monthly_data'] for r in route['results']}
            
            for m in range(max_month):
                m_num = m + 1
                m_name = month_names[m_num]
                
                total_data = route['total']['monthly_data'][m]
                venta = float(total_data['net_sale'])
                objetivo = float(total_data['target'])
                alcance = float(total_data['scope'])
                



                #clean formqts
                def parse_val(v):
                    if isinstance(v, str):
                        v = v.replace('$', '').replace('%', '').replace(',', '').strip()
                        return float(v) if v else 0.0
                    return float(v)
                
                margen = parse_val(res_lookup.get('margen', [{'value': 0}])[m]['value'])
                cxc = parse_val(res_lookup.get('cuentas por cobrar $', [{'value': 0}])[m]['value'])
                convenios = parse_val(res_lookup.get('convenios', [{'value': 0}])[m]['value'])
                clientes_nuevos = parse_val(res_lookup.get('clientes nuevos', [{'value': 0}])[m]['value'])
                
                sum_venta += venta
                sum_objetivo += objetivo
                sum_margen_pct += margen
                sum_cxc += cxc
                sum_convenios += convenios
                sum_clientes += clientes_nuevos
                rutas_meses_count += 1
                
                desempeño_por_clase = {}
                for pc in route['product_classes']:
                    pc_data = pc['monthly_data'][m]
                    desempeño_por_clase[pc['name'].lower()] = {
                        'objetivo': float(pc_data['target']),
                        'venta': float(pc_data['net_sale']),
                        'alcance': float(pc_data['scope'])
                    }
                
                metricas_mensuales[route_key][m_name] = {
                    'alcance': alcance,
                    'margen': margen,
                    'cuentas_por_cobrar': cxc,
                    'convenios': convenios,
                    'venta': venta,
                    'desempeño_por_clase': desempeño_por_clase
                }

        valid_months = max_month if max_month > 0 else 1
        alcance_promedio_global = (sum_venta / sum_objetivo * 100) if sum_objetivo > 0 else (100.0 if sum_venta > 0 else 0.0)

        summary = {
            'gerencia': f"{warehouse_info['warehouse_id']} - {warehouse_info['warehouse_name']}",
            'metricas_promedio_mensuales': {
                'venta_promedio': round(sum_venta / valid_months, 2),
                'alcance_promedio': round(alcance_promedio_global, 2),
                'margen_promedio': round(sum_margen_pct / rutas_meses_count, 2) if rutas_meses_count else 0.0,
                'clientes_nuevos_promedio': round(sum_clientes / valid_months, 2),
                'cuentas_por_cobrar_promedio': round(sum_cxc / valid_months, 2),
                'convenios_promedio': round(sum_convenios / valid_months, 2),
            },
            'metricas_mensuales': metricas_mensuales
        }

        return summary