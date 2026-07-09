import datetime
from decimal import Decimal
from django.db.models import Sum, Q
from apps.core.models import SaleTarget, SaleTransaction, Route, ProductClass, Customer
from dateutil.relativedelta import relativedelta

class SaleTargetsCalculatorService:
    """
    Service class to calculate sales target simulations when transferring or adjusting routes.
    """

    def __init__(self, mode, origin_route_id, destination_route_id=None, customer_ids=None, adjustment_direction='remove', transfer_growth_rule='exact'):
        """
        Initialize the calculator service with simulation parameters.

        Args:
            mode (str): Mode of the simulation, either 'transfer' (bewteen two routes) or 'adjustment' (by adding or removing from all customers).
            origin_route_id (str): ID of the origin route where the customers currently belong.
            destination_route_id (str, optional): ID of the destination route for transfers. Defaults to None.
            customer_ids (list, optional): List of customer IDs to be adjusted or transferred. Defaults to None.
            adjustment_direction (str, optional): Direction of adjustment, 'add' or 'remove'. Defaults to 'remove'.
            transfer_growth_rule (str, optional): Growth rule for transfers, 'exact' or 'dynamic'. Defaults to 'exact'.
        """
        self.mode = mode
        self.origin_route_id = origin_route_id
        self.destination_route_id = destination_route_id
        self.customer_ids = customer_ids or []
        self.adjustment_direction = adjustment_direction
        self.transfer_growth_rule = transfer_growth_rule
        self.errors = []
        
    def _parse_month(self, ym_str):
        """
        Parse a year-month string into a date object.

        Args:
            ym_str (str): The year-month string in 'YYYY-MM' format.

        Returns:
            datetime.date: Date object representing the first day of the parsed month, or None if invalid.
        """
        if not ym_str: return None
        try:
            return datetime.datetime.strptime(ym_str, '%Y-%m').date()
        except:
            return None

    def _months_diff(self, start_date, end_date):
        """
        Calculate the difference in months between two dates.

        Args:
            start_date (datetime.date): The start date.
            end_date (datetime.date): The end date.

        Returns:
            int: The difference in months, inclusive of the boundary months.
        """
        return (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month) + 1

    def calculate_simulation(self, target_year, effective_month, eval_customer_start, eval_customer_end, eval_route_start, eval_route_end, product_class_ids, calc_method):
        """
        Run the target adjustment simulation for origin and optionally destination routes.

        Args:
            target_year (int): The target calendar year for the simulation.
            effective_month (str): The month from which the adjustment is effective ('YYYY-MM').
            eval_customer_start (str): The start month of the customer evaluation period ('YYYY-MM').
            eval_customer_end (str): The end month of the customer evaluation period ('YYYY-MM').
            eval_route_start (str): The start month of the route evaluation period ('YYYY-MM').
            eval_route_end (str): The end month of the route evaluation period ('YYYY-MM').
            product_class_ids (list): List of product class IDs to consider.
            calc_method (str): The calculation method, either 'average' or 'contribution'.

        Returns:
            dict: Simulation results containing origin and destination details, and customer portfolio summary, or None.
        """
        if not self.origin_route_id or not self.customer_ids:
            self.errors.append("Falta seleccionar ruta origen o clientes.")
            return None
            
        if self.mode == 'transfer' and not self.destination_route_id:
            self.errors.append("Falta seleccionar la ruta destino.")
            return None

        eff_date = self._parse_month(effective_month)
        c_start = self._parse_month(eval_customer_start)
        c_end = self._parse_month(eval_customer_end)
        
        r_start = None
        r_end = None
        if calc_method == 'contribution':
            r_start = self._parse_month(eval_route_start)
            r_end = self._parse_month(eval_route_end)

        if not all([eff_date, c_start, c_end]):
            self.errors.append("Las fechas ingresadas no tienen un formato válido.")
            return None
            
        if calc_method == 'contribution' and not all([r_start, r_end]):
            self.errors.append("Las fechas de evaluación de ruta no tienen un formato válido.")
            return None
            
        c_end = c_end + relativedelta(day=31)
        if r_end:
            r_end = r_end + relativedelta(day=31)
            
        product_classes = ProductClass.objects.filter(id__in=product_class_ids)
        origin_route = Route.objects.filter(id=self.origin_route_id).first()
        dest_route = Route.objects.filter(id=self.destination_route_id).first() if self.mode == 'transfer' else None
        
        targets = self._get_route_targets([self.origin_route_id, self.destination_route_id] if self.mode == 'transfer' else [self.origin_route_id], target_year, product_classes)
        
        deltas = {} 
        if calc_method == 'average':
            deltas = self._calculate_average_deltas(c_start, c_end, product_classes)
        elif calc_method == 'contribution':
            deltas = self._calculate_contribution_deltas(c_start, c_end, r_start, r_end, product_classes)

        origin_targets = targets.get(self.origin_route_id, {})
        dest_targets = targets.get(self.destination_route_id, {}) if self.destination_route_id else {}
        
        def generate_computed_deltas(route_targets):
            cdeltas = {}
            for pc in product_classes:
                cdeltas[pc.id] = {}
                current_base = Decimal('0.00')
                
                for m in range(1, 13):
                    if m < eff_date.month:
                        cdeltas[pc.id][m] = Decimal('0.00')
                    else:
                        if m == eff_date.month:
                            if calc_method == 'average':
                                current_base = deltas.get(pc.id, Decimal('0.00'))
                            elif calc_method == 'contribution':
                                pct = deltas.get(pc.id, Decimal('0.00'))
                                current_base = origin_targets.get(pc.id, {}).get(m, Decimal('0.00')) * pct
                        else:
                            prev_val = route_targets.get(pc.id, {}).get(m - 1, Decimal('0.00'))
                            curr_val = route_targets.get(pc.id, {}).get(m, Decimal('0.00'))
                            
                            if prev_val > 0:
                                growth_factor = curr_val / prev_val
                                current_base = current_base * growth_factor
                            else:
                                if curr_val > 0:
                                    current_base = current_base * 2
                                    
                        cdeltas[pc.id][m] = current_base
            return cdeltas

        computed_deltas_origin = generate_computed_deltas(origin_targets)
        
        if self.mode == 'transfer' and self.transfer_growth_rule == 'dynamic':
            computed_deltas_dest = generate_computed_deltas(dest_targets)
        else:
            computed_deltas_dest = computed_deltas_origin

        months = [datetime.date(target_year, m, 1) for m in range(1, 13)]
        
        origin_result = self._build_route_result(origin_route, product_classes, origin_targets, computed_deltas_origin, months, is_origin=True)
        dest_result = None
        if self.mode == 'transfer':
            dest_result = self._build_route_result(dest_route, product_classes, dest_targets, computed_deltas_dest, months, is_origin=False)
            
        return {
            'origin': origin_result,
            'destination': dest_result,
            'customer_summary': self._calculate_customer_summary()
        }

    def _calculate_customer_summary(self):
        """
        Calculate the net changes in customer portfolio assignments for origin and destination routes.

        Returns:
            dict: Portfolio counts containing current, affected, final, and is_addition details.
        """
        origin_current = Customer.objects.filter(route_id=self.origin_route_id).count()
        customers_in_origin = Customer.objects.filter(id__in=self.customer_ids, route_id=self.origin_route_id).count()
        total_selected = len(self.customer_ids)
        
        summary = {}
        
        if self.mode == 'transfer':
            dest_current = Customer.objects.filter(route_id=self.destination_route_id).count()
            customers_in_dest = Customer.objects.filter(id__in=self.customer_ids, route_id=self.destination_route_id).count()
            
            origin_removed = customers_in_origin
            dest_added = total_selected - customers_in_dest
            
            summary['origin'] = {
                'current': origin_current,
                'affected': origin_removed,
                'is_addition': False,
                'final': origin_current - origin_removed
            }
            summary['destination'] = {
                'current': dest_current,
                'affected': dest_added,
                'is_addition': True,
                'final': dest_current + dest_added
            }
        else:
            if self.adjustment_direction == 'add':
                origin_added = total_selected - customers_in_origin
                summary['origin'] = {
                    'current': origin_current,
                    'affected': origin_added,
                    'is_addition': True,
                    'final': origin_current + origin_added
                }
            else:
                origin_removed = customers_in_origin
                summary['origin'] = {
                    'current': origin_current,
                    'affected': origin_removed,
                    'is_addition': False,
                    'final': origin_current - origin_removed
                }
                
        return summary

    def _get_route_targets(self, route_ids, target_year, product_classes):
        """
        Fetch the original target values for the given routes and year.

        Args:
            route_ids (list): List of route IDs.
            target_year (int): Calendar year.
            product_classes (QuerySet): QuerySet of ProductClass objects.

        Returns:
            dict: Nested dictionary mapping route_id -> product_class_id -> month -> target_amount.
        """
        start_y = datetime.date(target_year, 1, 1)
        end_y = datetime.date(target_year, 12, 31)
        
        qs = SaleTarget.objects.filter(
            route_id__in=[r for r in route_ids if r],
            product_class__in=product_classes,
            period__gte=start_y,
            period__lte=end_y
        )
        
        targets = {}
        for r_id in route_ids:
            if not r_id: continue
            targets[r_id] = {pc.id: {m: Decimal('0.00') for m in range(1, 13)} for pc in product_classes}
            
        for t in qs:
            targets[t.route_id][t.product_class_id][t.period.month] = t.target_amount
            
        return targets

    def _calculate_average_deltas(self, start_date, end_date, product_classes):
        """
        Calculate the average monthly sales for the selected customers during the evaluation period.

        Args:
            start_date (datetime.date): Start date of the evaluation period.
            end_date (datetime.date): End date of the evaluation period.
            product_classes (QuerySet): QuerySet of ProductClass objects.

        Returns:
            dict: Dictionary mapping product_class_id to the average monthly sales amount (Decimal).
        """
        months_count = self._months_diff(start_date, end_date)
        if months_count <= 0: months_count = 1
        
        sales = SaleTransaction.objects.filter(
            customer_id__in=self.customer_ids,
            product_class__in=product_classes,
            sale_date__gte=start_date,
            sale_date__lte=end_date
        ).values('product_class').annotate(total=Sum('net_amount'))
        
        sales_map = {item['product_class']: item['total'] for item in sales}
        
        deltas = {}
        for pc in product_classes:
            total = sales_map.get(pc.id, Decimal('0.00'))
            deltas[pc.id] = total / Decimal(months_count)
        return deltas

    def _calculate_contribution_deltas(self, c_start, c_end, r_start, r_end, product_classes):
        """
        Calculate the sales contribution ratio of the selected customers relative to the origin route.

        Args:
            c_start (datetime.date): Start date for customer sales evaluation.
            c_end (datetime.date): End date for customer sales evaluation.
            r_start (datetime.date): Start date for route sales evaluation.
            r_end (datetime.date): End date for route sales evaluation.
            product_classes (QuerySet): QuerySet of ProductClass objects.

        Returns:
            dict: Dictionary mapping product_class_id to the contribution percentage (Decimal).
        """
        c_sales = SaleTransaction.objects.filter(
            customer_id__in=self.customer_ids,
            product_class__in=product_classes,
            sale_date__gte=c_start,
            sale_date__lte=c_end
        ).values('product_class').annotate(total=Sum('net_amount'))
        c_map = {item['product_class']: item['total'] for item in c_sales}
        
        r_sales = SaleTransaction.objects.filter(
            route_id=self.origin_route_id,
            product_class__in=product_classes,
            sale_date__gte=r_start,
            sale_date__lte=r_end
        ).values('product_class').annotate(total=Sum('net_amount'))
        r_map = {item['product_class']: item['total'] for item in r_sales}
        
        deltas = {}
        for pc in product_classes:
            c_val = c_map.get(pc.id, Decimal('0.00'))
            r_val = r_map.get(pc.id, Decimal('0.00'))
            if r_val > 0:
                deltas[pc.id] = c_val / r_val
            else:
                deltas[pc.id] = Decimal('0.00')
        return deltas

    def _build_route_result(self, route, product_classes, targets, computed_deltas, months, is_origin):
        """
        Build the simulation targets breakdown and totals for a single route.

        Args:
            route (Route): The Route model instance.
            product_classes (QuerySet): QuerySet of ProductClass objects.
            targets (dict): Nested dict containing original target amounts.
            computed_deltas (dict): Nested dict containing calculated adjustments per month/class.
            months (list): List of datetime.date objects for each month of the target year.
            is_origin (bool): True if compiling the origin route results, False otherwise.

        Returns:
            dict: Structured dictionary with route targets, monthly/annual totals, and growth percentages.
        """
        result = {
            'route_name': f"{route.id.upper()} {route.name.title()}",
            'classes': [],
            'month_totals': [{'date': m, 'old_target': Decimal('0.00'), 'delta': Decimal('0.00'), 'new_target': Decimal('0.00')} for m in months],
            'grand_total': {'old_target': Decimal('0.00'), 'delta': Decimal('0.00'), 'new_target': Decimal('0.00')}
        }
        
        if self.mode == 'transfer':
            sign = -1 if is_origin else 1
        else:
            sign = 1 if self.adjustment_direction == 'add' else -1
        
        for pc in product_classes:
            pc_data = {
                'class_name': pc.name.title(),
                'months': [],
                'totals': {'old_target': Decimal('0.00'), 'delta': Decimal('0.00'), 'new_target': Decimal('0.00')}
            }
            
            for idx, m in enumerate(months):
                old_target = targets.get(pc.id, {}).get(m.month, Decimal('0.00'))
                
                growth = Decimal('0.00')
                if m.month > 1:
                    prev = targets.get(pc.id, {}).get(m.month - 1, Decimal('0.00'))
                    if prev > 0:
                        growth = ((old_target - prev) / prev) * 100
                    elif prev == 0 and old_target > 0:
                        growth = Decimal('100.00')
                        
                delta_val = computed_deltas.get(pc.id, {}).get(m.month, Decimal('0.00')) * sign
                
                new_target = old_target + delta_val
                if new_target < 0: new_target = Decimal('0.00')
                
                pc_data['totals']['old_target'] += old_target
                pc_data['totals']['delta'] += delta_val
                pc_data['totals']['new_target'] += new_target
                
                result['month_totals'][idx]['old_target'] += old_target
                result['month_totals'][idx]['delta'] += delta_val
                result['month_totals'][idx]['new_target'] += new_target
                
                result['grand_total']['old_target'] += old_target
                result['grand_total']['delta'] += delta_val
                result['grand_total']['new_target'] += new_target
                
                pc_data['months'].append({
                    'date': m,
                    'old_target': old_target,
                    'growth': growth,
                    'delta': delta_val,
                    'new_target': new_target
                })
            result['classes'].append(pc_data)
            
        for idx, mt in enumerate(result['month_totals']):
            if mt['date'].month > 1:
                prev = result['month_totals'][idx - 1]['old_target']
                if prev > 0:
                    mt['growth'] = ((mt['old_target'] - prev) / prev) * 100
                elif prev == 0 and mt['old_target'] > 0:
                    mt['growth'] = Decimal('100.00')
                else:
                    mt['growth'] = Decimal('0.00')
            else:
                mt['growth'] = Decimal('0.00')
            
        return result

    def export_data_report(self, results):
        """
        Generate a multi-sheet Excel report of the simulation targets and portfolio changes.

        Args:
            results (dict): The dictionary compiled by calculate_simulation.

        Returns:
            bytes: The binary content of the generated .xlsx workbook, or None.
        """
        if not results:
            return None
            
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        title_font = Font(bold=True, size=14)
        
        for key, r in results.items():
            if key == 'customer_summary':
                continue
            if not r:
                continue
                
            route_name = r['route_name'].replace('/', '-')
            sheet_title = f"{'Origen' if key == 'origin' else 'Destino'} - {route_name}"
            sheet_title = sheet_title.translate(str.maketrans('', '', '\\/*?:[]'))[:31]
            ws = wb.create_sheet(title=sheet_title)
            
            ws.append([f"Cálculo de Cuotas de Venta: {r['route_name']}"])
            ws.cell(row=ws.max_row, column=1).font = title_font
            ws.append([])
            
            summary = results.get('customer_summary', {}).get(key, {})
            if summary:
                ws.append(["Resumen de cartera de clientes"])
                ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
                ws.append(["Cartera actual:", f"{summary.get('current')} clientes"])
                
                affected_label = "Nuevos clientes a integrar:" if summary.get('is_addition') else "Clientes a transferir/remover:"
                affected_sign = "+" if summary.get('is_addition') else "-"
                ws.append([affected_label, f"{affected_sign}{summary.get('affected')} clientes"])
                
                ws.append(["Cartera proyectada:", f"{summary.get('final')} clientes"])
                ws.append([])
                
            ws.append([f"Desglose de cálculo (Objetivo actual, Crecimiento, Ajuste)"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
            ws.append([])
            
            if not r['classes']:
                ws.append(["No hay datos"])
                continue
                
            months = r['classes'][0]['months']
            
            header_row1 = ws.max_row + 1
            ws.cell(row=header_row1, column=1, value="Clase de producto").font = header_font
            ws.cell(row=header_row1, column=1).fill = header_fill
            ws.cell(row=header_row1, column=1).alignment = header_alignment
            ws.cell(row=header_row1, column=1).border = thin_border
            ws.merge_cells(start_row=header_row1, start_column=1, end_row=header_row1+1, end_column=1)
            
            col_idx = 2
            for m in months:
                month_name = m['date'].strftime('%b %Y').title()
                ws.cell(row=header_row1, column=col_idx, value=month_name).font = header_font
                ws.cell(row=header_row1, column=col_idx).fill = header_fill
                ws.cell(row=header_row1, column=col_idx).alignment = header_alignment
                
                for c in range(col_idx, col_idx + 3):
                    ws.cell(row=header_row1, column=c).border = thin_border
                    ws.cell(row=header_row1, column=c).fill = header_fill
                
                ws.merge_cells(start_row=header_row1, start_column=col_idx, end_row=header_row1, end_column=col_idx+2)
                col_idx += 3
                
            ws.cell(row=header_row1, column=col_idx, value="Total Anual").font = header_font
            ws.cell(row=header_row1, column=col_idx).fill = header_fill
            ws.cell(row=header_row1, column=col_idx).alignment = header_alignment
            for c in range(col_idx, col_idx + 3):
                ws.cell(row=header_row1, column=c).border = thin_border
                ws.cell(row=header_row1, column=c).fill = header_fill
            ws.merge_cells(start_row=header_row1, start_column=col_idx, end_row=header_row1, end_column=col_idx+2)
                
            header_row2 = header_row1 + 1
            col_idx = 2
            sub_headers = ["Obj. original", "Crecimiento", "Ajuste"]
            for m in months:
                for sub in sub_headers:
                    cell = ws.cell(row=header_row2, column=col_idx, value=sub)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    col_idx += 1
            for sub in sub_headers:
                cell = ws.cell(row=header_row2, column=col_idx, value=sub)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                col_idx += 1
                    
            for cls in r['classes']:
                current_row = ws.max_row + 1
                first_cell = ws.cell(row=current_row, column=1, value=cls['class_name'])
                first_cell.font = Font(bold=True)
                first_cell.border = thin_border
                
                col_idx = 2
                for m in cls['months']:
                    cell1 = ws.cell(row=current_row, column=col_idx, value=m['old_target'])
                    cell1.border = thin_border
                    cell1.number_format = '"$"#,##0.00'
                    col_idx += 1
                    
                    growth_val = float(m['growth']) / 100 if m['growth'] else 0.0
                    cell2 = ws.cell(row=current_row, column=col_idx, value=growth_val)
                    cell2.border = thin_border
                    cell2.number_format = '0.0%'
                    if growth_val > 0: cell2.font = Font(color="008000")
                    elif growth_val < 0: cell2.font = Font(color="FF0000")
                    col_idx += 1
                    
                    delta_val = m['delta']
                    cell3 = ws.cell(row=current_row, column=col_idx, value=delta_val)
                    cell3.border = thin_border
                    cell3.number_format = '"$"#,##0.00'
                    if delta_val > 0: cell3.font = Font(color="008000")
                    elif delta_val < 0: cell3.font = Font(color="FF0000")
                    col_idx += 1
                    
                t = cls['totals']
                cell1 = ws.cell(row=current_row, column=col_idx, value=t['old_target'])
                cell1.border = thin_border
                cell1.font = Font(bold=True)
                cell1.number_format = '"$"#,##0.00'
                col_idx += 1
                
                cell2 = ws.cell(row=current_row, column=col_idx, value="-")
                cell2.border = thin_border
                cell2.alignment = Alignment(horizontal="center")
                col_idx += 1
                
                cell3 = ws.cell(row=current_row, column=col_idx, value=t['delta'])
                cell3.border = thin_border
                cell3.font = Font(bold=True)
                cell3.number_format = '"$"#,##0.00'
                if t['delta'] > 0: cell3.font = Font(bold=True, color="008000")
                elif t['delta'] < 0: cell3.font = Font(bold=True, color="FF0000")
                col_idx += 1
                
            current_row = ws.max_row + 1
            cell = ws.cell(row=current_row, column=1, value="TOTAL RUTA")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            
            col_idx = 2
            for mt in r['month_totals']:
                cell1 = ws.cell(row=current_row, column=col_idx, value=mt['old_target'])
                cell1.border = thin_border
                cell1.font = Font(bold=True)
                cell1.number_format = '"$"#,##0.00'
                cell1.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                col_idx += 1
                
                growth_val = float(mt['growth']) / 100 if mt['growth'] else 0.0
                cell2 = ws.cell(row=current_row, column=col_idx, value=growth_val)
                cell2.border = thin_border
                cell2.number_format = '0.0%'
                cell2.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                if growth_val > 0: cell2.font = Font(bold=True, color="008000")
                elif growth_val < 0: cell2.font = Font(bold=True, color="FF0000")
                col_idx += 1
                
                delta_val = mt['delta']
                cell3 = ws.cell(row=current_row, column=col_idx, value=delta_val)
                cell3.border = thin_border
                cell3.number_format = '"$"#,##0.00'
                cell3.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                if delta_val > 0: cell3.font = Font(bold=True, color="008000")
                elif delta_val < 0: cell3.font = Font(bold=True, color="FF0000")
                col_idx += 1
                
            gt = r['grand_total']
            cell1 = ws.cell(row=current_row, column=col_idx, value=gt['old_target'])
            cell1.border = thin_border
            cell1.font = Font(bold=True)
            cell1.number_format = '"$"#,##0.00'
            cell1.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
            col_idx += 1
            
            cell2 = ws.cell(row=current_row, column=col_idx, value="-")
            cell2.border = thin_border
            cell2.alignment = Alignment(horizontal="center")
            cell2.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
            col_idx += 1
            
            cell3 = ws.cell(row=current_row, column=col_idx, value=gt['delta'])
            cell3.border = thin_border
            cell3.font = Font(bold=True)
            cell3.number_format = '"$"#,##0.00'
            cell3.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
            if gt['delta'] > 0: cell3.font = Font(bold=True, color="008000")
            elif gt['delta'] < 0: cell3.font = Font(bold=True, color="FF0000")
            col_idx += 1
                    
            ws.append([])
            ws.append([])
            ws.append([f"Objetivos Planificados (Objetivos finales): {r['route_name']}"])
            ws.cell(row=ws.max_row, column=1).font = title_font
            ws.append([])
            
            header_row3 = ws.max_row + 1
            ws.cell(row=header_row3, column=1, value="Clase de producto").font = header_font
            ws.cell(row=header_row3, column=1).fill = header_fill
            ws.cell(row=header_row3, column=1).alignment = header_alignment
            ws.cell(row=header_row3, column=1).border = thin_border
            
            col_idx = 2
            for m in months:
                month_name = m['date'].strftime('%b %Y').title()
                cell = ws.cell(row=header_row3, column=col_idx, value=month_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                col_idx += 1
                
            cell = ws.cell(row=header_row3, column=col_idx, value="Total Anual")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            col_idx += 1
                
            for cls in r['classes']:
                current_row = ws.max_row + 1
                first_cell = ws.cell(row=current_row, column=1, value=cls['class_name'])
                first_cell.font = Font(bold=True)
                first_cell.border = thin_border
                
                col_idx = 2
                for m in cls['months']:
                    cell = ws.cell(row=current_row, column=col_idx, value=m['new_target'])
                    cell.border = thin_border
                    cell.number_format = '"$"#,##0.00'
                    if m['new_target'] > m['old_target']:
                        cell.font = Font(color="008000", bold=True)
                    elif m['new_target'] < m['old_target']:
                        cell.font = Font(color="FF0000", bold=True)
                    col_idx += 1
                    
                t = cls['totals']
                cell = ws.cell(row=current_row, column=col_idx, value=t['new_target'])
                cell.border = thin_border
                cell.font = Font(bold=True)
                cell.number_format = '"$"#,##0.00'
                if t['new_target'] > t['old_target']: cell.font = Font(bold=True, color="008000")
                elif t['new_target'] < t['old_target']: cell.font = Font(bold=True, color="FF0000")
                col_idx += 1
                
            current_row = ws.max_row + 1
            cell = ws.cell(row=current_row, column=1, value="TOTAL RUTA")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            
            col_idx = 2
            for mt in r['month_totals']:
                cell1 = ws.cell(row=current_row, column=col_idx, value=mt['new_target'])
                cell1.border = thin_border
                cell1.font = Font(bold=True)
                cell1.number_format = '"$"#,##0.00'
                cell1.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                if mt['new_target'] > mt['old_target']: cell1.font = Font(bold=True, color="008000")
                elif mt['new_target'] < mt['old_target']: cell1.font = Font(bold=True, color="FF0000")
                col_idx += 1
                
            gt = r['grand_total']
            cell1 = ws.cell(row=current_row, column=col_idx, value=gt['new_target'])
            cell1.border = thin_border
            cell1.font = Font(bold=True)
            cell1.number_format = '"$"#,##0.00'
            cell1.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
            if gt['new_target'] > gt['old_target']: cell1.font = Font(bold=True, color="008000")
            elif gt['new_target'] < gt['old_target']: cell1.font = Font(bold=True, color="FF0000")
            col_idx += 1
            
            from apps.core.models import Customer
            
            ws.append([])
            ws.append([])
            ws.append(["Clientes actuales en la ruta"])
            ws.cell(row=ws.max_row, column=1).font = title_font
            ws.append([])
            
            header_row_curr = ws.max_row + 1
            headers_curr = ["ID Cliente", "Nombre", "Tipo de Cliente", "Límite de Crédito"]
            for c_idx, h in enumerate(headers_curr, 1):
                cell = ws.cell(row=header_row_curr, column=c_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                
            route_id = self.origin_route_id if key == 'origin' else self.destination_route_id
            current_customers = Customer.objects.select_related('customer_type').filter(route_id=route_id).order_by('name')
            
            if not current_customers:
                ws.append(["Sin clientes asignados"])
            else:
                for cust in current_customers:
                    r_idx = ws.max_row + 1
                    cell1 = ws.cell(row=r_idx, column=1, value=cust.id)
                    cell2 = ws.cell(row=r_idx, column=2, value=cust.name)
                    cell3 = ws.cell(row=r_idx, column=3, value=cust.customer_type.name if cust.customer_type else "")
                    cell4 = ws.cell(row=r_idx, column=4, value=cust.credit_limit)
                    cell4.number_format = '"$"#,##0.00'
                    for c in [cell1, cell2, cell3, cell4]:
                        c.border = thin_border
                        
            ws.append([])
            ws.append([])
            ws.append(["Clientes considerados en el cálculo (Seleccionados)"])
            ws.cell(row=ws.max_row, column=1).font = title_font
            ws.append([])
            
            header_row_cons = ws.max_row + 1
            headers_cons = ["ID Cliente", "Nombre", "Ruta Actual", "Tipo de Cliente"]
            for c_idx, h in enumerate(headers_cons, 1):
                cell = ws.cell(row=header_row_cons, column=c_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                
            considered_customers = Customer.objects.select_related('customer_type', 'route').filter(id__in=self.customer_ids).order_by('name')
            
            if not considered_customers:
                ws.append(["No se seleccionaron clientes"])
            else:
                for cust in considered_customers:
                    r_idx = ws.max_row + 1
                    cell1 = ws.cell(row=r_idx, column=1, value=cust.id)
                    cell2 = ws.cell(row=r_idx, column=2, value=cust.name)
                    cell3 = ws.cell(row=r_idx, column=3, value=cust.route.id.upper() if cust.route else "Sin ruta")
                    cell4 = ws.cell(row=r_idx, column=4, value=cust.customer_type.name if cust.customer_type else "")
                    for c in [cell1, cell2, cell3, cell4]:
                        c.border = thin_border
                    
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 35
            for c in range(3, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(c)].width = 18
                
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
