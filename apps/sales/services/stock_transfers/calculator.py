import datetime
from decimal import Decimal
from django.db.models import Sum
from apps.core.models import Product, Stock, SaleTransaction
from dateutil.relativedelta import relativedelta

class StockTransferCalculatorService:
    """
    Service class to calculate stock transfer suggestions between warehouses based on past sales.
    """

    def __init__(self, origin_warehouse_id, destination_warehouse_id, product_class_ids=None, rotation_level_ids=None):
        """
        Initialize the calculator service.

        Args:
            origin_warehouse_id (str): ID of the origin warehouse.
            destination_warehouse_id (str): ID of the destination warehouse.
            product_class_ids (list, optional): List of product class IDs to filter. Defaults to None.
            rotation_level_ids (list, optional): List of rotation level values (1, 2, 3). Defaults to None.
        """
        self.origin_warehouse_id = origin_warehouse_id
        self.destination_warehouse_id = destination_warehouse_id
        self.product_class_ids = product_class_ids or []
        self.rotation_level_ids = rotation_level_ids or []
        self.errors = []
        
    def _parse_month(self, ym_str):
        """
        Parse a year-month string into a date object.
        """
        if not ym_str: return None
        try:
            return datetime.datetime.strptime(ym_str, '%Y-%m').date()
        except:
            return None

    def _months_diff(self, start_date, end_date):
        """
        Calculate the difference in months between two dates.
        """
        return (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month) + 1

    def calculate_transfer(self, start_date_str, end_date_str):
        """
        Calculate the stock transfer metrics based on historical sales in the destination warehouse.

        Args:
            start_date_str (str): The start month of the evaluation ('YYYY-MM').
            end_date_str (str): The end month of the evaluation ('YYYY-MM').

        Returns:
            list: A list of dictionaries grouped by product class with calculated metrics.
        """
        if not self.origin_warehouse_id or not self.destination_warehouse_id:
            self.errors.append("Falta seleccionar almacén de origen o destino.")
            return None
            
        start_date = self._parse_month(start_date_str)
        end_date = self._parse_month(end_date_str)
        
        if not start_date or not end_date:
            self.errors.append("Las fechas de evaluación no tienen un formato válido.")
            return None
            
        end_date = end_date + relativedelta(day=31)
        
        months_count = self._months_diff(start_date, end_date)
        if months_count <= 0: months_count = 1
        
        # Filter products by classes if provided
        products_qs = Product.objects.select_related('product_class')
        # Clean product_class_ids list, removing empty strings
        clean_class_ids = [pid for pid in self.product_class_ids if pid]
        if clean_class_ids:
            products_qs = products_qs.filter(product_class_id__in=clean_class_ids)
            
        products = list(products_qs.order_by('name'))
        product_ids = [p.id for p in products]
        
        if not product_ids:
            return []
            
        # Query total sales quantity for destination warehouse in the period
        sales = SaleTransaction.objects.filter(
            warehouse_id=self.destination_warehouse_id,
            product_id__in=product_ids,
            sale_date__gte=start_date,
            sale_date__lte=end_date
        ).values('product_id').annotate(total_qty=Sum('quantity'))
        
        sales_map = {item['product_id']: item['total_qty'] or Decimal('0.00') for item in sales}
        
        # Query current stock in destination warehouse
        destination_stocks = Stock.objects.filter(
            warehouse_id=self.destination_warehouse_id,
            product_id__in=product_ids
        )
        stock_map = {stock.product_id: stock.quantity or Decimal('0.00') for stock in destination_stocks}
        
        grouped_results = {}
        
        # Clean rotation_level_ids
        valid_rotations = [str(r) for r in self.rotation_level_ids if r]
        
        for p in products:
            sold_qty = sales_map.get(p.id, Decimal('0.00'))
            avg_monthly = sold_qty / Decimal(months_count)
            current_stock = stock_map.get(p.id, Decimal('0.00'))
            
            # Since there's no explicitly implemented 'in_transit' field in the model currently, 
            # we will set it to 0 as a placeholder for the UI calculation.
            in_transit = Decimal('0.00') 
            
            # Optionally filter out products with no activity/stock to avoid clutter
            if sold_qty == 0 and current_stock == 0:
                continue
                
            # Rotation calculation
            if sold_qty > 50:
                rotation_level = '1'
                rotation_name = 'Alta'
            elif sold_qty >= 10:
                rotation_level = '2'
                rotation_name = 'Media'
            else:
                rotation_level = '3'
                rotation_name = 'Baja'
                
            # Filter by rotation if requested
            if valid_rotations and rotation_level not in valid_rotations:
                continue
                
            class_id = p.product_class_id if p.product_class_id else 'Sin Clase'
            class_name = p.product_class.name.title() if p.product_class else 'Sin Clase'
            
            if class_id not in grouped_results:
                grouped_results[class_id] = {
                    'class_name': class_name,
                    'products': []
                }
                
            grouped_results[class_id]['products'].append({
                'product_id': p.id,
                'product_name': (p.name or "").title(),
                'sold_qty': round(sold_qty, 2),
                'avg_monthly': round(avg_monthly, 2),
                'current_stock': round(current_stock, 2),
                'in_transit': round(in_transit, 2),
                'rotation_name': rotation_name,
                'rotation_level': rotation_level
            })
            
        final_results = []
        for class_id, data in grouped_results.items():
            # Sort products by sold_qty descending
            data['products'].sort(key=lambda x: x['sold_qty'], reverse=True)
            final_results.append(data)
            
        # Sort classes by name
        final_results.sort(key=lambda x: x['class_name'])
            
        return final_results

    def export_data_report(self, results, start_date_str, end_date_str, origin_name, destination_name, coverages=None):
        """
        Generate a multi-sheet or formatted Excel report of the stock transfers simulation.
        """
        if not results:
            return None
            
        if coverages is None:
            coverages = {}
            
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io
        from datetime import datetime
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transferencias de Stock"
        
        header_font = Font(bold=True, color="FFFFFF")
        dark_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid") # dark gray
        title_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        period_str = f"{start_date_str} a {end_date_str}"
        
        headers = ["ID producto", "Producto", "Unidades vendidas", "Promedio mensual", "Existencias", "En tránsito", "Cobertura (%)", "Transferencia"]
        
        # Escribir header global solo una vez
        ws.append([f"Periodo Evaluado para reposición: {period_str}"])
        ws.cell(row=ws.max_row, column=1).font = title_font
        
        ws.append([f"CEDIS Origen: {origin_name} | CEDIS Destino: {destination_name}"])
        ws.append([f"Fecha de cálculo: {current_time}"])
        ws.append([])
        
        for group in results:
            ws.append(["Clase de producto", group['class_name']])
            ws.cell(row=ws.max_row, column=1).font = bold_font
            ws.cell(row=ws.max_row, column=2).font = bold_font
            ws.append([])
            
            # Group products by rotation
            rotations = {'Alta': [], 'Media': [], 'Baja': []}
            for p in group['products']:
                if p['rotation_name'] in rotations:
                    rotations[p['rotation_name']].append(p)
                    
            for rot_name, prods in rotations.items():
                if not prods:
                    continue
                    
                # Rotation Header
                ws.append([f"{rot_name} rotación"])
                ws.cell(row=ws.max_row, column=1).font = header_font
                ws.cell(row=ws.max_row, column=1).fill = dark_fill
                
                # Table Headers
                ws.append(headers)
                header_row = ws.max_row
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=header_row, column=col_num)
                    cell.font = header_font
                    cell.fill = dark_fill
                    cell.border = thin_border
                    
                # Data Rows
                for p in prods:
                    prod_id_str = str(p['product_id'])
                    cov = coverages.get(prod_id_str, 1.0)
                    
                    row_data = [
                        p['product_id'],
                        p['product_name'],
                        p['sold_qty'],
                        p['avg_monthly'],
                        p['current_stock'],
                        p['in_transit'],
                        cov, # Custom individual coverage
                        0    # Transferencia (calculated by formula below)
                    ]
                    ws.append(row_data)
                    data_row = ws.max_row
                    
                    # Apply formula for Transferencia
                    # Formula: =MAX((Promedio*Cobertura)-Existencias-Transito, 0)
                    # Promedio is Col D, Cobertura is Col G, Existencias is Col E, Transito is Col F
                    formula = f"=MAX((D{data_row}*G{data_row})-E{data_row}-F{data_row}, 0)"
                    ws.cell(row=data_row, column=8, value=formula)
                    
                    for col_num in range(1, len(headers) + 1):
                        ws.cell(row=data_row, column=col_num).border = thin_border
                        
                ws.append([])
            
            ws.append([]) # Extra space between classes
            
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        for col_letter in ['C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col_letter].width = 18
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
