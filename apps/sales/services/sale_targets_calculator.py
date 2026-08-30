import io
import calendar
import datetime
from decimal import Decimal
from dataclasses import dataclass, field
from typing import ClassVar, Optional, Any
from collections import defaultdict

from django.utils import timezone
from django.db.models import Sum, Q, QuerySet, Prefetch
from django.core.exceptions import ValidationError

from apps.core.services.users import UsersService
from apps.sales.models import SaleTarget, SaleTransaction, Route, RouteAssignment
from apps.customers.models import Customer, CustomerAssignment
from apps.customers.services import CustomersService
from apps.products.models import ProductClass
from apps.sales.services.routes import RoutesService


class ServiceError(Exception):
    pass

class PermissionsError(ServiceError):
    pass

class TargetCalculatorError(ServiceError):
    pass

class BaseTargetCalculationStrategy:
    """
    abstract base strategy for target adjustment calculations
    """
    name: str = 'base'

    def calculate_deltas(
        self,
        *,
        product_classes: QuerySet,
        customer_ids: list[str],
        origin_route_id: str,
        destination_route_id: Optional[str],
        eff_date: datetime.date,
        target_year: int,
        origin_targets: dict,
        dest_targets: dict,
        eval_params: dict,
    ) -> tuple[dict, dict]:
        """
        calculates computed deltas for origin and destination routes.
        returns (computed_deltas_origin, computed_deltas_dest).
        each dict has structure: {product_class_id: {month_num (1..12): Decimal}}
        """
        raise NotImplementedError


class AverageMonthlyStrategy(BaseTargetCalculationStrategy):
    """
    calculates deltas based on the average monthly sales of the selected customers
    """
    name: str = 'average'

    def calculate_deltas(
        self,
        *,
        product_classes: QuerySet,
        customer_ids: list[str],
        origin_route_id: str,
        destination_route_id: Optional[str],
        eff_date: datetime.date,
        target_year: int,
        origin_targets: dict,
        dest_targets: dict,
        eval_params: dict,
    ) -> tuple[dict, dict]:
        c_start = eval_params.get('c_start')
        c_end = eval_params.get('c_end')
        transfer_growth_rule = eval_params.get('transfer_growth_rule', 'exact')
        mode = eval_params.get('mode', 'transfer')

        months_count = (c_end.year - c_start.year) * 12 + (c_end.month - c_start.month) + 1
        if months_count <= 0:
            months_count = 1

        sales_qs = SaleTransaction.objects.filter(
            customer_id__in=customer_ids,
            product_class__in=product_classes,
            sale_date__gte=c_start,
            sale_date__lte=c_end
        ).values('product_class_id').annotate(total=Sum('net_amount'))

        sales_map = {item['product_class_id']: item['total'] for item in sales_qs}

        base_deltas = {}
        for pc in product_classes:
            total_sales = sales_map.get(pc.id, Decimal('0.00')) or Decimal('0.00')
            base_deltas[pc.id] = total_sales / Decimal(months_count)

        def _generate_monthly_deltas(route_targets: dict, use_origin_base_pct: bool = False):
            cdeltas = {}
            for pc in product_classes:
                cdeltas[pc.id] = {}
                current_base = Decimal('0.00')

                for m in range(1, 13):
                    if m < eff_date.month:
                        cdeltas[pc.id][m] = Decimal('0.00')
                    elif m == eff_date.month:
                        current_base = base_deltas.get(pc.id, Decimal('0.00'))
                        cdeltas[pc.id][m] = current_base
                    else:
                        prev_val = route_targets.get(pc.id, {}).get(m - 1, Decimal('0.00'))
                        curr_val = route_targets.get(pc.id, {}).get(m, Decimal('0.00'))

                        if prev_val > 0:
                            growth_factor = curr_val / prev_val
                            current_base = current_base * growth_factor
                        else:
                            pass
                        cdeltas[pc.id][m] = current_base
            return cdeltas

        computed_origin = _generate_monthly_deltas(origin_targets)

        if mode == 'transfer' and transfer_growth_rule == 'dynamic' and dest_targets:
            computed_dest = _generate_monthly_deltas(dest_targets)
        else:
            computed_dest = computed_origin

        return computed_origin, computed_dest


class ContributionStrategy(BaseTargetCalculationStrategy):
    """
    calculates deltas based on the percentage contribution of customer sales relative to route sales
    """
    name: str = 'contribution'

    def calculate_deltas(
        self,
        *,
        product_classes: QuerySet,
        customer_ids: list[str],
        origin_route_id: str,
        destination_route_id: Optional[str],
        eff_date: datetime.date,
        target_year: int,
        origin_targets: dict,
        dest_targets: dict,
        eval_params: dict,
    ) -> tuple[dict, dict]:
        c_start = eval_params.get('c_start')
        c_end = eval_params.get('c_end')
        r_start = eval_params.get('r_start')
        r_end = eval_params.get('r_end')
        transfer_growth_rule = eval_params.get('transfer_growth_rule', 'exact')
        mode = eval_params.get('mode', 'transfer')

        # customer sales in customer evaluation window
        c_sales_qs = SaleTransaction.objects.filter(
            customer_id__in=customer_ids,
            product_class__in=product_classes,
            sale_date__gte=c_start,
            sale_date__lte=c_end
        ).values('product_class_id').annotate(total=Sum('net_amount'))
        c_map = {item['product_class_id']: item['total'] for item in c_sales_qs}

        # origin route sales in route evaluation window
        r_sales_qs = SaleTransaction.objects.filter(
            route_id=origin_route_id,
            product_class__in=product_classes,
            sale_date__gte=r_start,
            sale_date__lte=r_end
        ).values('product_class_id').annotate(total=Sum('net_amount'))
        r_map = {item['product_class_id']: item['total'] for item in r_sales_qs}

        contribution_pcts = {}
        for pc in product_classes:
            c_val = c_map.get(pc.id, Decimal('0.00')) or Decimal('0.00')
            r_val = r_map.get(pc.id, Decimal('0.00')) or Decimal('0.00')
            if r_val > 0:
                contribution_pcts[pc.id] = c_val / r_val
            else:
                contribution_pcts[pc.id] = Decimal('0.00')

        def _generate_monthly_deltas(route_targets: dict):
            cdeltas = {}
            for pc in product_classes:
                cdeltas[pc.id] = {}
                current_base = Decimal('0.00')
                pct = contribution_pcts.get(pc.id, Decimal('0.00'))

                for m in range(1, 13):
                    if m < eff_date.month:
                        cdeltas[pc.id][m] = Decimal('0.00')
                    elif m == eff_date.month:
                        eff_target = origin_targets.get(pc.id, {}).get(m, Decimal('0.00'))
                        current_base = eff_target * pct
                        cdeltas[pc.id][m] = current_base
                    else:
                        prev_val = route_targets.get(pc.id, {}).get(m - 1, Decimal('0.00'))
                        curr_val = route_targets.get(pc.id, {}).get(m, Decimal('0.00'))

                        if prev_val > 0:
                            growth_factor = curr_val / prev_val
                            current_base = current_base * growth_factor
                        else:
                            pass
                        cdeltas[pc.id][m] = current_base
            return cdeltas

        computed_origin = _generate_monthly_deltas(origin_targets)

        if mode == 'transfer' and transfer_growth_rule == 'dynamic' and dest_targets:
            computed_dest = _generate_monthly_deltas(dest_targets)
        else:
            computed_dest = computed_origin

        return computed_origin, computed_dest


# strategy registry for extensible plugandplay calculations
CALCULATION_STRATEGIES: dict[str, type[BaseTargetCalculationStrategy]] = {
    'average': AverageMonthlyStrategy,
    'contribution': ContributionStrategy,
}

@dataclass
class SaleTargetCalculatorService(UsersService):
    """
    enterprise calculator service for sales target simulations
    enforces route and customer permissions via usersservice / routeservice.
    """
    ACCESS_CONTEXTS: ClassVar[tuple[str, ...]] = (
        'acceso_total_ventas',
        'ventas',
        'acceso_total_rutas',
        'acceso_total_cuotas',
        'cuotas',
        'acceso_total',
    )

    def get_allowed_routes_qs(self) -> QuerySet:
        """
        returns active routes the user is allowed to view/simulate.
        If full access: returns all active routes.
        Otherwise: returns routes assigned to the user.
        """
        routes_service = RoutesService(user=self.user)
        return routes_service.get_allowed_routes(can_view=True, can_edit=False).filter(is_active=True).order_by('id')

    def get_route_customers(self, route_id: str, filter_type: str = 'assigned') -> QuerySet:
        """
        returns customers for the given route based on active assignments.
            'assigned': customers with active assignment on route_id (start_date <= today, end_date is null or > today).
            'all': all customers permitted to the user.
        """
        today = timezone.localdate()
        customers_service = CustomersService(user=self.user)
        base_qs = customers_service.read_customers()

        if filter_type == 'all':
            return base_qs.prefetch_related(
                Prefetch(
                    'assignments',
                    queryset=CustomerAssignment.objects.filter(
                        Q(end_date__isnull=True) | Q(end_date__gt=today),
                        start_date__lte=today
                    ).select_related('route'),
                    to_attr='active_assignments'
                )
            ).order_by('name', 'id')

        # filter customers actively assigned to the specific route
        return base_qs.filter(
            assignments__route_id=route_id,
            assignments__start_date__lte=today
        ).filter(
            Q(assignments__end_date__isnull=True) | Q(assignments__end_date__gt=today)
        ).distinct().order_by('name', 'id')

    def parse_month_bounds(self, ym_str: str) -> tuple[Optional[datetime.date], Optional[datetime.date]]:
        """
        parse 'YYYY-MM' string into exact (first_day, last_day) dates using calendar module.
        """
        if not ym_str:
            return None, None
        try:
            dt = datetime.datetime.strptime(ym_str.strip(), '%Y-%m').date()
            first_day = datetime.date(dt.year, dt.month, 1)
            _, last_day_num = calendar.monthrange(dt.year, dt.month)
            last_day = datetime.date(dt.year, dt.month, last_day_num)
            return first_day, last_day
        except (ValueError, TypeError):
            return None, None

    def get_route_targets(self, route_ids: list[str], target_year: int, product_classes: QuerySet) -> dict:
        """
        retrieves existing target amounts for given routes and year
        returns: {route_id: {product_class_id: {month: Decimal}}}
        """
        start_y = datetime.date(target_year, 1, 1)
        end_y = datetime.date(target_year, 12, 31)

        valid_route_ids = [r for r in route_ids if r]
        qs = SaleTarget.objects.filter(
            route_id__in=valid_route_ids,
            product_class__in=product_classes,
            period__gte=start_y,
            period__lte=end_y
        )

        targets = {}
        for r_id in valid_route_ids:
            targets[r_id] = {pc.id: {m: Decimal('0.00') for m in range(1, 13)} for pc in product_classes}

        for t in qs:
            m = t.period.month
            if t.route_id in targets and t.product_class_id in targets[t.route_id]:
                targets[t.route_id][t.product_class_id][m] = t.target_amount

        return targets

    def calculate_simulation(
        self,
        *,
        mode: str,
        calc_method: str,
        origin_route_id: str,
        destination_route_id: Optional[str] = None,
        customer_ids: list[str],
        adjustment_direction: str = 'remove',
        transfer_growth_rule: str = 'exact',
        target_year: int,
        effective_month: str,
        eval_customer_start: str,
        eval_customer_end: str,
        eval_route_start: Optional[str] = None,
        eval_route_end: Optional[str] = None,
        product_class_ids: list[str],
    ) -> dict:
        """
        Runs the full target simulation and returns structured results for templates and export.
        """
        errors = []
        if not origin_route_id:
            raise TargetCalculatorError("Debes seleccionar una ruta origen.")
        if not customer_ids:
            raise TargetCalculatorError("Debes seleccionar al menos un cliente para el cálculo.")

        if mode == 'transfer' and not destination_route_id:
            raise TargetCalculatorError("Debes seleccionar la ruta destino para la transferencia.")

        if mode == 'transfer' and origin_route_id == destination_route_id:
            raise TargetCalculatorError("La ruta origen y la ruta destino no pueden ser la misma.")

        eff_start, _ = self.parse_month_bounds(effective_month)
        c_start, c_end = self.parse_month_bounds(eval_customer_start)
        _, c_end_last = self.parse_month_bounds(eval_customer_end)

        if not eff_start or not c_start or not c_end_last:
            raise TargetCalculatorError("Las fechas ingresadas no tienen un formato válido (YYYY-MM).")

        r_start, r_end_last = None, None
        if calc_method == 'contribution':
            r_start, _ = self.parse_month_bounds(eval_route_start)
            _, r_end_last = self.parse_month_bounds(eval_route_end)
            if not r_start or not r_end_last:
                raise TargetCalculatorError("Debes ingresar fechas de evaluación de ruta válidas para el método de contribución.")

        # permission check
        if not self.has_full_access:
            allowed_routes = set(self.get_allowed_routes_qs().values_list('id', flat=True))
            if origin_route_id not in allowed_routes:
                raise PermissionsError(f"No tienes permisos sobre la ruta origen {origin_route_id}.")
            if mode == 'transfer' and destination_route_id and destination_route_id not in allowed_routes:
                raise PermissionsError(f"No tienes permisos sobre la ruta destino {destination_route_id}.")

        product_classes = ProductClass.objects.filter(id__in=product_class_ids).order_by('name')
        if not product_classes.exists():
            raise TargetCalculatorError("Debes seleccionar al menos una clase de producto.")

        origin_route = Route.objects.filter(id=origin_route_id).first()
        if not origin_route:
            raise TargetCalculatorError("La ruta origen especificada no existe.")

        dest_route = None
        if mode == 'transfer':
            dest_route = Route.objects.filter(id=destination_route_id).first()
            if not dest_route:
                raise TargetCalculatorError("La ruta destino especificada no existe.")

        routes_to_fetch = [origin_route_id]
        if mode == 'transfer' and destination_route_id:
            routes_to_fetch.append(destination_route_id)

        targets = self.get_route_targets(routes_to_fetch, target_year, product_classes)
        origin_targets = targets.get(origin_route_id, {})
        dest_targets = targets.get(destination_route_id, {}) if destination_route_id else {}

        strategy_cls = CALCULATION_STRATEGIES.get(calc_method)
        if not strategy_cls:
            raise TargetCalculatorError(f"Método de cálculo desconocido: '{calc_method}'.")

        strategy = strategy_cls()
        eval_params = {
            'c_start': c_start,
            'c_end': c_end_last,
            'r_start': r_start,
            'r_end': r_end_last,
            'transfer_growth_rule': transfer_growth_rule,
            'mode': mode,
        }

        computed_origin, computed_dest = strategy.calculate_deltas(
            product_classes=product_classes,
            customer_ids=customer_ids,
            origin_route_id=origin_route_id,
            destination_route_id=destination_route_id,
            eff_date=eff_start,
            target_year=target_year,
            origin_targets=origin_targets,
            dest_targets=dest_targets,
            eval_params=eval_params,
        )

        months = [datetime.date(target_year, m, 1) for m in range(1, 13)]

        origin_result = self._build_route_result(
            route=origin_route,
            product_classes=product_classes,
            targets=origin_targets,
            computed_deltas=computed_origin,
            months=months,
            mode=mode,
            adjustment_direction=adjustment_direction,
            is_origin=True
        )

        dest_result = None
        if mode == 'transfer' and dest_route:
            dest_result = self._build_route_result(
                route=dest_route,
                product_classes=product_classes,
                targets=dest_targets,
                computed_deltas=computed_dest,
                months=months,
                mode=mode,
                adjustment_direction=adjustment_direction,
                is_origin=False
            )

        customer_summary = self._calculate_customer_summary(
            mode=mode,
            origin_route_id=origin_route_id,
            destination_route_id=destination_route_id,
            customer_ids=customer_ids,
            adjustment_direction=adjustment_direction
        )

        return {
            'mode': mode,
            'calc_method': calc_method,
            'target_year': target_year,
            'effective_month': effective_month,
            'origin': origin_result,
            'destination': dest_result,
            'customer_summary': customer_summary,
            'origin_route_id': origin_route_id,
            'destination_route_id': destination_route_id,
            'customer_ids': customer_ids,
        }

    def _build_route_result(
        self,
        *,
        route: Route,
        product_classes: QuerySet,
        targets: dict,
        computed_deltas: dict,
        months: list[datetime.date],
        mode: str,
        adjustment_direction: str,
        is_origin: bool
    ) -> dict:
        """
        builds matrix breakdown of targets, deltas, and projected values per product class and month
        """
        route_display = f"{route.id.upper()} {route.name.title()}"
        if route.business_unit:
            route_display += f" ({route.business_unit.name.title()})"

        result = {
            'route_id': route.id,
            'route_name': route_display,
            'classes': [],
            'month_totals': [
                {'date': m, 'old_target': Decimal('0.00'), 'growth': Decimal('0.00'), 'delta': Decimal('0.00'), 'new_target': Decimal('0.00')}
                for m in months
            ],
            'grand_total': {'old_target': Decimal('0.00'), 'delta': Decimal('0.00'), 'new_target': Decimal('0.00')}
        }

        if mode == 'transfer':
            sign = Decimal('-1') if is_origin else Decimal('1')
        else:
            sign = Decimal('1') if adjustment_direction == 'add' else Decimal('-1')

        for pc in product_classes:
            pc_data = {
                'class_id': pc.id,
                'class_name': pc.name.title(),
                'months': [],
                'totals': {'old_target': Decimal('0.00'), 'delta': Decimal('0.00'), 'new_target': Decimal('0.00')}
            }

            for idx, m in enumerate(months):
                old_target = targets.get(pc.id, {}).get(m.month, Decimal('0.00'))

                growth = Decimal('0.00')
                if m.month > 1:
                    prev = targets.get(pc.id, {}).get(m.month - 1, Decimal('0.00'))
                    if prev > 0:
                        growth = ((old_target - prev) / prev) * Decimal('100.00')
                    elif prev == 0 and old_target > 0:
                        growth = Decimal('100.00')

                delta_val = computed_deltas.get(pc.id, {}).get(m.month, Decimal('0.00')) * sign
                new_target = old_target + delta_val
                if new_target < 0:
                    new_target = Decimal('0.00')

                pc_data['totals']['old_target'] += old_target
                pc_data['totals']['delta'] += delta_val
                pc_data['totals']['new_target'] += new_target

                result['month_totals'][idx]['old_target'] += old_target
                result['month_totals'][idx]['delta'] += delta_val
                result['month_totals'][idx]['new_target'] += new_target

                result['grand_total']['old_target'] += old_target
                result['grand_total']['delta'] += delta_val
                result['grand_total']['new_target'] += new_target

                pc_data['months'].append({
                    'date': m,
                    'old_target': old_target,
                    'growth': growth,
                    'delta': delta_val,
                    'new_target': new_target
                })
            result['classes'].append(pc_data)

        for idx, mt in enumerate(result['month_totals']):
            if mt['date'].month > 1:
                prev = result['month_totals'][idx - 1]['old_target']
                if prev > 0:
                    mt['growth'] = ((mt['old_target'] - prev) / prev) * Decimal('100.00')
                elif prev == 0 and mt['old_target'] > 0:
                    mt['growth'] = Decimal('100.00')
                else:
                    mt['growth'] = Decimal('0.00')
            else:
                mt['growth'] = Decimal('0.00')

        return result

    def _calculate_customer_summary(
        self,
        *,
        mode: str,
        origin_route_id: str,
        destination_route_id: Optional[str],
        customer_ids: list[str],
        adjustment_direction: str
    ) -> dict:
        """
        computes active customer portfolio changes using active CustomerAssignment records
        """
        today = timezone.localdate()

        def _get_active_count(route_id: str) -> int:
            return Customer.objects.filter(
                assignments__route_id=route_id,
                assignments__start_date__lte=today
            ).filter(
                Q(assignments__end_date__isnull=True) | Q(assignments__end_date__gt=today)
            ).distinct().count()

        def _get_selected_in_route_count(route_id: str) -> int:
            return Customer.objects.filter(
                id__in=customer_ids,
                assignments__route_id=route_id,
                assignments__start_date__lte=today
            ).filter(
                Q(assignments__end_date__isnull=True) | Q(assignments__end_date__gt=today)
            ).distinct().count()

        origin_current = _get_active_count(origin_route_id)
        customers_in_origin = _get_selected_in_route_count(origin_route_id)
        total_selected = len(customer_ids)

        summary = {}

        if mode == 'transfer' and destination_route_id:
            dest_current = _get_active_count(destination_route_id)
            customers_in_dest = _get_selected_in_route_count(destination_route_id)

            origin_removed = customers_in_origin
            dest_added = total_selected - customers_in_dest

            summary['origin'] = {
                'current': origin_current,
                'affected': origin_removed,
                'is_addition': False,
                'final': max(0, origin_current - origin_removed)
            }
            summary['destination'] = {
                'current': dest_current,
                'affected': dest_added,
                'is_addition': True,
                'final': dest_current + dest_added
            }
        else:
            if adjustment_direction == 'add':
                origin_added = total_selected - customers_in_origin
                summary['origin'] = {
                    'current': origin_current,
                    'affected': origin_added,
                    'is_addition': True,
                    'final': origin_current + origin_added
                }
            else:
                origin_removed = customers_in_origin
                summary['origin'] = {
                    'current': origin_current,
                    'affected': origin_removed,
                    'is_addition': False,
                    'final': max(0, origin_current - origin_removed)
                }

        return summary

    def export_simulation_excel(self, results: dict) -> io.BytesIO:
        """
        Generates multi-sheet Excel report matching Datall's export styling standards.
        """
        exporter = SaleTargetCalculatorExports(calculator_service=self)
        return exporter.export_simulation_report(results)


@dataclass
class SaleTargetCalculatorExports:
    calculator_service: SaleTargetCalculatorService

    def export_simulation_report(self, results: dict) -> io.BytesIO:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        if wb.worksheets:
            wb.remove(wb.active)

        # Style palette consistent with Datall reports
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        subheader_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        grand_total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

        white_bold = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
        black_bold = Font(name="Calibri", color="0F172A", bold=True, size=10)
        black_reg = Font(name="Calibri", color="0F172A", size=10)
        title_font = Font(name="Calibri", color="0F172A", bold=True, size=14)
        subtitle_font = Font(name="Calibri", color="64748B", italic=True, size=9)
        section_font = Font(name="Calibri", color="1E293B", bold=True, size=11)

        green_font = Font(name="Calibri", color="15803D", bold=True, size=10)
        red_font = Font(name="Calibri", color="B91C1C", bold=True, size=10)

        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        thin_border_side = Side(style='thin', color='CBD5E1')
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        currency_format = '"$"#,##0.00'
        pct_format = '0.00%'
        int_format = '#,##0'

        def style_cell(cell, fill=None, font=None, alignment=None, number_format=None):
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if number_format:
                cell.number_format = number_format
            cell.border = cell_border
            return cell

        routes_keys = [('origin', 'Origen')]
        if results.get('mode') == 'transfer' and results.get('destination'):
            routes_keys.append(('destination', 'Destino'))

        now_str = timezone.localtime().strftime('%Y-%m-%d %H:%M')

        for key, role_label in routes_keys:
            r = results.get(key)
            if not r:
                continue

            clean_route = r['route_name'].replace('/', '-').translate(str.maketrans('', '', '\\/*?:[]'))
            sheet_title = f"{role_label} - {clean_route}"[:31]
            ws = wb.create_sheet(title=sheet_title)
            ws.views.sheetView[0].showGridLines = True

            # title
            ws.append([f"SIMULACIÓN DE OBJETIVOS DE VENTA: {r['route_name'].upper()} ({role_label.upper()})"])
            ws.cell(row=ws.max_row, column=1).font = title_font
            ws.append([f"Generado el: {now_str} | Método: {results.get('calc_method', '').title()} | Modo: {results.get('mode', '').title()} | Año: {results.get('target_year')}"])
            ws.cell(row=ws.max_row, column=1).font = subtitle_font
            ws.append([])

            # customer portfolio summary
            summary = results.get('customer_summary', {}).get(key, {})
            if summary:
                ws.append(["Resumen de Cartera de Clientes"])
                ws.cell(row=ws.max_row, column=1).font = section_font
                ws.append(["Cartera actual:", f"{summary.get('current', 0)} clientes"])
                affected_label = "Nuevos clientes a integrar:" if summary.get('is_addition') else "Clientes a transferir/remover:"
                affected_sign = "+" if summary.get('is_addition') else "-"
                ws.append([affected_label, f"{affected_sign}{summary.get('affected', 0)} clientes"])
                ws.append(["Cartera proyectada:", f"{summary.get('final', 0)} clientes"])
                ws.append([])

            # breakdown table
            ws.append(["Desglose de Cálculo (Objetivo Original, Crecimiento %, Ajuste Delta)"])
            ws.cell(row=ws.max_row, column=1).font = section_font
            ws.append([])

            if not r['classes']:
                ws.append(["No hay clases de producto seleccionadas"])
                continue

            months = r['classes'][0]['months']

            h_row1 = ws.max_row + 1
            style_cell(ws.cell(row=h_row1, column=1, value="Clase de Producto"), fill=header_fill, font=white_bold, alignment=center_align)
            ws.merge_cells(start_row=h_row1, start_column=1, end_row=h_row1 + 1, end_column=1)

            col_idx = 2
            for m in months:
                month_label = m['date'].strftime('%b %Y').upper()
                cell = ws.cell(row=h_row1, column=col_idx, value=month_label)
                style_cell(cell, fill=header_fill, font=white_bold, alignment=center_align)
                for c in range(col_idx, col_idx + 3):
                    style_cell(ws.cell(row=h_row1, column=c), fill=header_fill)
                ws.merge_cells(start_row=h_row1, start_column=col_idx, end_row=h_row1, end_column=col_idx + 2)
                col_idx += 3

            cell = ws.cell(row=h_row1, column=col_idx, value="TOTAL ANUAL")
            style_cell(cell, fill=header_fill, font=white_bold, alignment=center_align)
            for c in range(col_idx, col_idx + 3):
                style_cell(ws.cell(row=h_row1, column=c), fill=header_fill)
            ws.merge_cells(start_row=h_row1, start_column=col_idx, end_row=h_row1, end_column=col_idx + 2)

            h_row2 = h_row1 + 1
            col_idx = 2
            sub_headers = ["Obj. Orig", "Crec %", "Ajuste"]
            for m in months:
                for sh in sub_headers:
                    style_cell(ws.cell(row=h_row2, column=col_idx, value=sh), fill=subheader_fill, font=white_bold, alignment=center_align)
                    col_idx += 1
            for sh in sub_headers:
                style_cell(ws.cell(row=h_row2, column=col_idx, value=sh), fill=subheader_fill, font=white_bold, alignment=center_align)
                col_idx += 1

            for cls in r['classes']:
                cur_row = ws.max_row + 1
                style_cell(ws.cell(row=cur_row, column=1, value=cls['class_name']), font=black_bold, alignment=left_align)

                col_idx = 2
                for m in cls['months']:
                    style_cell(ws.cell(row=cur_row, column=col_idx, value=m['old_target']), font=black_reg, alignment=right_align, number_format=currency_format)
                    col_idx += 1

                    g_val = float(m['growth']) / 100.0 if m['growth'] else 0.0
                    g_font = green_font if g_val > 0 else (red_font if g_val < 0 else black_reg)
                    style_cell(ws.cell(row=cur_row, column=col_idx, value=g_val), font=g_font, alignment=right_align, number_format=pct_format)
                    col_idx += 1

                    d_val = m['delta']
                    d_font = green_font if d_val > 0 else (red_font if d_val < 0 else black_reg)
                    style_cell(ws.cell(row=cur_row, column=col_idx, value=d_val), font=d_font, alignment=right_align, number_format=currency_format)
                    col_idx += 1

                t = cls['totals']
                style_cell(ws.cell(row=cur_row, column=col_idx, value=t['old_target']), fill=total_fill, font=black_bold, alignment=right_align, number_format=currency_format)
                col_idx += 1
                style_cell(ws.cell(row=cur_row, column=col_idx, value="-"), fill=total_fill, font=black_reg, alignment=center_align)
                col_idx += 1
                d_font = green_font if t['delta'] > 0 else (red_font if t['delta'] < 0 else black_bold)
                style_cell(ws.cell(row=cur_row, column=col_idx, value=t['delta']), fill=total_fill, font=d_font, alignment=right_align, number_format=currency_format)
                col_idx += 1

            tot_row = ws.max_row + 1
            style_cell(ws.cell(row=tot_row, column=1, value="TOTAL RUTA"), fill=total_fill, font=black_bold, alignment=left_align)

            col_idx = 2
            for mt in r['month_totals']:
                style_cell(ws.cell(row=tot_row, column=col_idx, value=mt['old_target']), fill=total_fill, font=black_bold, alignment=right_align, number_format=currency_format)
                col_idx += 1

                g_val = float(mt['growth']) / 100.0 if mt['growth'] else 0.0
                g_font = green_font if g_val > 0 else (red_font if g_val < 0 else black_bold)
                style_cell(ws.cell(row=tot_row, column=col_idx, value=g_val), fill=total_fill, font=g_font, alignment=right_align, number_format=pct_format)
                col_idx += 1

                d_val = mt['delta']
                d_font = green_font if d_val > 0 else (red_font if d_val < 0 else black_bold)
                style_cell(ws.cell(row=tot_row, column=col_idx, value=d_val), fill=total_fill, font=d_font, alignment=right_align, number_format=currency_format)
                col_idx += 1

            gt = r['grand_total']
            style_cell(ws.cell(row=tot_row, column=col_idx, value=gt['old_target']), fill=grand_total_fill, font=black_bold, alignment=right_align, number_format=currency_format)
            col_idx += 1
            style_cell(ws.cell(row=tot_row, column=col_idx, value="-"), fill=grand_total_fill, font=black_reg, alignment=center_align)
            col_idx += 1
            d_font = green_font if gt['delta'] > 0 else (red_font if gt['delta'] < 0 else black_bold)
            style_cell(ws.cell(row=tot_row, column=col_idx, value=gt['delta']), fill=grand_total_fill, font=d_font, alignment=right_align, number_format=currency_format)
            col_idx += 1

            ws.append([])
            ws.append([])
            ws.append(["Objetivos Planificados (Objetivos Finales con Ajuste Aplicado)"])
            ws.cell(row=ws.max_row, column=1).font = section_font
            ws.append([])

            h_row3 = ws.max_row + 1
            style_cell(ws.cell(row=h_row3, column=1, value="Clase de Producto"), fill=header_fill, font=white_bold, alignment=center_align)

            col_idx = 2
            for m in months:
                month_label = m['date'].strftime('%b %Y').upper()
                style_cell(ws.cell(row=h_row3, column=col_idx, value=month_label), fill=header_fill, font=white_bold, alignment=center_align)
                col_idx += 1

            style_cell(ws.cell(row=h_row3, column=col_idx, value="TOTAL ANUAL"), fill=header_fill, font=white_bold, alignment=center_align)

            for cls in r['classes']:
                cur_row = ws.max_row + 1
                style_cell(ws.cell(row=cur_row, column=1, value=cls['class_name']), font=black_bold, alignment=left_align)

                col_idx = 2
                for m in cls['months']:
                    c_font = green_font if m['new_target'] > m['old_target'] else (red_font if m['new_target'] < m['old_target'] else black_reg)
                    style_cell(ws.cell(row=cur_row, column=col_idx, value=m['new_target']), font=c_font, alignment=right_align, number_format=currency_format)
                    col_idx += 1

                t = cls['totals']
                c_font = green_font if t['new_target'] > t['old_target'] else (red_font if t['new_target'] < t['old_target'] else black_bold)
                style_cell(ws.cell(row=cur_row, column=col_idx, value=t['new_target']), fill=total_fill, font=c_font, alignment=right_align, number_format=currency_format)

            tot_row2 = ws.max_row + 1
            style_cell(ws.cell(row=tot_row2, column=1, value="TOTAL RUTA"), fill=total_fill, font=black_bold, alignment=left_align)

            col_idx = 2
            for mt in r['month_totals']:
                c_font = green_font if mt['new_target'] > mt['old_target'] else (red_font if mt['new_target'] < mt['old_target'] else black_bold)
                style_cell(ws.cell(row=tot_row2, column=col_idx, value=mt['new_target']), fill=total_fill, font=c_font, alignment=right_align, number_format=currency_format)
                col_idx += 1

            gt = r['grand_total']
            c_font = green_font if gt['new_target'] > gt['old_target'] else (red_font if gt['new_target'] < gt['old_target'] else black_bold)
            style_cell(ws.cell(row=tot_row2, column=col_idx, value=gt['new_target']), fill=grand_total_fill, font=c_font, alignment=right_align, number_format=currency_format)

        ws_cust = wb.create_sheet(title="Clientes Seleccionados")
        ws_cust.views.sheetView[0].showGridLines = True
        ws_cust.append(["Clientes considerados en la simulación"])
        ws_cust.cell(row=ws_cust.max_row, column=1).font = title_font
        ws_cust.append([])

        h_row_c = ws_cust.max_row + 1
        headers_cust = ["ID Cliente", "Nombre / Razón Social", "Tipo de Cliente", "Límite de Crédito", "Días Crédito", "Líder Opinión"]
        for c_idx, h_text in enumerate(headers_cust, 1):
            style_cell(ws_cust.cell(row=h_row_c, column=c_idx, value=h_text), fill=header_fill, font=white_bold, alignment=center_align)

        considered_customers = Customer.objects.filter(id__in=results.get('customer_ids', [])).select_related('customer_type').order_by('name')
        for cust in considered_customers:
            cur_row = ws_cust.max_row + 1
            style_cell(ws_cust.cell(row=cur_row, column=1, value=cust.id.upper()), font=black_reg, alignment=center_align)
            style_cell(ws_cust.cell(row=cur_row, column=2, value=cust.name.title()), font=black_bold, alignment=left_align)
            style_cell(ws_cust.cell(row=cur_row, column=3, value=cust.customer_type.name.title() if cust.customer_type else "-"), font=black_reg, alignment=left_align)
            style_cell(ws_cust.cell(row=cur_row, column=4, value=cust.credit_limit), font=black_reg, alignment=right_align, number_format=currency_format)
            style_cell(ws_cust.cell(row=cur_row, column=5, value=cust.credit_days), font=black_reg, alignment=center_align, number_format=int_format)
            style_cell(ws_cust.cell(row=cur_row, column=6, value="Sí" if cust.opinion_leader else "No"), font=black_reg, alignment=center_align)

        # autofit columns across all worksheets
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if '\n' in val_str:
                        val_str = max(val_str.split('\n'), key=len)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
