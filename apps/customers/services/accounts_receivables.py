import io
from dataclasses import dataclass
from decimal import Decimal
from typing import ClassVar

from django.db.models import (
    Q,
    QuerySet,
    Count,
    Sum,
    F,
    Case,
    When,
    Value,
    DecimalField,
)
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.core.services.users import UsersService
from apps.customers.models import Customer, CustomerAssignment, AccountsReceivable
from apps.sales.models import Route
from apps.sales.services.routes import RoutesService


class ServiceError(Exception):
    pass


class PermissionsError(ServiceError):
    pass


class AccountsReceivableNotFound(ServiceError):
    pass


@dataclass
class AccountsReceivablesService(UsersService):
    accounts_receivable_model: type = AccountsReceivable
    route_model: type = Route
    customer_model: type = Customer
    customer_assignment_model: type = CustomerAssignment

    ACCESS_CONTEXTS: ClassVar[tuple[str, ...]] = (
        'acceso_total_clientes',
        'clientes',
        'acceso_total_ventas',
        'ventas',
        'acceso_total_rutas',
    )

    def _get_allowed_routes_qs(self) -> QuerySet:
        """
        Helper to get allowed routes for the current user using RoutesService.
        """
        routes_service = RoutesService(user=self.user)
        return routes_service.get_allowed_routes(can_view=True, can_edit=False)

    def read_ars_by_allowed_routes(self) -> QuerySet:
        """
        Returns accounts receivable for invoices emitted by routes to which the user has access.
        Independent of whether the customer currently belongs to those routes or not.
        """
        base_qs = self.accounts_receivable_model.objects.select_related(
            'customer',
            'customer__customer_type',
            'route',
            'route__business_unit',
            'route__sale_channel',
        )

        if self.has_full_access:
            return base_qs.order_by('-due_date', '-issue_date', '-id')

        allowed_routes = self._get_allowed_routes_qs()
        return base_qs.filter(route__in=allowed_routes).order_by('-due_date', '-issue_date', '-id')

    def read_ars_by_allowed_customers(self) -> QuerySet:
        """
        Returns accounts receivable corresponding to customers currently assigned to routes
        to which the user has access, regardless of which route historically emitted the invoice.
        """
        base_qs = self.accounts_receivable_model.objects.select_related(
            'customer',
            'customer__customer_type',
            'route',
            'route__business_unit',
            'route__sale_channel',
        )

        if self.has_full_access:
            return base_qs.order_by('-due_date', '-issue_date', '-id')

        today = timezone.localdate()
        allowed_routes = self._get_allowed_routes_qs()

        allowed_customers_subquery = self.customer_assignment_model.objects.filter(
            route__in=allowed_routes
        ).filter(
            Q(end_date__isnull=True) | Q(end_date__gte=today)
        ).values('customer_id')

        return base_qs.filter(customer_id__in=allowed_customers_subquery).order_by('-due_date', '-issue_date', '-id')

    def read_ars(self) -> QuerySet:
        """
        Default accounts receivable listing perspective: based on current customer assignments.
        """
        return self.read_ars_by_allowed_customers()

    def read_ar(self, *, pk: str | int) -> AccountsReceivable:
        """
        Returns a single accounts receivable instance by ID, validated against permissions.
        """
        ar_obj = self.read_ars().filter(pk=pk).first()
        if ar_obj:
            return ar_obj

        if self.accounts_receivable_model.objects.filter(pk=pk).exists():
            raise PermissionsError(f'No tienes permiso para acceder a la cuenta por cobrar con ID "{pk}".')

        raise AccountsReceivableNotFound(f'No se encontró ninguna cuenta por cobrar con el ID "{pk}".')

    def _clean_ars(self, file_obj) -> tuple[bool, str | object]:
        try:
            import pandas as pd
        except ImportError:
            return False, "La librería 'pandas' no está instalada en el entorno."

        from apps.core.services.uploads import BaseETLHelper
        is_valid, df_or_err = BaseETLHelper.read_file_to_dataframe(file_obj)
        if not is_valid:
            return False, df_or_err

        df = df_or_err
        df = BaseETLHelper.apply_reference_column_mappings(
            df,
            self.accounts_receivable_model,
            submodule_name='importacion',
            context='columna'
        )
        df = BaseETLHelper.resolve_foreign_key_columns(df, self.accounts_receivable_model)

        if 'customer' in df.columns and 'customer_id' not in df.columns:
            df.rename(columns={'customer': 'customer_id'}, inplace=True)

        is_req_valid, req_msg = BaseETLHelper.validate_required_columns(df, {'customer_id': 'Cliente'})
        if not is_req_valid:
            return False, req_msg

        str_cols = ['customer_id', 'route_id', 'doc_id', 'description']
        for col in str_cols:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace({'nan': None, '': None, 'None': None, 'none': None, 'null': None, 'NULL': None})

        df = df.dropna(subset=['customer_id'])
        if df.empty:
            return False, "El archivo no contiene registros válidos después de descartar filas sin identificador de cliente."

        from django.contrib.contenttypes.models import ContentType
        from apps.core.models import Reference
        ar_ctype = ContentType.objects.get_for_model(self.accounts_receivable_model)

        if 'issue_date' in df.columns:
            issue_refs = Reference.objects.filter(content_type=ar_ctype, context__icontains='emision')
            for ref in issue_refs:
                k = str(ref.key)
                v = str(ref.value).strip().lower() if getattr(ref, 'value', '') else str(getattr(ref, 'reference', '')).strip().lower()
                if v == 'null':
                    df['issue_date'] = df['issue_date'].replace(k, None)
                elif k and v:
                    df['issue_date'] = df['issue_date'].replace(k, v)

        if 'due_date' in df.columns:
            due_refs = Reference.objects.filter(content_type=ar_ctype, context__icontains='pago')
            for ref in due_refs:
                k = str(ref.key)
                v = str(ref.value).strip().lower() if getattr(ref, 'value', '') else str(getattr(ref, 'reference', '')).strip().lower()
                if v == 'null':
                    df['due_date'] = df['due_date'].replace(k, None)
                elif k and v:
                    df['due_date'] = df['due_date'].replace(k, v)

        date_cols = ['issue_date', 'due_date']
        for col in date_cols:
            if col in df.columns:
                df[col] = df[col].replace({'null': None, 'NULL': None, 'nan': None, 'none': None, '': None})
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.date

        num_cols = ['total_balance', 'balance_15', 'balance_30', 'balance_60', 'past_due', 'current_balance']
        for c in num_cols:
            if c in df.columns:
                df[c] = pd.to_numeric(
                    df[c].astype(str).str.replace(r'[$, ]', '', regex=True),
                    errors='coerce'
                ).fillna(0).round(4)

        df = df.where(pd.notnull(df), None)
        return True, df

    def bulk_create_ars(self, file_obj) -> object:
        from apps.core.services.uploads import ImportResult, PermissionsError, BaseETLHelper
        from django.db import transaction

        if not self.has_full_access:
            raise PermissionsError('No tienes permisos suficientes para reemplazar la cartera de cuentas por cobrar.')

        is_valid, df_or_err = self._clean_ars(file_obj)
        if not is_valid:
            return ImportResult(success=False, message=df_or_err)

        df = df_or_err

        is_fk_valid, fk_msg = BaseETLHelper.validate_foreign_keys(df, self.accounts_receivable_model)
        if not is_fk_valid:
            return ImportResult(success=False, message=fk_msg)

        model_fields = [f.name for f in self.accounts_receivable_model._meta.get_fields() if not f.is_relation]
        model_fields.extend([f.attname for f in self.accounts_receivable_model._meta.get_fields() if f.is_relation and hasattr(f, 'attname')])
        
        valid_columns = [col for col in df.columns if col in model_fields and col != 'id']

        records_to_create = []
        total_processed = 0

        for _, row in df.iterrows():
            data = {}
            for col in valid_columns:
                val = row[col]
                if val is not None and str(val).lower() not in ['nan', '', 'none']:
                    data[col] = val
            
            if 'customer_id' not in data or not data['customer_id']:
                continue

            records_to_create.append(self.accounts_receivable_model(**data))
            total_processed += 1

        if not records_to_create:
            return ImportResult(success=False, message="El archivo no contiene cuentas por cobrar válidas para importar.")

        try:
            with transaction.atomic():
                self.accounts_receivable_model.objects.all().delete()
                
                self.accounts_receivable_model.objects.bulk_create(
                    records_to_create,
                    batch_size=5000
                )

            return ImportResult(
                success=True,
                message=f"Importación exitosa. Se reemplazó la cartera completa con {len(records_to_create)} nuevos registros.",
                total_processed=total_processed,
                created_count=len(records_to_create),
                updated_count=0
            )

        except Exception as e:
            humanized_msg = BaseETLHelper.humanize_database_error(e)
            return ImportResult(
                success=False,
                message=humanized_msg,
                total_processed=total_processed,
                errors=[str(e)]
            )


@dataclass
class AccountsReceivablesStats:
    '''Dedicated only to give general stats about accounts receivable'''
    accounts_receivables_service: AccountsReceivablesService

    @property
    def _base_qs(self) -> QuerySet:
        return self.accounts_receivables_service.read_ars()

    def stats(self, *, qs: QuerySet = None) -> dict:
        base_qs = qs if qs is not None else self._base_qs

        agg = base_qs.aggregate(
            ars_count=Count('pk', distinct=True),
            total_balance=Sum('total_balance'),
            current_balance=Sum('current_balance'),
            balance_15=Sum('balance_15'),
            balance_30=Sum('balance_30'),
            balance_60=Sum('balance_60'),
            past_due=Sum('past_due'),
            unique_customers_count=Count('customer', distinct=True),
            unique_routes_count=Count('route', distinct=True),
        )

        total_balance = agg['total_balance'] or Decimal('0.0000')
        current_balance = agg['current_balance'] or Decimal('0.0000')
        balance_15 = agg['balance_15'] or Decimal('0.0000')
        balance_30 = agg['balance_30'] or Decimal('0.0000')
        balance_60 = agg['balance_60'] or Decimal('0.0000')
        past_due = agg['past_due'] or Decimal('0.0000')

        agg['total_balance'] = total_balance
        agg['current_balance'] = current_balance
        agg['overdue_balance'] = balance_15 + balance_30 + balance_60 + past_due
        agg['balance_15'] = balance_15
        agg['balance_30'] = balance_30
        agg['balance_60'] = balance_60
        agg['past_due'] = past_due
        agg['accs_receivable_count'] = agg['unique_customers_count']

        #credit limit of customers with accounts receivable in current queryset
        ar_customer_ids = base_qs.values_list('customer_id', flat=True).distinct()
        credit_ar = Customer.objects.filter(id__in=ar_customer_ids).aggregate(total=Sum('credit_limit'))['total'] or Decimal('0.0000')

        #credit limit of entire allowed customer portfolio
        from apps.customers.services.customers import CustomersService
        customers_service = CustomersService(user=self.accounts_receivables_service.user)
        credit_ptf = customers_service.read_customers().aggregate(total=Sum('credit_limit'))['total'] or Decimal('0.0000')

        usage_ar = (total_balance / credit_ar * 100) if credit_ar > 0 else Decimal('0.00')
        usage_ptf = (total_balance / credit_ptf * 100) if credit_ptf > 0 else Decimal('0.00')

        agg['credit_ar'] = credit_ar
        agg['credit_ptf'] = credit_ptf
        agg['credit_usage_by_ar'] = usage_ar
        agg['credit_usage_by_ptf'] = usage_ptf

        return agg

    def customer_breakdown(self, *, qs: QuerySet = None) -> QuerySet:
        base_qs = qs if qs is not None else self._base_qs

        return base_qs.values(
            'customer__id',
            'customer__name',
            'customer__credit_limit',
            'route__id',
            'route__name',
            'route__business_unit__id',
            'route__business_unit__name',
        ).annotate(
            total_balance=Sum('total_balance'),
            current_balance=Sum('current_balance'),
            balance_15=Sum('balance_15'),
            balance_30=Sum('balance_30'),
            balance_60=Sum('balance_60'),
            past_due=Sum('past_due'),
            overdue_balance=F('total_balance') - F('current_balance'),
            credit_usage=Case(
                When(customer__credit_limit__gt=0, then=(F('total_balance') * 100.0) / F('customer__credit_limit')),
                default=Value(0.0),
                output_field=DecimalField()
            )
        ).order_by('-total_balance')

@dataclass
class AccountsReceivablesExports:
    '''dedicated to receive all exports request for accounts receivable objects and filters'''
    accounts_receivables_service: AccountsReceivablesService

    @property
    def _base_qs(self) -> QuerySet:
        return self.accounts_receivables_service.read_ars()

    def export_collections_report(self, *, qs: QuerySet = None, perspective: str = 'current_customers') -> io.BytesIO:
        stats_service = AccountsReceivablesStats(accounts_receivables_service=self.accounts_receivables_service)
        base_qs = qs if qs is not None else self._base_qs

        kpis = stats_service.stats(qs=base_qs)
        customer_breakdown = stats_service.customer_breakdown(qs=base_qs)

        wb = openpyxl.Workbook()

        #styles
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        section_font = Font(name="Calibri", size=12, bold=True, color="0F172A")
        title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
        subtitle_font = Font(name="Calibri", size=9, italic=True, color="64748B")
        data_font = Font(name="Calibri", size=10)
        bold_data_font = Font(name="Calibri", size=10, bold=True)

        thin_border_side = Side(style='thin', color='CBD5E1')
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        currency_format = '"$"#,##0.00'
        pct_format = '0.00%'
        int_format = '#,##0'

        #sheet 1 generals
        ws_summary = wb.active
        ws_summary.title = "Resumen General"
        ws_summary.views.sheetView[0].showGridLines = True

        #title
        ws_summary.cell(row=1, column=1, value="REPORTE EJECUTIVO DE COBRANZA").font = title_font
        now_str = timezone.localtime().strftime('%Y-%m-%d %H:%M')
        persp_str = "Clientes asignados" if perspective != 'emitting_routes' else "Ruta emisora"
        ws_summary.cell(row=2, column=1, value=f"Generado el: {now_str} | Modo: {persp_str}").font = subtitle_font

        #sect 1, gnrl metrics
        ws_summary.cell(row=4, column=1, value="1. Indicadores Generales").font = section_font
        general_headers = ["Indicador", "Monto / Cantidad"]
        for col_num, h_text in enumerate(general_headers, 1):
            cell = ws_summary.cell(row=5, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_num > 1 else "left")
            cell.border = cell_border

        general_rows = [
            ("Saldo total", float(kpis.get('total_balance') or 0), currency_format),
            ("Saldo al corriente", float(kpis.get('current_balance') or 0), currency_format),
            ("Saldo vencido", float(kpis.get('overdue_balance') or 0), currency_format),
            ("Clientes con saldo (cuentas por cobrar)", int(kpis.get('unique_customers_count') or 0), int_format),
            ("Total de facturas / documentos", int(kpis.get('ars_count') or 0), int_format),
        ]

        for row_idx, (label, val, num_fmt) in enumerate(general_rows, 6):
            c_lbl = ws_summary.cell(row=row_idx, column=1, value=label)
            c_lbl.font = data_font
            c_lbl.border = cell_border

            c_val = ws_summary.cell(row=row_idx, column=2, value=val)
            c_val.font = bold_data_font
            c_val.number_format = num_fmt
            c_val.alignment = Alignment(horizontal="right")
            c_val.border = cell_border

        #sect 2, aging buckets
        start_row_aging = 13
        ws_summary.cell(row=start_row_aging, column=1, value="2. Antigüedad de Saldos Vencidos").font = section_font
        aging_headers = ["Tramo de Vencimiento", "Monto", "% del Saldo Vencido"]
        for col_num, h_text in enumerate(aging_headers, 1):
            cell = ws_summary.cell(row=start_row_aging + 1, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_num > 1 else "left")
            cell.border = cell_border

        overdue_tot = float(kpis.get('overdue_balance') or 0)
        aging_rows = [
            ("1 a 15 días", float(kpis.get('balance_15') or 0)),
            ("16 a 30 días", float(kpis.get('balance_30') or 0)),
            ("31 a 60 días", float(kpis.get('balance_60') or 0)),
            ("Mayor a 60 días (+60)", float(kpis.get('past_due') or 0)),
        ]

        for idx, (lbl, amt) in enumerate(aging_rows, start_row_aging + 2):
            c_l = ws_summary.cell(row=idx, column=1, value=lbl)
            c_l.font = data_font
            c_l.border = cell_border

            c_a = ws_summary.cell(row=idx, column=2, value=amt)
            c_a.font = data_font
            c_a.number_format = currency_format
            c_a.alignment = Alignment(horizontal="right")
            c_a.border = cell_border

            pct_val = (amt / overdue_tot) if overdue_tot > 0 else 0.0
            c_p = ws_summary.cell(row=idx, column=3, value=pct_val)
            c_p.font = bold_data_font
            c_p.number_format = pct_format
            c_p.alignment = Alignment(horizontal="right")
            c_p.border = cell_border

        #sect 3, credit usage
        start_row_credit = 20
        ws_summary.cell(row=start_row_credit, column=1, value="3. Utilización de Líneas de Crédito").font = section_font
        credit_headers = ["Segmento", "Saldo Deudor Utilizado", "Límite de Crédito Autorizado", "% Utilización"]
        for col_num, h_text in enumerate(credit_headers, 1):
            cell = ws_summary.cell(row=start_row_credit + 1, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_num > 1 else "left")
            cell.border = cell_border

        tot_bal = float(kpis.get('total_balance') or 0)
        credit_ar_val = float(kpis.get('credit_ar') or 0)
        usage_ar_val = float(kpis.get('credit_usage_by_ar') or 0) / 100.0

        credit_ptf_val = float(kpis.get('credit_ptf') or 0)
        usage_ptf_val = float(kpis.get('credit_usage_by_ptf') or 0) / 100.0

        credit_rows = [
            ("Cuentas por Cobrar (Clientes con Saldo Activo)", tot_bal, credit_ar_val, usage_ar_val),
            ("Toda la Cartera (Total Clientes Asignados)", tot_bal, credit_ptf_val, usage_ptf_val),
        ]

        for idx, (seg, debt, cred, usg) in enumerate(credit_rows, start_row_credit + 2):
            c_s = ws_summary.cell(row=idx, column=1, value=seg)
            c_s.font = data_font
            c_s.border = cell_border

            c_d = ws_summary.cell(row=idx, column=2, value=debt)
            c_d.font = data_font
            c_d.number_format = currency_format
            c_d.alignment = Alignment(horizontal="right")
            c_d.border = cell_border

            c_c = ws_summary.cell(row=idx, column=3, value=cred)
            c_c.font = data_font
            c_c.number_format = currency_format
            c_c.alignment = Alignment(horizontal="right")
            c_c.border = cell_border

            c_u = ws_summary.cell(row=idx, column=4, value=usg)
            c_u.font = bold_data_font
            c_u.number_format = pct_format
            c_u.alignment = Alignment(horizontal="right")
            c_u.border = cell_border

        #sheet 2 breakdown customers
        ws_customers = wb.create_sheet(title="Desglose por Cliente")
        ws_customers.views.sheetView[0].showGridLines = True

        cust_headers = [
            "ID Cliente",
            "Nombre Cliente",
            "ID Ruta",
            "Nombre Ruta",
            "Gerencia / Unidad",
            "Límite de Crédito",
            "% Uso de Crédito",
            "Saldo Total",
            "Al Corriente",
            "Saldo Vencido",
            "1 a 15 Días",
            "16 a 30 Días",
            "31 a 60 Días",
            "+60 Días",
        ]

        for col_num, h_text in enumerate(cust_headers, 1):
            cell = ws_customers.cell(row=1, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border

        ws_customers.row_dimensions[1].height = 24

        for row_idx, item in enumerate(customer_breakdown, 2):
            cid = item.get('customer__id') or ''
            cname = (item.get('customer__name') or '').title()
            rid = item.get('route__id') or ''
            rname = (item.get('route__name') or '').title()
            bu_name = (item.get('route__business_unit__name') or '').title()
            c_limit = float(item.get('customer__credit_limit') or 0)
            c_usage = float(item.get('credit_usage') or 0) / 100.0
            tot = float(item.get('total_balance') or 0)
            curr = float(item.get('current_balance') or 0)
            ovd = float(item.get('overdue_balance') or 0)
            b15 = float(item.get('balance_15') or 0)
            b30 = float(item.get('balance_30') or 0)
            b60 = float(item.get('balance_60') or 0)
            past = float(item.get('past_due') or 0)

            values = [
                (cid, '@', "center"),
                (cname, None, "left"),
                (rid, '@', "center"),
                (rname, None, "left"),
                (bu_name, None, "left"),
                (c_limit, currency_format, "right"),
                (c_usage, pct_format, "right"),
                (tot, currency_format, "right"),
                (curr, currency_format, "right"),
                (ovd, currency_format, "right"),
                (b15, currency_format, "right"),
                (b30, currency_format, "right"),
                (b60, currency_format, "right"),
                (past, currency_format, "right"),
            ]

            for col_idx, (val, num_fmt, align_h) in enumerate(values, 1):
                c = ws_customers.cell(row=row_idx, column=col_idx, value=val)
                c.font = data_font
                c.border = cell_border
                c.alignment = Alignment(horizontal=align_h)
                if num_fmt:
                    c.number_format = num_fmt

        #sheet 3 base qs detail
        ws_detail = wb.create_sheet(title="Detalle Facturas y Doctos")
        ws_detail.views.sheetView[0].showGridLines = True

        detail_headers = [
            "Folio / Documento",
            "Concepto",
            "ID Cliente",
            "Nombre Cliente",
            "ID Ruta",
            "Nombre Ruta",
            "Gerencia",
            "Fecha de Emisión",
            "Fecha de Vencimiento",
            "Saldo Total",
            "Al Corriente",
            "1 a 15 Días",
            "16 a 30 Días",
            "31 a 60 Días",
            "+60 Días",
        ]

        for col_num, h_text in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border

        ws_detail.row_dimensions[1].height = 24

        detail_qs = base_qs.select_related('customer', 'route', 'route__business_unit').order_by('customer_id', '-issue_date', 'doc_id')

        for row_idx, ar in enumerate(detail_qs, 2):
            doc = ar.doc_id or str(ar.id)
            desc = (ar.description or '').title()
            cid = ar.customer_id or ''
            cname = (ar.customer.name or '').title() if ar.customer else ''
            rid = ar.route_id or ''
            rname = (ar.route.name or '').title() if ar.route else ''
            bu_name = (ar.route.business_unit.name or '').title() if ar.route and ar.route.business_unit else ''
            f_emision = ar.issue_date.strftime('%Y-%m-%d') if ar.issue_date else ''
            f_venc = ar.due_date.strftime('%Y-%m-%d') if ar.due_date else ''
            tot = float(ar.total_balance or 0)
            curr = float(ar.current_balance or 0)
            b15 = float(ar.balance_15 or 0)
            b30 = float(ar.balance_30 or 0)
            b60 = float(ar.balance_60 or 0)
            past = float(ar.past_due or 0)

            values = [
                (doc, '@', "center"),
                (desc, None, "left"),
                (cid, '@', "center"),
                (cname, None, "left"),
                (rid, '@', "center"),
                (rname, None, "left"),
                (bu_name, None, "left"),
                (f_emision, 'yyyy-mm-dd', "center"),
                (f_venc, 'yyyy-mm-dd', "center"),
                (tot, currency_format, "right"),
                (curr, currency_format, "right"),
                (b15, currency_format, "right"),
                (b30, currency_format, "right"),
                (b60, currency_format, "right"),
                (past, currency_format, "right"),
            ]

            for col_idx, (val, num_fmt, align_h) in enumerate(values, 1):
                c = ws_detail.cell(row=row_idx, column=col_idx, value=val)
                c.font = data_font
                c.border = cell_border
                c.alignment = Alignment(horizontal=align_h)
                if num_fmt:
                    c.number_format = num_fmt

        #auto fix width columns
        for sheet in [ws_summary, ws_customers, ws_detail]:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if cell.number_format == currency_format:
                        val_str = f"${val_str}"
                    max_len = max(max_len, len(val_str))
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
